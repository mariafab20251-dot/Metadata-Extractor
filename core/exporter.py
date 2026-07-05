import csv
import json
from pathlib import Path
from config import CSV_PATH, JSON_PATH, DATA_DIR

class DataExporter:
    def __init__(self):
        self.ensure_csv_header()
        self.reports_dir = DATA_DIR / "reports"
        self.reports_dir.mkdir(parents=True, exist_ok=True)

    CSV_COLUMNS = [
        'video_id', 'platform', 'channel_name', 'video_title',
        'video_description', 'url', 'overlay_text', 'speech_text',
        'captions', 'hashtags', 'timestamp',
        'transcript_word_count', 'transcript_wpm', 'video_duration',
        'view_count',
        'generated_script', 'generated_word_count', 'suggested_title',
        'hashtag_1', 'hashtag_2'
    ]

    def ensure_csv_header(self):
        if not CSV_PATH.exists():
            with open(CSV_PATH, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(self.CSV_COLUMNS)

    def export_to_csv(self, data, channel_folder=None):
        import time
        csv_path = CSV_PATH
        if channel_folder:
            csv_path = channel_folder / "results.csv"
            if not csv_path.exists():
                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(self.CSV_COLUMNS)

        max_retries = 5
        for attempt in range(max_retries):
            try:
                with open(csv_path, 'a', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow([
                        data['video_id'],
                        data['platform'],
                        data.get('channel_name', ''),
                        data.get('video_title', ''),
                        data.get('video_description', ''),
                        data['url'],
                        data['overlay_text'],
                        data['speech_text'],
                        data['captions'],
                        data['hashtags'],
                        data['timestamp'],
                        data.get('transcript_word_count', 0),
                        data.get('transcript_wpm', 0.0),
                        data.get('video_duration', 0.0),
                        data.get('view_count', 0),
                        data.get('generated_script', ''),
                        data.get('generated_word_count', 0),
                        data.get('suggested_title', ''),
                        data.get('hashtag_1', ''),
                        data.get('hashtag_2', '')
                    ])
                break
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                else:
                    raise

    def export_to_json(self, data, channel_folder=None):
        json_path = JSON_PATH
        if channel_folder:
            json_path = channel_folder / "results.json"

        existing_data = []
        if json_path.exists():
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    content = f.read().strip()
                    if content:  # Only parse if file has content
                        existing_data = json.loads(content)
            except (json.JSONDecodeError, ValueError):
                # File is corrupted or empty, start fresh
                existing_data = []

        existing_data.append(data)

        # Upsert by video_id: keep only the latest entry per video (reprocessing
        # should replace stale data, not accumulate duplicates).
        deduped = {}
        for item in existing_data:
            deduped[item.get('video_id')] = item
        existing_data = list(deduped.values())

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, indent=2, ensure_ascii=False)

    def clean_text(self, text):
        """Remove duplicate lines and clean text"""
        if not text:
            return "(No text detected)"

        lines = text.split('\n')
        seen = set()
        unique_lines = []
        for line in lines:
            line = line.strip()
            if line and line not in seen:
                seen.add(line)
                unique_lines.append(line)

        return ' '.join(unique_lines) if unique_lines else "(No text detected)"

    def save_results(self, video_id, url, platform, overlay_text, speech_text, captions, hashtags,
                     channel_folder=None, channel_name="", video_title="", video_description="",
                     transcript_word_count=0, transcript_wpm=0.0, video_duration=0.0,
                     view_count=0):
        from datetime import datetime
        data = {
            'video_id': video_id,
            'platform': platform,
            'channel_name': channel_name,
            'video_title': video_title,
            'video_description': video_description,
            'url': url,
            'overlay_text': overlay_text,
            'speech_text': speech_text,
            'captions': captions,
            'hashtags': hashtags,
            'timestamp': datetime.now().isoformat(),
            'transcript_word_count': transcript_word_count,
            'transcript_wpm': transcript_wpm,
            'video_duration': video_duration,
            'view_count': view_count,
            'generated_script': '',
            'generated_word_count': 0,
            'suggested_title': '',
            'hashtag_1': '',
            'hashtag_2': ''
        }
        self.export_to_csv(data, channel_folder)
        self.export_to_json(data, channel_folder)
        self.export_to_txt(data, channel_folder)

    def export_to_txt(self, data, channel_folder=None):
        """Export individual TXT report for each video"""
        video_id = data['video_id']
        platform = data['platform'].upper()

        report = f"""
{'='*80}
VIDEO EXTRACTION REPORT
{'='*80}

Video ID:       {video_id}
Platform:       {platform}
Channel:        {data.get('channel_name', '') or '(Unknown)'}
Title:          {data.get('video_title', '') or '(Unknown)'}
URL:            {data['url']}
Date Processed: {data['timestamp'].split('T')[0]}

{'='*80}
DESCRIPTION
{'='*80}

{data.get('video_description', '') or '(No description)'}

{'='*80}
OVERLAY TEXT (OCR)
{'='*80}

{self.clean_text(data['overlay_text'])}

{'='*80}
SPEECH TRANSCRIPT (WHISPER)
{'='*80}

{data['speech_text'] or '(No speech detected)'}

{'='*80}
TRANSCRIPT STATS
{'='*80}

Word Count:     {data.get('transcript_word_count', 0):,}
WPM (Words/Min): {data.get('transcript_wpm', 0):.1f}
Duration (sec):  {data.get('video_duration', 0):.1f}
View Count:     {data.get('view_count', 0):,}

{'='*80}
CAPTIONS & METADATA
{'='*80}

Captions:  {data['captions'] or '(None)'}
Hashtags:  {data['hashtags'] or '(None)'}

{'='*80}
"""

        reports_dir = self.reports_dir
        if channel_folder:
            reports_dir = channel_folder / "reports"
            reports_dir.mkdir(exist_ok=True)

        txt_path = reports_dir / f"{platform}_{video_id}.txt"
        with open(txt_path, 'w', encoding='utf-8') as f:
            f.write(report.strip())

    def generate_excel_report(self, json_path=None, channel_folder=None):
        """Generate clean Excel report from JSON data.

        Upsert semantics: new videos are appended; videos already in the sheet
        are updated in place with the latest extracted data (so reprocessing
        fixes bad rows). The user-editable Custom Script / Custom Title columns
        are preserved for existing rows.
        """
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment

        # Determine paths
        if channel_folder:
            source_json = channel_folder / "results.json"
            # Add channel name to filename for dedicated channel folders
            folder_name = channel_folder.name
            if folder_name != "_general_downloads":
                output_excel = channel_folder / f"results_clean_{folder_name}.xlsx"
            else:
                output_excel = channel_folder / "results_clean.xlsx"
        else:
            source_json = json_path or JSON_PATH
            output_excel = DATA_DIR / "results_clean.xlsx"

        if not source_json.exists():
            return None

        try:
            # Load JSON data
            with open(source_json, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if not data:
                return None

            # Dedupe JSON by video_id, keeping the LAST (newest) entry — the
            # exporter appends on every reprocess, so later entries are fresher.
            latest_by_id = {}
            for item in data:
                latest_by_id[item['video_id']] = item

            def _row(item):
                return {
                    'Video ID': item['video_id'],
                    'Platform': item['platform'].upper(),
                    'Channel Name': item.get('channel_name', ''),
                    'Title': item.get('video_title', ''),
                    'Description': item.get('video_description', ''),
                    'URL': item['url'],
                    'Overlay Text': self.clean_text(item['overlay_text']),
                    'Speech Transcript': item['speech_text'] or '(No speech detected)',
                    'Captions': item['captions'] or '(None)',
                    'Hashtags': item['hashtags'] or '(None)',
                    'Date Processed': item['timestamp'].split('T')[0],
                    'View Count': item.get('view_count', 0),
                    'Duration (sec)': item.get('video_duration', 0),
                    'Custom Script': '',
                    'Custom Title': '',
                }

            headers = list(_row(next(iter(latest_by_id.values()))).keys())
            # Columns to leave untouched when updating an existing row
            preserve_cols = {'Custom Script', 'Custom Title'}

            if output_excel.exists():
                try:
                    wb = load_workbook(output_excel)
                except Exception:
                    wb = None
                if wb is not None:
                    ws = wb.active
                    header_row = [cell.value for cell in ws[1]]
                    try:
                        vid_col = header_row.index('Video ID') + 1
                    except ValueError:
                        vid_col = 1
                    # Map existing video_id -> row number
                    row_by_id = {}
                    for r in range(2, ws.max_row + 1):
                        vid = ws.cell(row=r, column=vid_col).value
                        if vid:
                            row_by_id[str(vid)] = r

                    next_row = ws.max_row + 1
                    for vid, item in latest_by_id.items():
                        rowdata = _row(item)
                        if str(vid) in row_by_id:
                            # Update in place, preserving user-edited columns
                            r = row_by_id[str(vid)]
                            for col_idx, key in enumerate(headers, 1):
                                if key in preserve_cols:
                                    continue
                                ws.cell(row=r, column=col_idx, value=rowdata[key])
                        else:
                            # Append new row
                            for col_idx, key in enumerate(headers, 1):
                                ws.cell(row=next_row, column=col_idx, value=rowdata[key])
                            row_by_id[str(vid)] = next_row
                            next_row += 1

                    for row in ws.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    wb.save(output_excel)
                    wb.close()
                    return str(output_excel)

            # Fresh file
            df = pd.DataFrame([_row(i) for i in latest_by_id.values()])
            with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Video Extractions')
                worksheet = writer.sheets['Video Extractions']
                # Set column widths
                for col_letter, w in [('A',15),('B',12),('C',22),('D',50),('E',60),('F',50),
                                       ('G',60),('H',60),('I',40),('J',30),('K',15),('L',12),
                                       ('M',14),('N',60),('O',50)]:
                    worksheet.column_dimensions[col_letter].width = w
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

            return str(output_excel)

        except ImportError:
            return None
        except Exception as e:
            print(f"Excel generation error: {e}")
            return None

    def export_scripts_excel(self, videos, channel_name=None, output_path=None):
        """
        Export script data to Excel for the Script Generator tab.

        Columns: video_id | title | transcript | transcript_word_count
                 | transcript_wpm | suggested_title | hashtag_1
                 | hashtag_2 | generated_script | generated_word_count

        Args:
            videos: List of dicts with keys matching column names above
            channel_name: Channel name (used for folder path if no output_path)
            output_path: Explicit output path (overrides channel-based path)

        Returns:
            Path to saved Excel file, or None on failure
        """
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment

        COLUMNS = [
            ("video_id", "Video ID"),
            ("title", "Title"),
            ("transcript", "Transcript"),
            ("transcript_word_count", "Word Count"),
            ("transcript_wpm", "WPM"),
            ("suggested_title", "Suggested Title"),
            ("hashtag_1", "Hashtag 1"),
            ("hashtag_2", "Hashtag 2"),
            ("generated_script", "Generated Script"),
            ("generated_word_count", "Generated Word Count"),
        ]

        COL_WIDTHS = {
            "A": 15, "B": 50, "C": 80, "D": 12, "E": 10,
            "F": 50, "G": 20, "H": 20, "I": 80, "J": 15,
        }

        if not output_path and channel_name:
            safe_channel = re.sub(r'[^a-zA-Z0-9_-]', '_', channel_name)[:50]
            output_path = (
                self.reports_dir.parent / "channels" / safe_channel / "reports"
                / f"{safe_channel}_scripts.xlsx"
            )
        elif not output_path:
            output_path = DATA_DIR / "scripts_export.xlsx"

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            # Build a set of existing video_ids if file exists (avoid duplicates)
            existing_ids = set()
            if output_path.exists():
                try:
                    existing_wb = load_workbook(output_path)
                    existing_ws = existing_wb.active
                    # Find the video_id column index
                    header_row = [cell.value for cell in existing_ws[1]]
                    vid_idx = None
                    for i, h in enumerate(header_row):
                        if h == "Video ID":
                            vid_idx = i
                            break
                    if vid_idx is not None:
                        for row in existing_ws.iter_rows(
                            min_row=2, values_only=True
                        ):
                            if row[vid_idx]:
                                existing_ids.add(str(row[vid_idx]))
                    existing_wb.close()
                except Exception:
                    pass  # File exists but can't be read — overwrite it

            # Filter out videos already in the sheet
            new_videos = [v for v in videos if v.get("video_id") not in existing_ids]

            if not new_videos and existing_ids:
                return str(output_path)  # Nothing new to add

            # Determine write mode
            if existing_ids:
                # Append mode — load existing workbook
                wb = load_workbook(output_path)
                ws = wb.active
                start_row = ws.max_row + 1
            else:
                # Fresh file
                wb = Workbook()
                ws = wb.active
                ws.title = "Scripts"
                # Write header
                for col_idx, (_, label) in enumerate(COLUMNS, 1):
                    ws.cell(row=1, column=col_idx, value=label)
                start_row = 2

            # Write new data rows
            for row_offset, video in enumerate(new_videos):
                row_num = start_row + row_offset
                data_key = next(
                    (k for k, _ in [
                        ("video_id", None), ("title", None),
                        ("speech_text", None), ("transcript_word_count", None),
                        ("transcript_wpm", None), ("suggested_title", None),
                        ("hashtag_1", None), ("hashtag_2", None),
                        ("generated_script", None), ("generated_word_count", None),
                    ]),
                    None,
                )
                ws.cell(row=row_num, column=1, value=video.get("video_id", ""))
                ws.cell(row=row_num, column=2, value=video.get("title", ""))
                ws.cell(row=row_num, column=3, value=video.get("speech_text", "") or video.get("transcript", ""))
                ws.cell(row=row_num, column=4, value=video.get("transcript_word_count", 0))
                ws.cell(row=row_num, column=5, value=video.get("transcript_wpm", 0.0))
                ws.cell(row=row_num, column=6, value=video.get("suggested_title", ""))
                ws.cell(row=row_num, column=7, value=video.get("hashtag_1", ""))
                ws.cell(row=row_num, column=8, value=video.get("hashtag_2", ""))
                ws.cell(row=row_num, column=9, value=video.get("generated_script", ""))
                ws.cell(row=row_num, column=10, value=video.get("generated_word_count", 0))

            # Set column widths
            for col_letter, width in COL_WIDTHS.items():
                ws.column_dimensions[col_letter].width = width

            # Enable text wrapping on all data cells
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            wb.save(output_path)
            wb.close()
            return str(output_path)

        except Exception as e:
            print(f"Scripts Excel export error: {e}")
            return None
