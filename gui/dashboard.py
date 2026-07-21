import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
from datetime import datetime
from pathlib import Path
import re
import os

# Import TTS calibration tables from script_generator (single source of truth)
from core.script_generator import (
    _VOICE_WPM, _TONE_PACING, _BASE_TTS_SPEED, calc_target_word_count
)

# ═══════════════════════════════════════════════════════════════
class ScrollableFrame(tk.Frame):
    """A frame that wraps its contents in a canvas with a vertical scrollbar.

    Usage — replace:
        parent = tk.Frame(tab)
        parent.pack(fill=tk.BOTH, expand=True)
    with:
        parent = ScrollableFrame(tab)
        parent.pack(fill=tk.BOTH, expand=True)
        # then pack/grid onto parent.inner instead of parent
    """
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)

        self.canvas = tk.Canvas(self, highlightthickness=0, borderwidth=0)
        self.vsb = tk.Scrollbar(self, orient=tk.VERTICAL, command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Inner frame that holds the actual content
        self.inner = tk.Frame(self.canvas)
        self.inner_id = self.canvas.create_window((0, 0), window=self.inner, anchor=tk.NW)

        # Resize the inner frame's width to match the canvas
        def _configure_canvas(event):
            self.canvas.itemconfig(self.inner_id, width=event.width)
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.canvas.bind("<Configure>", _configure_canvas)

        # Propagate inner frame's height changes to the scroll region
        def _configure_inner(event):
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.inner.bind("<Configure>", _configure_inner)

        # Mouse-wheel scrolling — bind on the canvas (not bind_all) so it only
        # fires when the mouse is actually over this scrollable area.
        def _on_mousewheel(event):
            # event.delta: Windows (+/-120), event.num: Linux 4/5
            delta = -1 * (event.delta // 120) if event.delta else (-1 if event.num == 4 else 1)
            self.canvas.yview_scroll(delta, tk.UNITS)
        self.canvas.bind("<Enter>", lambda e: self.canvas.bind_all("<MouseWheel>", _on_mousewheel))
        self.canvas.bind("<Leave>", lambda e: self.canvas.unbind_all("<MouseWheel>"))
        # Also bind Linux scroll events directly (bind_all won't catch Button-4/5)
        self.canvas.bind("<Button-4>", _on_mousewheel)
        self.canvas.bind("<Button-5>", _on_mousewheel)

    def destroy(self):
        """Unbind mousewheel when this frame is destroyed."""
        try:
            self.canvas.unbind_all("<MouseWheel>")
            self.canvas.unbind_all("<Button-4>")
            self.canvas.unbind_all("<Button-5>")
        except Exception:
            pass
        super().destroy()


def _extract_yt_id(url: str) -> str:
    """Extract YouTube video ID from various URL formats."""
    if not url:
        return ""
    m = re.search(r'(?:youtube\.com/watch\?v=|youtu\.be/|youtube\.com/shorts/)([a-zA-Z0-9_-]{11})', url)
    return m.group(1) if m else ""


def _extract_video_id(name: str) -> str:
    """Extract actual video ID from filename like 'Title [abc123].mp4' or 'abc123.mp4'."""
    stem = name.rsplit('.', 1)[0] if '.' in name else name
    m = re.search(r'\[([a-zA-Z0-9_-]{6,})\]', stem)
    if m:
        return m.group(1)
    return stem


class Dashboard:
    def __init__(self, root, processor):
        self.root = root
        self.processor = processor
        self.root.title("Video Text Extractor - Professional Edition")
        self.root.geometry("1000x700")
        self.root.configure(bg='#f0f0f0')
        self.stop_processing = False
        self._last_excel_dir = None

        # Feature selections
        self.features = {
            'download_video': tk.BooleanVar(value=True),
            'download_audio': tk.BooleanVar(value=False),
            'voice_only': tk.BooleanVar(value=False),
            'extract_ocr': tk.BooleanVar(value=True),
            'extract_speech': tk.BooleanVar(value=True),
            'extract_captions': tk.BooleanVar(value=False),
            'extract_metadata': tk.BooleanVar(value=True),
            'auto_excel': tk.BooleanVar(value=True),
            'force_reprocess': tk.BooleanVar(value=False),
        }
        self.audio_format = tk.StringVar(value="mp3")
        self.video_quality = tk.StringVar(value="best")

        # Model selection
        self.gemini_model = tk.StringVar(value="gemini-3.5-flash")
        self.whisper_model = tk.StringVar(value="base")

        self.create_modern_ui()
        self.check_instagram_auth()
        self._load_gemini_config()
        # Apply default model selections
        self._on_gemini_model_change()
        self._on_whisper_model_change()

    def create_modern_ui(self):
        # Header
        header_frame = tk.Frame(self.root, bg='#2c3e50', height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)

        tk.Label(header_frame, text="🎥 Video Text Extractor",
                font=("Arial", 18, "bold"), bg='#2c3e50', fg='white').pack(side=tk.LEFT, padx=20, pady=15)

        # Platform selector in header
        tk.Label(header_frame, text="Platform:", bg='#2c3e50', fg='white',
                font=("Arial", 10)).pack(side=tk.LEFT, padx=(0, 5))
        self.platform_var = tk.StringVar(value="youtube")
        platform_combo = ttk.Combobox(header_frame, textvariable=self.platform_var,
                                     values=["youtube", "tiktok", "instagram", "facebook", "xiaohongshu", "other", "bilibili"],
                                     state="readonly", width=14)
        platform_combo.pack(side=tk.LEFT, padx=5)

        # Instagram auth status in header
        self.auth_label = tk.Label(header_frame, text="", bg='#2c3e50', fg='white', font=("Arial", 9))
        self.auth_label.pack(side=tk.RIGHT, padx=20)

        # Main content area with tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Tab 1: Quick Process
        self.quick_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.quick_tab, text="  Quick Process  ")
        self.create_quick_tab()

        # Tab 2: Metadata Scan
        self.metadata_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.metadata_tab, text="  Metadata Scan  ")
        self.create_metadata_tab()

        # Tab 3: Script Studio (merged: Rewrite transcript + Write Story from Video)
        self.create_script_studio_tab()

        # Tab 5: Case Commentary
        self.case_commentary_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.case_commentary_tab, text="  Case Commentary  ")
        self._cc_scroller = ScrollableFrame(self.case_commentary_tab)
        self._cc_scroller.pack(fill=tk.BOTH, expand=True)
        self.create_case_commentary_tab()

        # Tab 6: Settings
        self.settings_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.settings_tab, text="  Settings  ")
        self._settings_scroller = ScrollableFrame(self.settings_tab)
        self._settings_scroller.pack(fill=tk.BOTH, expand=True)
        self.create_settings_tab()

        # Tab 7: Activity Log
        self.log_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.log_tab, text="  Activity Log  ")
        self.create_log_tab()

        # Status bar at bottom
        self.create_status_bar()

    def create_quick_tab(self):
        """Quick Process tab - for immediate video processing"""
        # Scrollable wrapper so everything fits any window size
        scroller = ScrollableFrame(self.quick_tab)
        scroller.pack(fill=tk.BOTH, expand=True)
        SW = scroller.inner  # shorthand — pack/grid onto this instead of self.quick_tab

        # Input Section
        input_section = tk.LabelFrame(SW, text="📥 Input Source",
                                     font=("Arial", 11, "bold"), padx=15, pady=15)
        input_section.pack(fill=tk.X, padx=10, pady=10)

        # URL Input
        tk.Label(input_section, text="Video URLs or Channel:",
                font=("Arial", 10)).grid(row=0, column=0, sticky=tk.W, pady=5)

        url_frame = tk.Frame(input_section)
        url_frame.grid(row=1, column=0, columnspan=3, sticky=tk.EW, pady=5)

        # Create a Text widget to support multiple lines (instead of Entry)
        input_text_frame = tk.Frame(url_frame)
        input_text_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.input_text = tk.Text(input_text_frame, font=("Arial", 10), height=3, wrap=tk.WORD)
        self.input_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Add scrollbar for the text widget
        scrollbar = tk.Scrollbar(input_text_frame, command=self.input_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.input_text.config(yscrollcommand=scrollbar.set)

        tk.Button(url_frame, text="Browse File", command=self.browse_url_file,
                 bg='#607D8B', fg='white', width=12).pack(side=tk.LEFT, padx=(5, 0))

        tk.Button(url_frame, text="Browse Folder", command=self.browse_folder,
                 bg='#FF9800', fg='white', width=12).pack(side=tk.LEFT, padx=(5, 0))

        # Processing Options Section
        options_section = tk.LabelFrame(SW, text="⚙️ Processing Options",
                                       font=("Arial", 11, "bold"), padx=15, pady=15)
        options_section.pack(fill=tk.X, padx=10, pady=10)

        # Left column - Extraction features
        left_col = tk.Frame(options_section)
        left_col.grid(row=0, column=0, sticky=tk.W, padx=10)

        tk.Label(left_col, text="Select Features to Extract:",
                font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(0, 5))

        tk.Checkbutton(left_col, text="📥 Download Videos",
                      variable=self.features['download_video'],
                      font=("Arial", 10), command=self.update_feature_status).pack(anchor=tk.W, pady=2)

        tk.Checkbutton(left_col, text="🔍 OCR - Overlay Text",
                      variable=self.features['extract_ocr'],
                      font=("Arial", 10), command=self.update_feature_status).pack(anchor=tk.W, pady=2)

        tk.Checkbutton(left_col, text="🎤 Speech Transcription",
                      variable=self.features['extract_speech'],
                      font=("Arial", 10), command=self.update_feature_status).pack(anchor=tk.W, pady=2)

        tk.Checkbutton(left_col, text="📝 YouTube Captions (no download)",
                      variable=self.features['extract_captions'],
                      font=("Arial", 10), command=self.update_feature_status).pack(anchor=tk.W, pady=2)

        tk.Checkbutton(left_col, text="📋 Metadata (Captions/Hashtags)",
                      variable=self.features['extract_metadata'],
                      font=("Arial", 10), command=self.update_feature_status).pack(anchor=tk.W, pady=2)

        # Right column - Audio & Output features
        right_col = tk.Frame(options_section)
        right_col.grid(row=0, column=1, sticky=tk.W, padx=50)

        tk.Label(right_col, text="Audio & Output Options:",
                font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(0, 5))

        tk.Checkbutton(right_col, text="🎵 Download Audio Only",
                      variable=self.features['download_audio'],
                      font=("Arial", 10), command=self.update_feature_status).pack(anchor=tk.W, pady=2)

        audio_format_frame = tk.Frame(right_col)
        audio_format_frame.pack(anchor=tk.W, padx=20, pady=2)
        tk.Label(audio_format_frame, text="Format:", font=("Arial", 9)).pack(side=tk.LEFT)
        tk.Radiobutton(audio_format_frame, text="MP3", variable=self.audio_format,
                      value="mp3", font=("Arial", 8)).pack(side=tk.LEFT, padx=3)
        tk.Radiobutton(audio_format_frame, text="WAV", variable=self.audio_format,
                      value="wav", font=("Arial", 8)).pack(side=tk.LEFT, padx=3)

        # Voice-only option (sub-option under Download Audio)
        tk.Checkbutton(right_col, text="🎙️ Voice Only (Skip if no voiceover)",
                      variable=self.features['voice_only'],
                      font=("Arial", 9), state=tk.NORMAL).pack(anchor=tk.W, padx=20, pady=2)

        tk.Checkbutton(right_col, text="📊 Auto-generate Excel Report",
                      variable=self.features['auto_excel'],
                      font=("Arial", 10)).pack(anchor=tk.W, pady=2)

        # Video Quality selector
        quality_frame = tk.Frame(right_col)
        quality_frame.pack(anchor=tk.W, padx=0, pady=(4, 2))
        tk.Label(quality_frame, text="🎬 Video Quality:",
                font=("Arial", 9, "bold")).pack(side=tk.LEFT)
        quality_combo = ttk.Combobox(quality_frame, textvariable=self.video_quality,
                                    values=["best", "2160p (4K)", "1440p", "1080p", "720p",
                                            "480p", "360p"],
                                    state="readonly", width=12)
        quality_combo.pack(side=tk.LEFT, padx=(5, 0))

        tk.Checkbutton(right_col, text="🔄 Force Reprocess (ignore processed status)",
                      variable=self.features['force_reprocess'],
                      font=("Arial", 9), fg='#FF5722').pack(anchor=tk.W, pady=2)

        # Feature status label
        self.feature_status = tk.Label(options_section, text="",
                                      font=("Arial", 9), fg='#666')
        self.feature_status.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(10, 0))
        self.update_feature_status()

        # Model Selection row
        model_frame = tk.Frame(options_section)
        model_frame.grid(row=2, column=0, columnspan=2, sticky=tk.EW, pady=(4, 0))

        # Gemini Model
        tk.Label(model_frame, text="🤖 Gemini Model:",
                font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(10, 2))
        gemini_model_combo = ttk.Combobox(model_frame, textvariable=self.gemini_model,
                                           values=["gemini-3.5-flash", "gemini-2.5-flash",
                                                   "gemini-3.1-flash-lite", "gemini-2.0-flash-lite",
                                                   "gemini-2.0-flash", "gemini-1.5-flash"],
                                           state="readonly", width=18)
        gemini_model_combo.pack(side=tk.LEFT, padx=(0, 20))
        gemini_model_combo.bind("<<ComboboxSelected>>", self._on_gemini_model_change)

        # Whisper Model
        tk.Label(model_frame, text="🎙️ Whisper Model:",
                font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(0, 2))
        whisper_model_combo = ttk.Combobox(model_frame, textvariable=self.whisper_model,
                                            values=["tiny", "base", "small", "medium", "large"],
                                            state="readonly", width=10)
        whisper_model_combo.pack(side=tk.LEFT, padx=(0, 10))
        whisper_model_combo.bind("<<ComboboxSelected>>", self._on_whisper_model_change)

        tk.Label(model_frame, text="(tiny=fast/low-quality, large=slow/high-quality)",
                font=("Arial", 8, "italic"), fg="#888").pack(side=tk.LEFT)

        # ── Gemini API Configuration ────────────────────────────
        api_frame = tk.LabelFrame(SW, text="🔑 Gemini API Configuration",
                                  font=("Arial", 11, "bold"), padx=15, pady=15)
        api_frame.pack(fill=tk.X, padx=10, pady=10)

        # Row: entry + action buttons
        api_row = tk.Frame(api_frame)
        api_row.pack(fill=tk.X, pady=5)

        tk.Label(api_row, text="🔑 Gemini API Keys:", font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        self.gemini_api_key_entry = tk.Entry(api_row, font=("Arial", 10), width=40)
        self.gemini_api_key_entry.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)

        tk.Button(api_row, text="➕ Add Key", font=("Arial", 9),
                 bg='#4CAF50', fg='white',
                 command=self._add_api_key).pack(side=tk.LEFT, padx=2)
        tk.Button(api_row, text="🔍 Test All", font=("Arial", 9),
                 bg='#FF9800', fg='white',
                 command=self._test_all_keys).pack(side=tk.LEFT, padx=2)
        tk.Button(api_row, text="💾 Save", font=("Arial", 9),
                 bg='#2196F3', fg='white',
                 command=self._save_gemini_config).pack(side=tk.LEFT, padx=2)

        # Label: hint
        tk.Label(api_row, text="(paste key → Add)", font=("Arial", 8), fg="#888").pack(side=tk.LEFT, padx=2)

        # ── Key listbox + controls ──────────────────────────
        keys_frame = tk.Frame(api_frame)
        keys_frame.pack(fill=tk.X, pady=2)

        # Listbox (left side)
        list_frame = tk.Frame(keys_frame)
        list_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        scrollbar = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        self.keys_listbox = tk.Listbox(list_frame, font=("Consolas", 9),
                                       height=4, yscrollcommand=scrollbar.set,
                                       selectmode=tk.SINGLE, exportselection=False)
        scrollbar.config(command=self.keys_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.keys_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Buttons (right side)
        btn_frame = tk.Frame(keys_frame)
        btn_frame.pack(side=tk.RIGHT, fill=tk.Y)

        self.key_remove_btn = tk.Button(btn_frame, text="🗑 Remove", font=("Arial", 8),
                                        command=self._remove_api_key, width=10)
        self.key_remove_btn.pack(pady=2)
        self.key_up_btn = tk.Button(btn_frame, text="▲ Move Up", font=("Arial", 8),
                                    command=self._move_key_up, width=10)
        self.key_up_btn.pack(pady=2)
        self.key_down_btn = tk.Button(btn_frame, text="▼ Move Down", font=("Arial", 8),
                                      command=self._move_key_down, width=10)
        self.key_down_btn.pack(pady=2)

        # Active key status
        self.gemini_status_label = tk.Label(api_frame, text="", font=("Arial", 9))
        self.gemini_status_label.pack(anchor=tk.W, pady=(2, 2))

        # ── Service Account JSON file row ───────────────────
        sa_row = tk.Frame(api_frame)
        sa_row.pack(fill=tk.X, pady=2)

        tk.Label(sa_row, text="OR Service Account JSON:", font=("Arial", 9, "bold")).pack(side=tk.LEFT)
        self.sa_path_var = tk.StringVar(value="")
        tk.Label(sa_row, textvariable=self.sa_path_var, font=("Arial", 8), fg="#666",
                width=40, anchor=tk.W).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        tk.Button(sa_row, text="Browse", font=("Arial", 8),
                 bg='#607D8B', fg='white', command=self._browse_service_account).pack(side=tk.LEFT, padx=2)

        # Action Buttons Section
        action_section = tk.Frame(SW)
        action_section.pack(fill=tk.X, padx=10, pady=10)

        tk.Button(action_section, text="▶ Process Now", command=self.process_input,
                 bg='#4CAF50', fg='white', font=("Arial", 12, "bold"),
                 height=2, width=15).pack(side=tk.LEFT, padx=5)

        tk.Button(action_section, text="⏹ Stop", command=self.stop_process,
                 bg='#f44336', fg='white', font=("Arial", 12, "bold"),
                 height=2, width=12).pack(side=tk.LEFT, padx=5)

        tk.Button(action_section, text="📊 Export Results", command=self.export_data,
                 bg='#2196F3', fg='white', font=("Arial", 11),
                 height=2, width=12).pack(side=tk.LEFT, padx=5)

        # Progress
        progress_frame = tk.LabelFrame(SW, text="Progress",
                                      font=("Arial", 10, "bold"), padx=10, pady=10)
        progress_frame.pack(fill=tk.X, padx=10, pady=10)

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var,
                                           maximum=100, length=400)
        self.progress_bar.pack(fill=tk.X)

        self.progress_label = tk.Label(progress_frame, text="Ready",
                                      font=("Arial", 9), fg='#666')
        self.progress_label.pack(pady=(5, 0))

    def create_metadata_tab(self):
        """Metadata Scan tab - for bulk metadata extraction"""
        # Scrollable wrapper
        scroller = ScrollableFrame(self.metadata_tab)
        scroller.pack(fill=tk.BOTH, expand=True)
        SW = scroller.inner

        # Info banner
        info_frame = tk.Frame(SW, bg='#e3f2fd', relief=tk.RAISED, borderwidth=1)
        info_frame.pack(fill=tk.X, padx=10, pady=10)

        tk.Label(info_frame, text="💡 Metadata Scan: Quickly extract metadata from entire channels without downloading videos",
                bg='#e3f2fd', font=("Arial", 10, "italic"), fg='#1976d2').pack(pady=10, padx=10)

        # Scan section
        scan_section = tk.LabelFrame(SW, text="🔍 Metadata Scanner",
                                    font=("Arial", 11, "bold"), padx=15, pady=15)
        scan_section.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        tk.Label(scan_section, text="Channel/Profile URL:",
                font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=5)

        self.metadata_url = tk.Entry(scan_section, font=("Arial", 10))
        self.metadata_url.pack(fill=tk.X, pady=5)

        # Platform info
        platform_info_frame = tk.Frame(scan_section)
        platform_info_frame.pack(fill=tk.X, pady=10)

        tk.Label(platform_info_frame, text="✅ YouTube: Unlimited videos, fast",
                font=("Arial", 9), fg='green').pack(anchor=tk.W, padx=20)
        tk.Label(platform_info_frame, text="✅ TikTok: Unlimited videos, fast",
                font=("Arial", 9), fg='green').pack(anchor=tk.W, padx=20)
        tk.Label(platform_info_frame, text="⚠️ Instagram: Max 50 videos, requires login",
                font=("Arial", 9), fg='orange').pack(anchor=tk.W, padx=20)
        tk.Label(platform_info_frame, text="⚠️ Facebook: Public pages only, less reliable",
                font=("Arial", 9), fg='orange').pack(anchor=tk.W, padx=20)

        # Column selection
        columns_frame = tk.LabelFrame(scan_section, text="📋 Excel Columns to Include",
                                     font=("Arial", 10, "bold"))
        columns_frame.pack(fill=tk.X, pady=10)

        # Initialize column selection vars
        self.metadata_columns = {
            'video_id': tk.BooleanVar(value=True),
            'title': tk.BooleanVar(value=True),
            'channel_name': tk.BooleanVar(value=True),
            'description': tk.BooleanVar(value=True),
            'url': tk.BooleanVar(value=True),
            'duration': tk.BooleanVar(value=True),
            'view_count': tk.BooleanVar(value=True),
        }

        # Create checkboxes in grid layout
        col_checkboxes_frame = tk.Frame(columns_frame)
        col_checkboxes_frame.pack(fill=tk.X, padx=10, pady=10)

        column_labels = {
            'video_id': 'Video ID',
            'title': 'Title',
            'channel_name': 'Channel/Username',
            'description': 'Description/Caption',
            'url': 'URL',
            'duration': 'Duration',
            'view_count': 'View Count'
        }

        row = 0
        col = 0
        for key, label in column_labels.items():
            cb = tk.Checkbutton(col_checkboxes_frame, text=label,
                               variable=self.metadata_columns[key],
                               font=("Arial", 9))
            cb.grid(row=row, column=col, sticky=tk.W, padx=15, pady=3)
            col += 1
            if col > 2:  # 3 columns per row
                col = 0
                row += 1

        # Select/Deselect all buttons
        btn_frame = tk.Frame(columns_frame)
        btn_frame.pack(pady=(0, 10))

        tk.Button(btn_frame, text="Select All", command=lambda: self.toggle_all_columns(True),
                 bg='#4CAF50', fg='white', font=("Arial", 8), width=10).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Deselect All", command=lambda: self.toggle_all_columns(False),
                 bg='#f44336', fg='white', font=("Arial", 8), width=10).pack(side=tk.LEFT, padx=5)

        # Popularity filter (YouTube viral video filter)
        pop_frame = tk.LabelFrame(scan_section, text="🔥 Popularity Filter (YouTube only)",
                                 font=("Arial", 10, "bold"))
        pop_frame.pack(fill=tk.X, pady=10)

        pop_inner = tk.Frame(pop_frame)
        pop_inner.pack(fill=tk.X, padx=10, pady=8)

        # Min views
        tk.Label(pop_inner, text="Min Views:", font=("Arial", 9, "bold")).grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.min_views_var = tk.StringVar(value="0")
        tk.Entry(pop_inner, textvariable=self.min_views_var, font=("Arial", 9), width=14).grid(row=0, column=1, padx=(0, 5))
        tk.Label(pop_inner, text="Type 1000000 for 1M, 500000 for 500K", font=("Arial", 8), fg="#666").grid(row=0, column=2, sticky=tk.W)

        # Top N
        tk.Label(pop_inner, text="Top N Videos:", font=("Arial", 9, "bold")).grid(row=1, column=0, sticky=tk.W, padx=(0, 5), pady=(5, 0))
        self.top_n_var = tk.StringVar(value="0")
        tk.Entry(pop_inner, textvariable=self.top_n_var, font=("Arial", 9), width=14).grid(row=1, column=1, sticky=tk.W, padx=(0, 5), pady=(5, 0))
        tk.Label(pop_inner, text="0 = all matching videos, e.g. 10 = only top 10", font=("Arial", 8), fg="#666").grid(row=1, column=2, sticky=tk.W, pady=(5, 0))

        # Info text
        tk.Label(pop_frame, text="💡 Leave as 0 to process ALL videos (no popularity filter). " +
                "Only works for YouTube channels.",
                font=("Arial", 8, "italic"), fg="#888").pack(anchor=tk.W, padx=10, pady=(0, 5))

        # Scan button
        scan_btn_frame = tk.Frame(scan_section)
        scan_btn_frame.pack(pady=20)

        tk.Button(scan_btn_frame, text="🚀 Start Metadata Scan",
                 command=self.start_metadata_scan,
                 bg='#673AB7', fg='white', font=("Arial", 12, "bold"),
                 height=2, width=20).pack()

        # Results info
        results_frame = tk.LabelFrame(scan_section, text="📊 What You'll Get",
                                     font=("Arial", 10, "bold"))
        results_frame.pack(fill=tk.X, pady=10)

        tk.Label(results_frame, text="• Excel file with all video metadata (titles, URLs, durations)",
                font=("Arial", 9)).pack(anchor=tk.W, padx=10, pady=2)
        tk.Label(results_frame, text="• TXT file with all URLs for batch processing",
                font=("Arial", 9)).pack(anchor=tk.W, padx=10, pady=2)
        tk.Label(results_frame, text="• Load URLs with 'Browse File' for selective processing",
                font=("Arial", 9)).pack(anchor=tk.W, padx=10, pady=2)

    def create_script_studio_tab(self):
        """Unified Script Studio tab - merges Rewrite and Write-Story modes."""
        # Create the master tab
        self.script_studio_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.script_studio_tab, text="  Script Studio  ")

        # Top frame: Mode selector
        top_bar = tk.Frame(self.script_studio_tab, bg='#E3F2FD', padx=15, pady=12)
        top_bar.pack(fill=tk.X, side=tk.TOP)

        tk.Label(top_bar, text="📋 What do you want to do?",
                 font=("Arial", 11, "bold"), bg='#E3F2FD').pack(side=tk.LEFT, padx=(0, 10))

        # Display values (user-friendly); the handler maps them back to a mode key
        self._script_mode_labels = [
            "Rewrite Existing Script — video already has narration/commentary",
            "Write Story from Video — Gemini watches the video (for dialogue-only clips)",
        ]
        self.script_mode_var = tk.StringVar(value=self._script_mode_labels[0])
        mode_combo = ttk.Combobox(top_bar, textvariable=self.script_mode_var,
                                   values=self._script_mode_labels,
                                   state="readonly", width=60, font=("Arial", 10))
        mode_combo.current(0)
        mode_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)

        tk.Label(top_bar, text="   ", bg='#E3F2FD').pack(side=tk.LEFT)

        # Switch UI when the user picks a different mode
        mode_combo.bind("<<ComboboxSelected>>", lambda _e: self._switch_script_mode())

        # Container for the two mode UIs
        self.script_modes_container = tk.Frame(self.script_studio_tab)
        self.script_modes_container.pack(fill=tk.BOTH, expand=True)

        # Build the two sub-frames (reuse existing builders)
        self.script_tab = tk.Frame(self.script_modes_container)
        self.video_script_tab = tk.Frame(self.script_modes_container)

        self.create_script_tab()          # builds into self.script_tab
        self.create_video_script_tab()    # builds into self.video_script_tab

        # Show the initial mode
        self._switch_script_mode()

    def _switch_script_mode(self):
        """Show the active mode's UI, hide the other."""
        raw = self.script_mode_var.get()
        # Map display string back to mode key
        if "Rewrite Existing" in raw:
            mode = "rewrite"
        elif "Write Story from Video" in raw:
            mode = "write_story"
        else:
            mode = raw  # fallback to raw value

        # Hide both
        self.script_tab.pack_forget()
        self.video_script_tab.pack_forget()

        # Show the selected one
        if mode == "rewrite":
            self.script_tab.pack(fill=tk.BOTH, expand=True)
        elif mode == "write_story":
            self.video_script_tab.pack(fill=tk.BOTH, expand=True)

    def create_script_tab(self):
        """Rewrite mode UI - Gemini-powered cinematic narration from transcripts.

        Built into self.script_tab, which is a sub-frame of the merged
        Script Studio tab (set up by create_script_studio_tab).
        """
        # Scrollable wrapper for this sub-tab
        _scr = ScrollableFrame(self.script_tab)
        _scr.pack(fill=tk.BOTH, expand=True)
        _SW = _scr.inner

        # Info banner
        info_frame = tk.Frame(_SW, bg='#fff3e0', relief=tk.RAISED, borderwidth=1)
        info_frame.pack(fill=tk.X, padx=10, pady=10)

        tk.Label(info_frame, text="🤖 Script Generator: Convert transcripts into cinematic narration using Gemini AI",
                bg='#fff3e0', font=("Arial", 10, "italic"), fg='#e65100').pack(pady=10, padx=10)

        # Language + Preset row
        settings_frame = tk.LabelFrame(_SW, text="⚙️ Script Settings",
                                      font=("Arial", 11, "bold"), padx=15, pady=15)
        settings_frame.pack(fill=tk.X, padx=10, pady=10)

        settings_row = tk.Frame(settings_frame)
        settings_row.pack(fill=tk.X, pady=5)

        tk.Label(settings_row, text="Language:", font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        self.script_lang_var = tk.StringVar(value="English")
        LANGUAGES = [
            "English", "Russian", "German", "Arabic",
            "Hindi", "Korean", "Japanese",
        ]
        lang_combo = ttk.Combobox(settings_row, textvariable=self.script_lang_var,
                                   values=LANGUAGES,
                                   state="readonly", width=14)
        lang_combo.pack(side=tk.LEFT, padx=5)

        tk.Label(settings_row, text="Preset:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=(20, 0))
        self.script_style_var = tk.StringVar(value="Movie/Story Recap (dramatic)")
        self.script_style_combo = ttk.Combobox(settings_row, textvariable=self.script_style_var,
                                                values=["Movie/Story Recap (dramatic)"],
                                                state="readonly", width=42)
        self.script_style_combo.pack(side=tk.LEFT, padx=10)

        tk.Button(settings_row, text="✏️ Manage Presets",
                  font=("Arial", 8),
                  bg='#455A64', fg='white',
                  command=lambda: self._open_prompt_manager(
                      self.REWRITE_MODE_SLUGS,
                      self._slug_for_name(self.script_style_var.get(),
                                          self.REWRITE_MODE_SLUGS))
                  ).pack(side=tk.LEFT, padx=2)

        # Update the details panel whenever a preset is picked
        self.script_style_combo.bind(
            "<<ComboboxSelected>>", lambda _e: self._update_rewrite_preset_help())

        # Internal state: niche_var still exists (consumed downstream) but no UI
        self.niche_var = tk.StringVar(value="")

        # ── TTS Engine selector ─────────────────────────────────
        tts_frame = tk.LabelFrame(_SW, text="🎤 TTS Engine",
                                  font=("Arial", 10, "bold"), padx=10, pady=5)
        tts_frame.pack(fill=tk.X, padx=10, pady=(2, 6))
        self.tts_engine_var = tk.StringVar(value="cloud")
        tk.Radiobutton(tts_frame, text="Cloud TTS (Standard — 160 WPM)",
                       variable=self.tts_engine_var,
                       value="cloud", font=("Arial", 9)).pack(anchor=tk.W, padx=5, pady=1)
        tk.Radiobutton(tts_frame, text="Qwen3 TTS (Slower — ~110 WPM)",
                       variable=self.tts_engine_var,
                       value="qwen3", font=("Arial", 9)).pack(anchor=tk.W, padx=5, pady=1)
        tk.Radiobutton(tts_frame, text="Gemini TTS (auto speed)",
                       variable=self.tts_engine_var,
                       value="gemini", font=("Arial", 9)).pack(anchor=tk.W, padx=5, pady=1)

        # ── About this preset (details panel) ───────────────────
        help_box = tk.LabelFrame(_SW, text="📖 About this preset",
                                 font=("Arial", 10, "bold"), fg="#1565C0")
        help_box.pack(fill=tk.X, padx=10, pady=(0, 6))
        self.rewrite_help_label = tk.Label(
            help_box, text="Select a preset to see what it does.",
            font=("Arial", 9), justify=tk.LEFT, anchor=tk.W,
            wraplength=900, fg="#333")
        self.rewrite_help_label.pack(fill=tk.X, padx=10, pady=8)

        # Action buttons
        action_frame = tk.Frame(_SW)
        action_frame.pack(fill=tk.X, padx=10, pady=15)

        tk.Button(action_frame, text="🚀 Generate Scripts + Export Excel",
                 command=self._start_script_generation,
                 bg='#FF6F00', fg='white', font=("Arial", 12, "bold"),
                 height=2, width=35).pack(side=tk.LEFT, padx=5)

        # Info text
        info_text = tk.LabelFrame(_SW, text="ℹ️ How It Works",
                                 font=("Arial", 10, "bold"))
        info_text.pack(fill=tk.X, padx=10, pady=10)

        instructions = """1. Add your Gemini API Key(s) above — paste a key and click "Add Key"
   • Multiple keys from different accounts auto-fallback when quota runs out
   • Use ▲ ▼ to prioritise which key is tried first
2. Pick a Preset (the script style) and target language from the dropdowns
   • Each preset is a complete prompt — it already defines the niche & style
3. Click "Generate Scripts + Export Excel" — it will:
   a. Ask you to select an Excel file (the results_clean.xlsx from your downloads)
   b. Read all video transcripts from the Excel file
   c. Generate cinematic narration in the selected language via Gemini
   d. Save to new columns: {language}script, {language} title, {language} description
   e. Output a new Excel file with "_with_scripts" suffix
4. To add another language later — select it and click Generate again
5. Your video automation tool reads from the appropriate language column (e.g. "custom script" or "englishscript")"""
        tk.Label(info_text, text=instructions, font=("Arial", 9),
                justify=tk.LEFT, padx=10, pady=10).pack()

        # Fresh prompt styles from config
        self.root.after(100, self._refresh_script_styles)

    # ── Multi-key management ───────────────────────────────

    def _refresh_keys_listbox(self):
        """Refresh the listbox display from the saved keys list"""
        self.keys_listbox.delete(0, tk.END)
        from core.script_generator import ScriptGenerator
        gen = ScriptGenerator(api_keys=getattr(self, '_saved_api_keys', []))
        for i, key in enumerate(getattr(self, '_saved_api_keys', [])):
            masked = gen._mask_key(key)
            self.keys_listbox.insert(tk.END, f"  {masked}")
        # Update active key status
        if getattr(self, '_saved_api_keys', []):
            self.gemini_status_label.config(
                text=f"✅ {len(self._saved_api_keys)} key(s) loaded • Active: #{gen._current_key_index + 1} ({gen.active_key_label})",
                fg="green"
            )
        else:
            self.gemini_status_label.config(text="No API keys configured", fg="gray")

    def _add_api_key(self):
        """Add the key from the entry field to the saved list"""
        key = self.gemini_api_key_entry.get().strip()
        if not key:
            messagebox.showwarning("Empty Key", "Paste an API key into the field first.")
            return
        if not hasattr(self, '_saved_api_keys'):
            self._saved_api_keys = []
        if key in self._saved_api_keys:
            messagebox.showinfo("Duplicate", "This API key is already in the list.")
            return
        self._saved_api_keys.append(key)
        self.gemini_api_key_entry.delete(0, tk.END)
        self._refresh_keys_listbox()

    def _remove_api_key(self):
        """Remove the selected key from the list"""
        sel = self.keys_listbox.curselection()
        if not sel:
            messagebox.showinfo("Select First", "Select a key from the list to remove.")
            return
        idx = sel[0]
        if hasattr(self, '_saved_api_keys') and 0 <= idx < len(self._saved_api_keys):
            self._saved_api_keys.pop(idx)
            self._refresh_keys_listbox()

    def _move_key_up(self):
        """Move selected key up in priority"""
        sel = self.keys_listbox.curselection()
        if not sel or sel[0] == 0:
            return
        idx = sel[0]
        if hasattr(self, '_saved_api_keys') and 1 <= idx < len(self._saved_api_keys):
            self._saved_api_keys[idx], self._saved_api_keys[idx - 1] = \
                self._saved_api_keys[idx - 1], self._saved_api_keys[idx]
            self._refresh_keys_listbox()
            self.keys_listbox.selection_set(idx - 1)

    def _move_key_down(self):
        """Move selected key down in priority"""
        sel = self.keys_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        n = len(getattr(self, '_saved_api_keys', []))
        if hasattr(self, '_saved_api_keys') and 0 <= idx < n - 1:
            self._saved_api_keys[idx], self._saved_api_keys[idx + 1] = \
                self._saved_api_keys[idx + 1], self._saved_api_keys[idx]
            self._refresh_keys_listbox()
            self.keys_listbox.selection_set(idx + 1)

    def _browse_service_account(self):
        """Browse and select a service account JSON key file"""
        from tkinter import filedialog
        file_path = filedialog.askopenfilename(
            title="Select Service Account JSON Key",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if file_path:
            self.sa_path_var.set(file_path)
            self.gemini_status_label.config(
                text=f"✅ Service account selected: {Path(file_path).name}", fg="green"
            )

    def _get_gemini_config_path(self):
        """Path to the Gemini config JSON file"""
        from config import DATA_DIR
        return DATA_DIR / "gemini_config.json"

    def _save_gemini_config(self):
        """Save Gemini API keys and/or service account path to config file"""
        import json
        api_keys = getattr(self, '_saved_api_keys', [])
        sa_path = self.sa_path_var.get().strip()
        if not api_keys and not sa_path:
            messagebox.showwarning("Input Required",
                                   "Please add at least one API key or select a Service Account JSON file")
            return
        config_path = self._get_gemini_config_path()
        try:
            config_path.parent.mkdir(parents=True, exist_ok=True)
            # Preserve existing CC settings (tone, voice model) if present
            config = {}
            if config_path.exists():
                with open(config_path, 'r') as f:
                    existing = json.load(f)
                for k in ('cc_tone', 'cc_voice_model'):
                    if k in existing:
                        config[k] = existing[k]
            config['gemini_api_keys'] = api_keys
            if sa_path:
                config['service_account_path'] = sa_path
            with open(config_path, 'w') as f:
                json.dump(config, f)
            count = len(api_keys)
            self.gemini_status_label.config(
                text=f"✅ Config saved ({count} key{'s' if count != 1 else ''})",
                fg="green"
            )
        except Exception as e:
            self.gemini_status_label.config(text=f"❌ Save failed: {e}", fg="red")

    def _load_gemini_config(self):
        """Load Gemini API keys and service account path from config file (backward compatible)"""
        import json
        config_path = self._get_gemini_config_path()
        if config_path.exists():
            try:
                with open(config_path, 'r') as f:
                    config = json.load(f)

                # Try loading as list first (new format)
                keys = config.get('gemini_api_keys', [])
                if not keys:
                    # Fallback: single key format (old format)
                    single_key = config.get('gemini_api_key', '')
                    if single_key:
                        keys = [single_key]

                sa_path = config.get('service_account_path', '')

                # Portable self-heal: the stored path may be an absolute path
                # from the machine where the config was first created. If it
                # does not exist here, fall back to the service-account key
                # bundled next to this install (data/service-account-key.json).
                try:
                    from config import DATA_DIR
                    bundled_sa = DATA_DIR / "service-account-key.json"
                    if (not sa_path or not os.path.isfile(sa_path)) and bundled_sa.is_file():
                        sa_path = str(bundled_sa)
                except Exception:
                    pass

                self._saved_api_keys = keys
                self._refresh_keys_listbox()

                if sa_path:
                    self.sa_path_var.set(sa_path)
                return keys
            except:
                pass
        return []

    def _test_all_keys(self):
        """Test all configured API keys individually"""
        from core.script_generator import ScriptGenerator

        api_keys = getattr(self, '_saved_api_keys', [])
        sa_path = self.sa_path_var.get().strip()

        if not api_keys and not sa_path:
            messagebox.showwarning("Input Required",
                                   "Add at least one API key or select a Service Account JSON file first.")
            return

        self.gemini_status_label.config(text="⏳ Testing keys...", fg="orange")
        self.root.update()

        results = []
        try:
            if sa_path and not api_keys:
                gen = ScriptGenerator(service_account_path=sa_path)
                ok, msg = gen.test_connection()
                results.append(f"SA: {'OK' if ok else 'FAIL'} — {msg}")
            else:
                gen = ScriptGenerator(api_keys=api_keys)
                for i, key in enumerate(api_keys):
                    # Configure with just this key for isolated testing
                    gen.configure(key)
                    ok, msg = gen.test_connection()
                    icon = "✓" if ok else "✗"
                    results.append(f"Key #{i+1}: {icon} {msg}")

            # Show summary
            ok_count = sum(1 for r in results if "✓" in r or "OK" in r)
            self.gemini_status_label.config(
                text=f"{'✅' if ok_count > 0 else '❌'} {ok_count}/{len(results)} keys OK",
                fg="green" if ok_count > 0 else "red"
            )
            # Show details in a message box if only 1-2 keys
            if len(results) <= 4:
                messagebox.showinfo("Key Test Results", "\n\n".join(results))
        except Exception as e:
            self.gemini_status_label.config(text=f"❌ Error: {str(e)}", fg="red")

    def create_video_script_tab(self):
        """Write-Story mode UI - Gemini watches the video and generates a story script.

        Built into self.video_script_tab, a sub-frame of the merged
        Script Studio tab (set up by create_script_studio_tab).
        """
        # Session cache: resolved video path → {file_uri, mime_type}. Lets a
        # re-run of the SAME video (e.g. a second language) skip the re-upload.
        self._vs_upload_cache = {}
        # Scrollable wrapper for this sub-tab
        _scr = ScrollableFrame(self.video_script_tab)
        _scr.pack(fill=tk.BOTH, expand=True)
        _SW = _scr.inner

        main_frame = tk.Frame(_SW)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # ── Left panel: inputs ────────────────────────────────
        left = tk.Frame(main_frame)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=False, padx=(0, 5))

        # Video source
        src_box = tk.LabelFrame(left, text="🎬 Video Source", font=("Arial", 10, "bold"),
                                padx=10, pady=10)
        src_box.pack(fill=tk.X, pady=(0, 8))

        tk.Label(src_box, text="YouTube URLs (one per line):").pack(anchor=tk.W)
        self.vs_url_text = tk.Text(src_box, height=4, width=50, wrap=tk.WORD,
                                    font=("Arial", 9))
        self.vs_url_text.pack(fill=tk.X, pady=(2, 5))

        tk.Label(src_box, text="— OR —").pack(anchor=tk.CENTER, pady=2)
        btn_row = tk.Frame(src_box)
        btn_row.pack(fill=tk.X, pady=(2, 0))
        self.vs_file_label = tk.Label(btn_row, text="No file selected", fg="gray",
                                      anchor=tk.W, width=35)
        self.vs_file_label.pack(side=tk.LEFT)
        tk.Button(btn_row, text="📁 Browse File",
                  command=self._vs_browse_file).pack(side=tk.RIGHT, padx=1)
        tk.Button(btn_row, text="📂 From Channels",
                  command=self._vs_browse_channel).pack(side=tk.RIGHT, padx=1)

        # Options
        opt_box = tk.LabelFrame(left, text="⚙️ Prompt Settings", font=("Arial", 10, "bold"),
                                padx=10, pady=10)
        opt_box.pack(fill=tk.X, pady=(0, 8))

        # Preset (which video-based prompt to use)
        tk.Label(opt_box, text="Preset:").pack(anchor=tk.W)
        self.vs_prompt_type_var = tk.StringVar(value="Dialogue-Only Clip to Story Script")
        self.vs_prompt_type_menu = ttk.Combobox(opt_box, textvariable=self.vs_prompt_type_var,
                                                 state="readonly", width=42)
        self.vs_prompt_type_menu.pack(fill=tk.X, pady=(2, 5))
        self.vs_prompt_type_menu.bind(
            "<<ComboboxSelected>>", lambda _e: self._update_vs_preset_help())

        # About this preset (details panel)
        vs_help_box = tk.LabelFrame(opt_box, text="📖 About this preset",
                                    font=("Arial", 9, "bold"), fg="#1565C0")
        vs_help_box.pack(fill=tk.X, pady=(2, 6))
        self.vs_help_label = tk.Label(
            vs_help_box, text="Select a preset to see what it does.",
            font=("Arial", 8), justify=tk.LEFT, anchor=tk.W,
            wraplength=300, fg="#333")
        self.vs_help_label.pack(fill=tk.X, padx=6, pady=6)

        tk.Button(opt_box, text="✏️ Manage Presets",
                  font=("Arial", 8), bg='#455A64', fg='white',
                  command=lambda: self._open_prompt_manager(
                      self.WRITE_STORY_MODE_SLUGS,
                      self._slug_for_name(self.vs_prompt_type_var.get(),
                                          self.WRITE_STORY_MODE_SLUGS),
                      prefer_master=True)
                  ).pack(anchor=tk.W, pady=(0, 4))

        # Language
        tk.Label(opt_box, text="Target Language:").pack(anchor=tk.W)
        self.vs_lang_var = tk.StringVar(value="english")
        vs_lang_menu = ttk.Combobox(opt_box, textvariable=self.vs_lang_var,
                                     values=["english", "russian", "arabic", "spanish",
                                             "french", "german", "hindi", "urdu",
                                             "chinese", "japanese"],
                                     state="readonly", width=30)
        vs_lang_menu.pack(fill=tk.X, pady=(2, 5))

        # Internal state: niche/style still consumed downstream but no UI
        # (the chosen Preset already defines niche and style).
        self.vs_niche_var = tk.StringVar(value="")
        self.vs_style_pref_var = tk.StringVar(value="")

        # ── TTS Engine selector ─────────────────────────────────
        tts_frame = tk.LabelFrame(left, text="🎤 TTS Engine",
                                  font=("Arial", 10, "bold"), padx=10, pady=5)
        tts_frame.pack(fill=tk.X, pady=(2, 6))
        self.vs_tts_engine_var = tk.StringVar(value="cloud")
        tk.Radiobutton(tts_frame, text="Cloud TTS (Standard — 160 WPM)",
                       variable=self.vs_tts_engine_var,
                       value="cloud", font=("Arial", 9)).pack(anchor=tk.W, padx=5, pady=1)
        tk.Radiobutton(tts_frame, text="Qwen3 TTS (Slower — ~110 WPM)",
                       variable=self.vs_tts_engine_var,
                       value="qwen3", font=("Arial", 9)).pack(anchor=tk.W, padx=5, pady=1)
        tk.Radiobutton(tts_frame, text="Gemini TTS (auto speed)",
                       variable=self.vs_tts_engine_var,
                       value="gemini", font=("Arial", 9)).pack(anchor=tk.W, padx=5, pady=1)

        # ── Voice Model + Tone/Style selectors (Gemini TTS calibration) ──
        cal_frame = tk.LabelFrame(left, text="🎙️ Voice & Tone (Gemini TTS)",
                                  font=("Arial", 10, "bold"), padx=10, pady=5)
        cal_frame.pack(fill=tk.X, pady=(2, 6))

        row1 = tk.Frame(cal_frame)
        row1.pack(fill=tk.X, padx=5, pady=2)
        tk.Label(row1, text="Voice Model:", font=("Arial", 9)).pack(side=tk.LEFT, padx=(0, 5))
        self.vs_voice_model_var = tk.StringVar(value="Zephyr")
        voice_models = sorted(_VOICE_WPM.keys())
        voice_menu = tk.OptionMenu(row1, self.vs_voice_model_var, *voice_models)
        voice_menu.config(font=("Arial", 9))
        voice_menu.pack(side=tk.LEFT, fill=tk.X, expand=True)

        row2 = tk.Frame(cal_frame)
        row2.pack(fill=tk.X, padx=5, pady=2)
        tk.Label(row2, text="Tone / Style:", font=("Arial", 9)).pack(side=tk.LEFT, padx=(0, 5))
        self.vs_tone_var = tk.StringVar(value="Storytelling")
        tone_names = sorted(_TONE_PACING.keys())
        tone_menu = tk.OptionMenu(row2, self.vs_tone_var, *tone_names)
        tone_menu.config(font=("Arial", 9))
        tone_menu.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Context / Backstory
        ctx_box = tk.LabelFrame(left, text="📖 Context / Backstory (optional)",
                                font=("Arial", 10, "bold"), padx=10, pady=10)
        ctx_box.pack(fill=tk.X, pady=(0, 8))
        tk.Label(ctx_box, text="2-3 sentences about what happened BEFORE the clip:",
                font=("Arial", 9), fg="#555").pack(anchor=tk.W)
        self.vs_context_text = tk.Text(ctx_box, height=3, width=50, wrap=tk.WORD,
                                       font=("Arial", 9))
        self.vs_context_text.pack(fill=tk.X, pady=(5, 0))

        # Optional transcript
        trans_box = tk.LabelFrame(left, text="📝 Optional Transcript",
                                  font=("Arial", 10, "bold"), padx=10, pady=10)
        trans_box.pack(fill=tk.X, pady=(0, 8))

        self.vs_use_transcript = tk.BooleanVar(value=False)
        tk.Checkbutton(trans_box, text="Include transcript text",
                       variable=self.vs_use_transcript,
                       command=self._vs_toggle_transcript).pack(anchor=tk.W)

        self.vs_transcript_text = tk.Text(trans_box, height=5, width=50,
                                          state=tk.DISABLED, bg="#f0f0f0")
        self.vs_transcript_text.pack(fill=tk.X, pady=(5, 0))

        # Generate button
        tk.Button(left, text="🎬 Generate Script from Video",
                  font=("Arial", 11, "bold"), bg="#4CAF50", fg="white",
                  command=self._vs_generate).pack(fill=tk.X, pady=(0, 5))
        self.vs_status_label = tk.Label(left, text="", fg="blue", wraplength=400)
        self.vs_status_label.pack(anchor=tk.W)

        # ── Batch Process Section ─────────────────────────────
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=6)

        batch_box = tk.LabelFrame(left, text="📦 Batch Process (Channel)",
                                   font=("Arial", 10, "bold"), padx=10, pady=8)
        batch_box.pack(fill=tk.X, pady=(0, 6))

        # Channel selection
        sel_row = tk.Frame(batch_box)
        sel_row.pack(fill=tk.X, pady=(0, 4))
        tk.Button(sel_row, text="📁 Select Channel",
                  command=self._vs_select_channel,
                  bg="#607D8B", fg="white", font=("Arial", 9)).pack(side=tk.LEFT)
        self._batch_info_label = tk.Label(sel_row, text="No channel selected",
                                          fg="gray", font=("Arial", 9))
        self._batch_info_label.pack(side=tk.LEFT, padx=8)

        # Video list (text widget with status colors)
        self._batch_list_text = tk.Text(batch_box, height=5, width=50,
                                         state=tk.DISABLED, font=("Consolas", 9),
                                         bg="#fafafa", wrap=tk.NONE)
        self._batch_list_text.pack(fill=tk.X, pady=(0, 4))
        self._batch_list_text.tag_configure("waiting", foreground="gray")
        self._batch_list_text.tag_configure("processing", foreground="#1565C0")
        self._batch_list_text.tag_configure("done", foreground="#2E7D32")
        self._batch_list_text.tag_configure("error", foreground="red")

        # Buttons
        bbtn_row = tk.Frame(batch_box)
        bbtn_row.pack(fill=tk.X)
        self._batch_gen_btn = tk.Button(
            bbtn_row, text="▶ Generate All", state=tk.DISABLED,
            font=("Arial", 9, "bold"), bg="#2196F3", fg="white",
            command=self._vs_generate_all)
        self._batch_gen_btn.pack(side=tk.LEFT, padx=2)
        self._batch_save_btn = tk.Button(
            bbtn_row, text="💾 Save All to Excel", state=tk.DISABLED,
            font=("Arial", 9, "bold"), bg="#4CAF50", fg="white",
            command=self._vs_save_batch_excel)
        self._batch_save_btn.pack(side=tk.LEFT, padx=2)
        self._batch_prog_label = tk.Label(bbtn_row, text="", fg="blue",
                                          font=("Arial", 9))
        self._batch_prog_label.pack(side=tk.RIGHT, padx=4)

        # Batch state
        self._batch_videos = []      # {path, name, status, result}
        self._batch_channel_path = None

        # ── Right panel: output ───────────────────────────────
        right = tk.Frame(main_frame)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))

        out_box = tk.LabelFrame(right, text="📜 Generated Script",
                                font=("Arial", 10, "bold"), padx=10, pady=10)
        out_box.pack(fill=tk.BOTH, expand=True)

        self.vs_output_text = tk.Text(out_box, height=25, width=60, wrap=tk.WORD,
                                       font=("Arial", 10))
        self.vs_output_text.pack(fill=tk.BOTH, expand=True)

        # Scrollbar
        scroll = tk.Scrollbar(self.vs_output_text)
        self.vs_output_text.config(yscrollcommand=scroll.set)
        scroll.config(command=self.vs_output_text.yview)

        # Action buttons below output
        btn_frame = tk.Frame(right)
        btn_frame.pack(fill=tk.X, pady=(5, 0))
        tk.Button(btn_frame, text="Copy to Clipboard",
                  command=self._vs_copy_output).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_frame, text="Save to Excel",
                  command=self._vs_save_excel).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_frame, text="Clear",
                  command=lambda: self.vs_output_text.delete("1.0", tk.END)).pack(side=tk.LEFT, padx=2)

        # Internal state
        self._vs_file_path = None
        self._vs_video_url = None
        self._vs_last_result = None
        self._vs_all_results = []  # list of {url, result} dicts for multi-URL exports

        # Populate prompt type dropdown from ScriptGenerator prompts
        self._vs_populate_prompt_types()

    def _vs_browse_file(self):
        """Browse and select local video file(s) — supports multi-select."""
        from tkinter import filedialog
        from pathlib import Path
        paths = filedialog.askopenfilenames(
            title="Select Video File(s) (Ctrl+click for multiple)",
            filetypes=[("Video files", "*.mp4 *.mov *.avi *.webm *.mkv *.m4v"),
                       ("All files", "*.*")]
        )
        if not paths:
            return

        # Single file → normal single-file mode
        if len(paths) == 1:
            self._vs_file_path = paths[0]
            self._vs_video_url = None
            self.vs_url_text.delete("1.0", tk.END)
            self.vs_file_label.config(
                text=Path(paths[0]).name, fg="black")
            return

        # Multiple files → fill batch list (for Excel export)
        from pathlib import Path as _P
        first_parent = _P(paths[0]).parent
        self._batch_channel_path = first_parent
        self._batch_videos = [
            {"path": p, "name": _P(p).name, "status": "waiting", "result": None}
            for p in paths
        ]
        folder_name = first_parent.name or "Selected"
        self._batch_info_label.config(
            text=f"📁 {folder_name} ({len(self._batch_videos)} videos)", fg="black")
        self._vs_refresh_batch_list()
        self._batch_gen_btn.config(state=tk.NORMAL)
        self._batch_save_btn.config(state=tk.DISABLED)
        self._batch_prog_label.config(
            text=f"Ready • 0/{len(self._batch_videos)}", fg="blue")
        self.vs_status_label.config(
            text=f"✅ {len(self._batch_videos)} videos loaded for batch processing",
            fg="green")

    def _vs_toggle_transcript(self):
        """Enable/disable the transcript text area."""
        if self.vs_use_transcript.get():
            self.vs_transcript_text.config(state=tk.NORMAL, bg="white")
        else:
            self.vs_transcript_text.config(state=tk.DISABLED, bg="#f0f0f0")

    def _vs_populate_prompt_types(self):
        """Fill the Write-Story preset dropdown — only video-based prompts.

        Transcript-only and Case-Commentary prompts are excluded so the list
        only shows presets that actually work by watching the video.
        """
        try:
            from core.script_generator import ScriptGenerator
            gen = ScriptGenerator(api_keys=[])
            all_prompts = gen.get_prompt_list()  # [(slug, display_name, desc), ...]
            # Keep only WRITE_STORY_MODE_SLUGS, preserving that order
            by_slug = {slug: dname for slug, dname, _ in all_prompts}
            display_names = [by_slug[s] for s in self.WRITE_STORY_MODE_SLUGS if s in by_slug]
            if display_names:
                self.vs_prompt_type_menu['values'] = display_names
                self.vs_prompt_type_var.set(display_names[0])
                self._update_vs_preset_help()
        except Exception:
            pass

    def _update_vs_preset_help(self):
        """Show the help text for the selected Write-Story-mode preset."""
        if not hasattr(self, "vs_help_label"):
            return
        text = self._preset_help_for_name(
            self.vs_prompt_type_var.get(), self.WRITE_STORY_MODE_SLUGS)
        self.vs_help_label.config(text=text)

    def _vs_browse_channel(self):
        """Browse the channels folder — supports multi-select (Ctrl+click)."""
        from tkinter import filedialog
        from pathlib import Path

        channels_dir = Path(__file__).parent.parent / "channels"
        channels_dir.mkdir(parents=True, exist_ok=True)

        paths = filedialog.askopenfilenames(
            title="Select Video(s) from Channels (Ctrl+click for multiple)",
            initialdir=str(channels_dir),
            filetypes=[("Video files", "*.mp4 *.mov *.avi *.webm *.mkv *.m4v"),
                       ("All files", "*.*")]
        )
        if not paths:
            return

        # Single file → normal single-file mode
        if len(paths) == 1:
            path = paths[0]
            self._vs_file_path = path
            self._vs_video_url = None
            self.vs_url_text.delete("1.0", tk.END)
            try:
                rel = Path(path).relative_to(channels_dir)
                label = f"📂 {rel}"
            except ValueError:
                label = Path(path).name
            self.vs_file_label.config(text=label, fg="green")
            return

        # Multiple files → fill batch list
        first_parent = Path(paths[0]).parent
        self._batch_channel_path = first_parent
        self._batch_videos = [
            {"path": str(p), "name": Path(p).name, "status": "waiting", "result": None}
            for p in paths
        ]
        try:
            rel = first_parent.relative_to(channels_dir)
            label = str(rel)
        except ValueError:
            label = first_parent.name
        self._batch_info_label.config(
            text=f"📂 {label} ({len(self._batch_videos)} videos)", fg="black")
        self._vs_refresh_batch_list()
        self._batch_gen_btn.config(state=tk.NORMAL)
        self._batch_save_btn.config(state=tk.DISABLED)
        self._batch_prog_label.config(
            text=f"Ready • 0/{len(self._batch_videos)}", fg="blue")
        self.vs_status_label.config(
            text=f"✅ {len(self._batch_videos)} videos loaded for batch processing",
            fg="green")

    # ── Batch processing methods ─────────────────────────────

    def _vs_select_channel(self):
        """Select a channel folder and scan for video files."""
        from tkinter import filedialog
        from pathlib import Path

        base_dir = Path(__file__).parent.parent / "channels"
        base_dir.mkdir(parents=True, exist_ok=True)

        folder = filedialog.askdirectory(
            title="Select Channel Folder (e.g. youtube/ChannelName)",
            initialdir=str(base_dir),
        )
        if not folder:
            return

        folder = Path(folder)

        # Look for videos/ subfolder
        if folder.name == "videos":
            videos_dir = folder
            channel_root = folder.parent
        else:
            videos_dir = folder / "videos"
            channel_root = folder

        if not videos_dir.is_dir():
            self.vs_status_label.config(
                text="❌ No 'videos/' folder found in this channel.", fg="red")
            return

        # Scan for video files
        video_files = []
        for ext in ("*.mp4", "*.mov", "*.avi", "*.webm", "*.mkv", "*.m4v"):
            video_files.extend(sorted(videos_dir.glob(ext)))

        if not video_files:
            self.vs_status_label.config(
                text="❌ No video files found in channel.", fg="red")
            return

        self._batch_channel_path = channel_root
        self._batch_videos = [
            {"path": str(v), "name": v.name, "status": "waiting", "result": None}
            for v in video_files
        ]

        # Show relative path
        try:
            rel = channel_root.relative_to(base_dir.parent)
            label = str(rel)
        except ValueError:
            label = channel_root.name
        self._batch_info_label.config(
            text=f"📂 {label} ({len(self._batch_videos)} videos)", fg="black",
        )
        self._vs_refresh_batch_list()
        self._batch_gen_btn.config(state=tk.NORMAL)
        self._batch_save_btn.config(state=tk.DISABLED)
        self._batch_prog_label.config(
            text=f"Ready • 0/{len(self._batch_videos)}", fg="blue")
        self.vs_status_label.config(
            text=f"✅ Channel loaded: {len(self._batch_videos)} videos", fg="green")

    def _vs_refresh_batch_list(self):
        """Rebuild the batch video list Text widget from _batch_videos."""
        w = self._batch_list_text
        w.config(state=tk.NORMAL)
        w.delete("1.0", tk.END)
        icons = {"waiting": "⏳", "processing": "🔄", "done": "✅", "error": "❌"}
        for v in self._batch_videos:
            icon = icons.get(v["status"], "⏳")
            w.insert(tk.END, f"{icon} {v['name']}\n", v["status"])
        w.config(state=tk.DISABLED)

    def _vs_generate_all(self):
        """Batch-process every video in the selected channel."""
        if not self._batch_videos:
            return

        # Reset state
        for v in self._batch_videos:
            v["status"] = "waiting"
            v["result"] = None
        self._vs_refresh_batch_list()
        self._batch_gen_btn.config(state=tk.DISABLED)
        self._batch_save_btn.config(state=tk.DISABLED)

        lang = self.vs_lang_var.get().strip()
        niche = self.vs_niche_var.get().strip()
        prompt_type_name = self.vs_prompt_type_var.get().strip()
        style_pref = self.vs_style_pref_var.get().strip()
        context = self.vs_context_text.get("1.0", tk.END).strip()

        def _task():
            try:
                from core.script_generator import ScriptGenerator

                api_keys = getattr(self, '_saved_api_keys', [])
                sa_path = self.sa_path_var.get().strip()

                if not api_keys and not sa_path:
                    self.root.after(0, lambda: self._batch_prog_label.config(
                        text="❌ No API keys configured", fg="red"))
                    self.root.after(0, lambda: self._batch_gen_btn.config(
                        state=tk.NORMAL))
                    return

                if sa_path and not api_keys:
                    gen = ScriptGenerator(service_account_path=sa_path)
                elif api_keys:
                    gen = ScriptGenerator(api_keys=api_keys)
                else:
                    return

                # Set active prompt from dropdown
                slug = None
                for s, dname, _ in gen.get_prompt_list():
                    if dname == prompt_type_name:
                        slug = s
                        break
                if slug and gen.get_prompt_data(slug):
                    gen._active_prompt_key = slug  # in-memory only — don't save to JSON

                total = len(self._batch_videos)
                done = 0
                errors = 0

                for i, v in enumerate(self._batch_videos):
                    v["status"] = "processing"
                    self.root.after(0, lambda: self._vs_refresh_batch_list())
                    self.root.after(0, lambda i=i: self._batch_prog_label.config(
                        text=f"⏳ Processing {i+1}/{total}...", fg="blue"))

                    try:
                        result = gen.generate_script_from_video(
                            video_path=v["path"],
                            language=lang,
                            niche_angle=niche,
                            style_preference=style_pref,
                            context=context,
                            upload_cache=self._vs_upload_cache,
                        )
                        v["result"] = result
                        if "error" in result:
                            v["status"] = "error"
                            errors += 1
                        else:
                            v["status"] = "done"
                            done += 1
                    except Exception as e:
                        v["status"] = "error"
                        v["result"] = {"error": str(e)}
                        errors += 1

                    self.root.after(0, lambda: self._vs_refresh_batch_list())

                # All done
                self.root.after(0, lambda d=done, e=errors: self._batch_prog_label.config(
                    text=f"✅ {d} success, {e} error{'s' if e != 1 else ''}",
                    fg="green" if errors == 0 else "#E65100"))
                self.root.after(0, lambda: self._batch_save_btn.config(
                    state=tk.NORMAL))

            except Exception as e:
                self.root.after(0, lambda: self._batch_prog_label.config(
                    text=f"❌ Batch error: {str(e)}", fg="red"))
                self.root.after(0, lambda: self._batch_gen_btn.config(
                    state=tk.NORMAL))

        import threading
        threading.Thread(target=_task, daemon=True).start()

    def _vs_save_batch_excel(self):
        """Save all batch results to a single Excel file."""
        from tkinter import filedialog, messagebox
        from datetime import datetime
        from pathlib import Path

        results = [v for v in self._batch_videos
                   if v["result"] and "error" not in v["result"]]
        if not results:
            messagebox.showwarning("No Results",
                                   "No successful results to save.")
            return

        default_name = "batch_scripts"
        if self._batch_channel_path:
            default_name = f"{self._batch_channel_path.name}_scripts"
        default_name += f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        path = filedialog.asksaveasfilename(
            title="Save Batch Scripts to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=str(Path(__file__).parent.parent / "channels"),
            initialfile=default_name,
        )
        if not path:
            return

        try:
            import pandas as pd
            rows = []
            for v in results:
                r = v["result"]
                video_name = v["name"]
                rows.append({
                    "Video ID": _extract_video_id(video_name),
                    "Title": r.get("suggested_title", r.get("title", "")),
                    "Video File": video_name,
                    "Language": self.vs_lang_var.get().strip(),
                    "Niche": self.vs_niche_var.get().strip(),
                    "Prompt Type": self.vs_prompt_type_var.get().strip(),
                    "Custom Title": r.get("suggested_title", ""),
                    # Clean the raw Gemini output down to narration-only before
                    # saving — the batch generate path stores the raw result, so
                    # without this the Dialogue-Only preset's deliverables
                    # package (analysis report, sync checklist, effects notes)
                    # would leak into Custom Script and Tool 2 would read it aloud.
                    "Custom Script": self._clean_narration_script(r.get("script", "")),
                    "Word Count": r.get("generated_word_count", 0),
                    "Hashtag 1": r.get("hashtag_1", ""),
                    "Hashtag 2": r.get("hashtag_2", ""),
                    "Voiceover Style": r.get("voiceover_style", ""),
                    "Voiceover Speed (WPM)": r.get("voiceover_speed", ""),
                    "Date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                })
            # Append/merge with existing file (skip duplicate Video IDs)
            existing_path = Path(path)
            if existing_path.exists():
                try:
                    old_df = pd.read_excel(path)
                    old_rows = old_df.to_dict("records")
                    existing_ids = {str(r.get("Video ID", "")) for r in old_rows if r.get("Video ID")}
                    deduped_new = [r for r in rows if str(r.get("Video ID", "")) not in existing_ids]
                    rows = old_rows + deduped_new
                except Exception:
                    pass  # If read fails, just save new rows
            df = pd.DataFrame(rows)
            df.to_excel(path, index=False)
            messagebox.showinfo("Saved",
                                f"✅ {len(rows)} scripts saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {str(e)}")

    def _vs_generate(self):
        """Generate scripts from YouTube URL(s) or a video file using Gemini."""
        import threading

        # Get all URLs (one per line)
        raw_urls = self.vs_url_text.get("1.0", tk.END).strip()
        urls = [u.strip() for u in raw_urls.split("\n") if u.strip()]
        file_path = self._vs_file_path
        lang = self.vs_lang_var.get().strip()
        niche = self.vs_niche_var.get().strip()
        transcript = self.vs_transcript_text.get("1.0", tk.END).strip() if self.vs_use_transcript.get() else ""
        prompt_type_name = self.vs_prompt_type_var.get().strip()
        style_pref = self.vs_style_pref_var.get().strip()
        context = self.vs_context_text.get("1.0", tk.END).strip()

        use_file = bool(file_path)
        use_urls = bool(urls)

        if not use_urls and not use_file:
            self.vs_status_label.config(text="❌ Enter at least one YouTube URL or select a video file.", fg="red")
            return

        if use_file:
            # Single file mode
            self.vs_status_label.config(text="⏳ Generating script from file...", fg="blue")
            self.vs_output_text.delete("1.0", tk.END)
            self.vs_last_result = None

            # Single file mode — clear multi-URL results
            self._vs_all_results = []
            def single_task():
                try:
                    from core.script_generator import ScriptGenerator
                    api_keys = getattr(self, '_saved_api_keys', [])
                    sa_path = self.sa_path_var.get().strip()
                    if not api_keys and not sa_path:
                        self.root.after(0, lambda: self.vs_status_label.config(
                            text="❌ Gemini API key not configured. Go to Settings.", fg="red"))
                        return
                    if sa_path and not api_keys:
                        gen = ScriptGenerator(service_account_path=sa_path)
                    elif api_keys:
                        gen = ScriptGenerator(api_keys=api_keys)
                    else:
                        self.root.after(0, lambda: self.vs_status_label.config(
                            text="❌ No API key or service account configured.", fg="red"))
                        return
                    slug = None
                    for s, dname, _ in gen.get_prompt_list():
                        if dname == prompt_type_name:
                            slug = s
                            break
                    if slug and gen.get_prompt_data(slug):
                        gen._active_prompt_key = slug
                    # Calculate effective WPM based on TTS engine selection
                    _vs_engine = getattr(self, 'vs_tts_engine_var', None)
                    if _vs_engine is not None and _vs_engine.get() == "qwen3":
                        wpm_val = 110
                    elif _vs_engine is not None and _vs_engine.get() == "gemini":
                        # Use calibration tables with selected voice & tone
                        _vm = getattr(self, 'vs_voice_model_var', None)
                        _ts = getattr(self, 'vs_tone_var', None)
                        vm = _vm.get() if _vm else 'Zephyr'
                        ts = _ts.get() if _ts else 'Storytelling'
                        wpm_at_1x = _VOICE_WPM.get(vm, 105)
                        tone_factor = _TONE_PACING.get(ts, 0.85)
                        wpm_val = round(wpm_at_1x * _BASE_TTS_SPEED * tone_factor)
                    else:
                        wpm_val = 165
                    result = gen.generate_script_from_video(
                        video_url=None,
                        video_path=file_path,
                        language=lang,
                        niche_angle=niche,
                        style_preference=style_pref,
                        context=context,
                        transcript=transcript if transcript else None,
                        wpm=wpm_val,
                        progress_callback=lambda msg: self.root.after(0, lambda m=msg: self.vs_status_label.config(text=m, fg="blue")),
                        upload_cache=self._vs_upload_cache,
                    )
                    self.root.after(0, lambda: self._vs_handle_result(result))
                except Exception as e:
                    self.root.after(0, lambda: self.vs_status_label.config(
                        text=f"❌ Error: {str(e)}", fg="red"))
            threading.Thread(target=single_task, daemon=True).start()
            return

        # Multiple URLs mode
        self.vs_output_text.delete("1.0", tk.END)
        self.vs_last_result = None

        def multi_task():
            try:
                from core.script_generator import ScriptGenerator
                api_keys = getattr(self, '_saved_api_keys', [])
                sa_path = self.sa_path_var.get().strip()
                if not api_keys and not sa_path:
                    self.root.after(0, lambda: self.vs_status_label.config(
                        text="❌ Gemini API key not configured. Go to Settings.", fg="red"))
                    return
                if sa_path and not api_keys:
                    gen = ScriptGenerator(service_account_path=sa_path)
                elif api_keys:
                    gen = ScriptGenerator(api_keys=api_keys)
                else:
                    self.root.after(0, lambda: self.vs_status_label.config(
                        text="❌ No API key or service account configured.", fg="red"))
                    return

                slug = None
                for s, dname, _ in gen.get_prompt_list():
                    if dname == prompt_type_name:
                        slug = s
                        break
                if slug and gen.get_prompt_data(slug):
                    gen._active_prompt_key = slug

                all_results = []
                total = len(urls)
                for i, url in enumerate(urls):
                    self.root.after(0, lambda i=i, url=url: (
                        self.vs_status_label.config(
                            text=f"⏳ Processing URL {i+1}/{total}...", fg="blue"),
                    ))
                    self._append_output(
                        f"\n{'='*60}\n📌 URL {i+1}/{total}: {url}\n{'='*60}\n")

                    # Calculate effective WPM based on TTS engine selection
                    _vs_engine = getattr(self, 'vs_tts_engine_var', None)
                    if _vs_engine is not None and _vs_engine.get() == "qwen3":
                        wpm_val = 110
                    elif _vs_engine is not None and _vs_engine.get() == "gemini":
                        # Use calibration tables with selected voice & tone
                        _vm = getattr(self, 'vs_voice_model_var', None)
                        _ts = getattr(self, 'vs_tone_var', None)
                        vm = _vm.get() if _vm else 'Zephyr'
                        ts = _ts.get() if _ts else 'Storytelling'
                        wpm_at_1x = _VOICE_WPM.get(vm, 105)
                        tone_factor = _TONE_PACING.get(ts, 0.85)
                        wpm_val = round(wpm_at_1x * _BASE_TTS_SPEED * tone_factor)
                    else:
                        wpm_val = 165
                    result = gen.generate_script_from_video(
                        video_url=url,
                        video_path=None,
                        language=lang,
                        niche_angle=niche,
                        style_preference=style_pref,
                        context=context,
                        transcript=transcript if transcript else None,
                        wpm=wpm_val,
                        progress_callback=lambda msg: self.root.after(0, lambda m=msg: self.vs_status_label.config(text=m, fg="blue")),
                        upload_cache=self._vs_upload_cache,
                    )
                    all_results.append({"url": url, "result": result})

                    # Clean the script for display and Excel export
                    if result and "script" in result:
                        result["script"] = self._clean_narration_script(result["script"])

                    if result:
                        text = result.get("script_text", "") or result.get("script", "") or str(result)
                    else:
                        text = "❌ No result returned."

                    self.root.after(0, lambda t=text: self._append_output(t + "\n\n"))

                # Store ALL results for Excel export
                self._vs_all_results = all_results
                self.vs_last_result = all_results[-1]["result"] if all_results else None
                self.root.after(0, lambda: self.vs_status_label.config(
                    text=f"✅ Done! Processed {total} URL(s).", fg="green"))
            except Exception as e:
                import traceback
                traceback.print_exc()
                self.root.after(0, lambda: self.vs_status_label.config(
                    text=f"❌ Error: {str(e)}", fg="red"))

        threading.Thread(target=multi_task, daemon=True).start()

    def _append_output(self, text):
        """Append text to the output widget safely from any thread."""
        try:
            self.vs_output_text.insert(tk.END, text)
            self.vs_output_text.see(tk.END)
        except Exception:
            pass

    @staticmethod
    def _clean_narration_script(raw_script: str) -> str:
        """Strip analysis reports, deliverables packages, and other extra content
        from Gemini output — keep only the clean narration script with timestamps.

        Handles three Gemini output shapes:
          1. Pipe format (courtroom / movie recap):
             ``[00:00] | text | (words) | [beat]`` — kept as-is.
          2. Dash / "visual sync checklist" format (the Dialogue-Only preset):
             ``[00:00] — "spoken text" syncs with <english note>`` — the spoken
             text is pulled out of the quotes and the "syncs with…" English note
             is dropped, then rebuilt into clean ``[00:00] | text`` pipe lines.
             Without this, the whole deliverables package (SCENE ANALYSIS
             REPORT, WORD COUNT VERIFICATION, PRODUCTION EFFECTS NOTES, etc.)
             leaked into the Custom Script column and Tool 2 read it all aloud —
             timestamps, English descriptions and all.
          3. No timestamps at all — returned as-is.
        """
        if not raw_script:
            return ""
        lines = raw_script.split("\n")

        # Pattern: optional bracket, timestamp, optional bracket, pipe
        ts_pattern = r"^\s*(?:\[)?\d{1,2}:\d{2}(?:\])?\s*\|"

        first_ts = None
        last_ts = None
        for i, line in enumerate(lines):
            stripped = line.strip()
            if re.match(ts_pattern, stripped):
                if first_ts is None:
                    first_ts = i
                last_ts = i

        if first_ts is not None:
            # ── Pipe format found — keep only the real narration lines ──────
            # Several movie/CCTV presets interleave progress / confirmation
            # lines BETWEEN the narration ones:
            #   // Running total: 120 of 270 words //
            #   // Watched up to 03:15 — next segment: 03:20 //
            #   [SYNC: matches visual at this timestamp — door opens]
            # and end with a "PHASE 6 DELIVERABLES" report. The [first_ts:
            # last_ts] slice drops the trailing report, but these inline lines
            # would otherwise leak into Custom Script and get read aloud by
            # Tool 2. Drop anything that isn't an actual "[MM:SS] | …" line
            # (plus markdown headings / separators).
            narration = lines[first_ts:last_ts + 1]
            clean = []
            for line in narration:
                stripped = line.strip()
                if not stripped:
                    clean.append(line)        # keep blank spacers
                    continue
                # Drop markdown headings and separator lines
                if stripped.startswith("###") or stripped.startswith("---"):
                    continue
                # Drop // running total // and // watched up to … // markers
                if stripped.startswith("//") or stripped.startswith("＃"):
                    continue
                # Drop standalone [SYNC: …] / [NOTE: …] production notes
                if re.match(r"^\[(?:SYNC|NOTE|BEAT)\b", stripped, re.IGNORECASE):
                    continue
                # Keep only genuine timestamped narration lines; anything else
                # inside the block (stray prose, labels) is noise.
                if re.match(ts_pattern, stripped):
                    clean.append(line)
            return "\n".join(clean).strip()

        # ── No pipe timestamps — try the dash / sync-checklist format ───────
        # Only lines whose content STARTS with a double-quote right after the
        # dash are narration; effects notes ("[00:08] — Zoom in: …") have no
        # leading quote and are correctly skipped. The inner dialogue in the
        # narration uses single quotes, so the first closing double-quote is
        # the true end of the spoken text.
        dash_pattern = re.compile(
            r'^\s*\[?(\d{1,2}:\d{2})\]?\s*[—–\-]\s*'   # [MM:SS] — / - / –
            r'["“”]([^"“”]+)["“”]'                        # "spoken text"
        )
        dash_lines = []
        for line in lines:
            m = dash_pattern.match(line.strip())
            if m:
                ts = m.group(1)
                spoken = m.group(2).strip()
                if spoken:
                    dash_lines.append(f"[{ts}] | {spoken}")

        if dash_lines:
            return "\n".join(dash_lines).strip()

        # No recognisable timestamps at all — return as-is (some courtroom
        # scripts are plain prose without timestamps).
        return raw_script.strip()

    def _vs_handle_result(self, result):
        """Handle the video script generation result."""
        if "error" in result:
            self.vs_status_label.config(text=f"❌ {result['error']}", fg="red")
            return

        script = self._clean_narration_script(result.get("script", ""))
        title = result.get("suggested_title", "")
        h1 = result.get("hashtag_1", "")
        h2 = result.get("hashtag_2", "")
        wc = result.get("generated_word_count", 0)

        # Store the cleaned script back so Excel export also gets it clean
        result["script"] = script

        output = []
        if title:
            output.append(f"Suggested Title: {title}")
            output.append("")
        output.append(script)
        if h1 or h2:
            output.append("")
            output.append(f"#{h1} #{h2}")

        self.vs_output_text.delete("1.0", tk.END)
        self.vs_output_text.insert("1.0", "\n".join(output))
        self.vs_last_result = result

        parts = []
        if title:
            parts.append(f'"{title}"')
        parts.append(f"{wc} words")
        self.vs_status_label.config(
            text=f"✅ Script generated! ({', '.join(parts)})", fg="green")

    def _vs_copy_output(self):
        """Copy the generated script to clipboard."""
        text = self.vs_output_text.get("1.0", tk.END).strip()
        if text:
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            self.vs_status_label.config(text="📋 Copied to clipboard!", fg="green")

    def _vs_save_excel(self):
        """Save the generated script(s) to an Excel file."""
        from tkinter import filedialog, messagebox
        from datetime import datetime
        from pathlib import Path

        # Use multi-URL results if available, otherwise fall back to single result
        rows = []
        if self._vs_all_results:
            for entry in self._vs_all_results:
                r = entry.get("result") or {}
                url = entry.get("url", "")
                rows.append({
                    "Video ID": _extract_yt_id(url) or (_extract_video_id(Path(url).name) if url else ""),
                    "Title": r.get("suggested_title", r.get("title", "")),
                    "Video URL": url,
                    "Language": self.vs_lang_var.get().strip(),
                    "Niche": self.vs_niche_var.get().strip(),
                    "Custom Title": r.get("suggested_title", ""),
                    "Custom Script": r.get("script", ""),
                    "Word Count": r.get("generated_word_count", 0),
                    "Hashtag 1": r.get("hashtag_1", ""),
                    "Hashtag 2": r.get("hashtag_2", ""),
                    "Voiceover Style": r.get("voiceover_style", ""),
                    "Voiceover Speed (WPM)": r.get("voiceover_speed", ""),
                    "Date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                })
        elif self.vs_last_result:
            r = self.vs_last_result
            first_url = (self.vs_url_text.get("1.0", tk.END).strip().split("\n")[0]
                         if self.vs_url_text.get("1.0", tk.END).strip() else "")
            if not first_url and self._vs_file_path:
                first_url = self._vs_file_path
            # For local files, extract video ID from bracket notation; for URLs use _extract_yt_id
            vid_id = _extract_yt_id(first_url)
            if not vid_id and first_url:
                vid_id = _extract_video_id(Path(first_url).name)
            vid_id = vid_id or ""
            rows.append({
                "Video ID": vid_id,
                "Title": r.get("suggested_title", r.get("title", "")),
                "Video URL": first_url,
                "Language": self.vs_lang_var.get().strip(),
                "Niche": self.vs_niche_var.get().strip(),
                "Custom Title": r.get("suggested_title", ""),
                "Custom Script": r.get("script", ""),
                "Word Count": r.get("generated_word_count", 0),
                "Hashtag 1": r.get("hashtag_1", ""),
                "Hashtag 2": r.get("hashtag_2", ""),
                "Voiceover Style": r.get("voiceover_style", ""),
                "Voiceover Speed (WPM)": r.get("voiceover_speed", ""),
                "Date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
        else:
            self.vs_status_label.config(text="❌ Nothing to save — generate a script first.", fg="red")
            return

        # Build a default filename matching the Case Commentary style:
        # <niche-slug>_<video-id>_<language>.xlsx  (falls back gracefully).
        prompt_type_name = self.vs_prompt_type_var.get().strip()
        _slug = None
        try:
            from core.script_generator import ScriptGenerator
            for s, dname, _ in ScriptGenerator(api_keys=[]).get_prompt_list():
                if dname == prompt_type_name:
                    _slug = s
                    break
        except Exception:
            pass
        _niche_part = _slug or (self.vs_niche_var.get().strip() or "VideoScript")
        _lang = self.vs_lang_var.get().strip() or "lang"
        _ids = [str(r.get("Video ID", "")).strip() for r in rows if str(r.get("Video ID", "")).strip()]
        _vid_part = _ids[0] if len(_ids) == 1 else ("batch" if _ids else "unknown")
        # Sanitize for a filesystem-safe name.
        def _safe(s):
            return re.sub(r'[^A-Za-z0-9._-]+', '_', str(s)).strip('_') or "x"
        default_name = f"{_safe(_niche_part)}_{_safe(_vid_part)}_{_safe(_lang)}.xlsx"

        path = filedialog.asksaveasfilename(
            title="Save Script to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=str(Path(__file__).parent.parent / "channels"),
            initialfile=default_name
        )
        if not path:
            return

        try:
            import pandas as pd
            # Append/merge with existing file (skip duplicate Video IDs)
            existing_path = Path(path)
            if existing_path.exists():
                try:
                    old_df = pd.read_excel(path)
                    old_rows = old_df.to_dict("records")
                    existing_ids = {str(r.get("Video ID", "")) for r in old_rows if r.get("Video ID")}
                    deduped_new = [r for r in rows if str(r.get("Video ID", "")) not in existing_ids]
                    rows = old_rows + deduped_new
                except Exception:
                    pass
            df = pd.DataFrame(rows)
            df.to_excel(path, index=False)
            messagebox.showinfo("Saved", f"{len(rows)} script(s) saved to:\n{path}")
            self.vs_status_label.config(text=f"✅ {len(rows)} script(s) saved to: {path}", fg="green")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {str(e)}")

    # ── Case Commentary Tab ──────────────────────────────────────

    def create_case_commentary_tab(self):
        """Case Commentary tab — Gemini watches a courtroom video and outputs
        a summary, montage clip timestamps, and commentary spots.
        """
        self._cc_upload_cache = {}  # session cache: video_path → {file_uri, mime_type}
        main_frame = tk.Frame(self._cc_scroller.inner)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # ── Left panel: inputs ──────────────────────────────
        left = tk.Frame(main_frame)
        left.pack(side=tk.LEFT, fill=tk.Y, expand=False, padx=(0, 5))

        # URL input
        src_box = tk.LabelFrame(left, text="🎬 YouTube Video", font=("Arial", 10, "bold"),
                                padx=10, pady=10)
        src_box.pack(fill=tk.X, pady=(0, 8))

        tk.Label(src_box, text="YouTube URL:").pack(anchor=tk.W)
        self.cc_url_text = tk.Text(src_box, height=3, width=50, wrap=tk.WORD,
                                   font=("Arial", 9))
        self.cc_url_text.pack(fill=tk.X, pady=(2, 5))

        # Optional: file browse + channel browse
        btn_row = tk.Frame(src_box)
        btn_row.pack(fill=tk.X, pady=(2, 0))
        self.cc_file_label = tk.Label(btn_row, text="No file selected", fg="gray",
                                      anchor=tk.W, width=35)
        self.cc_file_label.pack(side=tk.LEFT)
        tk.Button(btn_row, text="📁 Browse File",
                  command=self._cc_browse_file).pack(side=tk.RIGHT, padx=1)
        tk.Button(btn_row, text="📂 From Channels",
                  command=self._cc_browse_channel).pack(side=tk.RIGHT, padx=1)

        # Context / Backstory (optional)
        ctx_box = tk.LabelFrame(left, text="📖 Context / Backstory (optional)",
                                font=("Arial", 10, "bold"), padx=10, pady=8)
        ctx_box.pack(fill=tk.X, pady=(0, 6))
        tk.Label(ctx_box,
                 text="Describe what happened before/after the clip, location details, or any background info to help Gemini understand the footage.",
                 font=("Arial", 8), fg="#666", justify=tk.LEFT, wraplength=300).pack(anchor=tk.W)
        self.cc_context_text = tk.Text(ctx_box, height=3, width=50, wrap=tk.WORD,
                                       font=("Arial", 9))
        self.cc_context_text.pack(fill=tk.X, pady=(2, 0))

        # Optional Script / Description (optional)
        script_box = tk.LabelFrame(left, text="📝 Optional Script / Description (optional)",
                                   font=("Arial", 10, "bold"), padx=10, pady=8)
        script_box.pack(fill=tk.X, pady=(0, 6))
        tk.Label(script_box,
                 text="Optionally provide your own script or a detailed description of what happens in the video. Gemini will use this as a reference.",
                 font=("Arial", 8), fg="#666", justify=tk.LEFT, wraplength=300).pack(anchor=tk.W)
        self.cc_script_text = tk.Text(script_box, height=4, width=50, wrap=tk.WORD,
                                      font=("Arial", 9))
        self.cc_script_text.pack(fill=tk.X, pady=(2, 0))

        # Options
        opt_box = tk.LabelFrame(left, text="⚙️ Settings", font=("Arial", 10, "bold"),
                                padx=10, pady=10)
        opt_box.pack(fill=tk.X, pady=(0, 8))

        # Language
        tk.Label(opt_box, text="Target Language:").pack(anchor=tk.W)
        self.cc_lang_var = tk.StringVar(value="english")
        cc_lang_menu = ttk.Combobox(opt_box, textvariable=self.cc_lang_var,
                                     values=["english", "russian", "arabic", "spanish",
                                             "french", "german", "hindi", "urdu",
                                             "chinese", "japanese"],
                                     state="readonly", width=30)
        cc_lang_menu.pack(fill=tk.X, pady=(2, 5))

        # Niche / Commentary preset — each niche is its own editable prompt.
        # Any prompt whose slug starts with "case_commentary" shows up here,
        # so adding a new niche is just "Add New" in the editor (no code).
        tk.Label(opt_box, text="Niche / Commentary Type:").pack(anchor=tk.W)
        self.cc_niche_var = tk.StringVar(value="")
        self.cc_niche_menu = ttk.Combobox(opt_box, textvariable=self.cc_niche_var,
                                           values=[], state="readonly", width=30)
        self.cc_niche_menu.pack(fill=tk.X, pady=(2, 4))
        self.cc_niche_menu.bind("<<ComboboxSelected>>",
                                lambda _e: self._cc_update_preset_help())

        tk.Button(opt_box, text="✏️ Edit / Add Niche Prompt",
                  font=("Arial", 8), bg="#455A64", fg="white",
                  command=self._cc_edit_prompt).pack(anchor=tk.W, pady=(0, 4))

        # ── Vertical / Shorts reframe flag ──────────────────────
        # Writes a "Vertical Format" column into the Excel. Tool 2 (Video
        # Automation Studio) reads it and reframes the stitched video to 9:16
        # with speaking-face tracking. The long-video story-cut niche turns
        # this on automatically at save time even if left unchecked.
        self.cc_vertical_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            opt_box,
            text="📱 Vertical / Shorts (9:16) with face tracking",
            variable=self.cc_vertical_var, font=("Arial", 9),
            anchor=tk.W).pack(fill=tk.X, pady=(0, 4))
        tk.Label(opt_box,
                 text="Tool 2 crops to 9:16 and keeps the speaker in frame. "
                      "Auto-on for the Long Video niche.",
                 font=("Arial", 7), fg="#666", justify=tk.LEFT,
                 wraplength=300).pack(anchor=tk.W, pady=(0, 4))

        # ── Aspect ratio (Movie Recap niche: 16:9 / 9:16 / 1:1) ──────────────
        # Written to the Excel "Aspect Ratio" column. Tool 2 renders the
        # stitched cut in this shape. 16:9 = native (no reframe), 9:16 = vertical
        # reframe with face tracking, 1:1 = square. Only meaningful for the
        # story-cut niches (Movie Recap / Long Video); harmless otherwise.
        asp_row = tk.Frame(opt_box)
        asp_row.pack(fill=tk.X, pady=(0, 4))
        tk.Label(asp_row, text="🖼 Aspect ratio:",
                 font=("Arial", 9)).pack(side=tk.LEFT, padx=(0, 6))
        self.cc_aspect_var = tk.StringVar(value="16:9 (horizontal)")
        self.cc_aspect_menu = ttk.Combobox(
            asp_row, textvariable=self.cc_aspect_var,
            values=["16:9 (horizontal)", "9:16 (vertical)", "1:1 (square)"],
            state="readonly", width=18)
        self.cc_aspect_menu.pack(side=tk.LEFT)

        # ── Target duration for story-cut mode (MM:SS format) ────────────────
        dur_row = tk.Frame(opt_box)
        dur_row.pack(fill=tk.X, pady=(0, 4))
        tk.Label(dur_row, text="⏱ Target duration (MM:SS):",
                 font=("Arial", 9)).pack(side=tk.LEFT, padx=(0, 6))
        self.cc_duration_var = tk.StringVar(value="3:00")
        self.cc_duration_entry = tk.Entry(
            dur_row, textvariable=self.cc_duration_var, width=8,
            font=("Arial", 9), justify=tk.CENTER)
        self.cc_duration_entry.pack(side=tk.LEFT)
        tk.Label(dur_row, text="  e.g. 2:50 or 3:00 (target cut length)",
                 font=("Arial", 7), fg="#666").pack(side=tk.LEFT, padx=(4, 0))

        # ── TTS Engine selector ─────────────────────────────────
        tts_frame = tk.LabelFrame(opt_box, text="🎤 TTS Engine",
                                  font=("Arial", 10, "bold"), padx=10, pady=5)
        tts_frame.pack(fill=tk.X, pady=(2, 6))
        self.cc_tts_engine_var = tk.StringVar(value="cloud")
        tk.Radiobutton(tts_frame, text="Cloud TTS (Standard — 160 WPM)",
                       variable=self.cc_tts_engine_var,
                       value="cloud", font=("Arial", 9)).pack(anchor=tk.W, padx=5, pady=1)
        tk.Radiobutton(tts_frame, text="Qwen3 TTS (Slower — ~110 WPM)",
                       variable=self.cc_tts_engine_var,
                       value="qwen3", font=("Arial", 9)).pack(anchor=tk.W, padx=5, pady=1)
        tk.Radiobutton(tts_frame, text="Gemini TTS (auto speed)",
                       variable=self.cc_tts_engine_var,
                       value="gemini", font=("Arial", 9)).pack(anchor=tk.W, padx=5, pady=1)

        # ── Voice Model + Tone/Style selectors (Gemini TTS calibration) ──
        cal_frame = tk.LabelFrame(opt_box, text="🎙️ Voice & Tone (Gemini TTS)",
                                  font=("Arial", 10, "bold"), padx=10, pady=5)
        cal_frame.pack(fill=tk.X, pady=(2, 6))

        row1 = tk.Frame(cal_frame)
        row1.pack(fill=tk.X, padx=5, pady=2)
        tk.Label(row1, text="Voice Model:", font=("Arial", 9)).pack(side=tk.LEFT, padx=(0, 5))
        self.cc_voice_model_var = tk.StringVar(value="Zephyr")
        voice_models = sorted(_VOICE_WPM.keys())
        voice_menu = tk.OptionMenu(row1, self.cc_voice_model_var, *voice_models)
        voice_menu.config(font=("Arial", 9))
        voice_menu.pack(side=tk.LEFT, fill=tk.X, expand=True)
        # Auto-save voice model preference when user changes it
        self.cc_voice_model_var.trace_add("write", lambda *_: self._cc_save_tone_preference())

        row2 = tk.Frame(cal_frame)
        row2.pack(fill=tk.X, padx=5, pady=2)
        tk.Label(row2, text="Tone / Style:", font=("Arial", 9)).pack(side=tk.LEFT, padx=(0, 5))
        self.cc_tone_var = tk.StringVar(value="Storytelling")
        tone_names = sorted(_TONE_PACING.keys())
        tone_menu = tk.OptionMenu(row2, self.cc_tone_var, *tone_names)
        tone_menu.config(font=("Arial", 9))
        tone_menu.pack(side=tk.LEFT, fill=tk.X, expand=True)
        # Auto-save tone preference when user changes it
        self.cc_tone_var.trace_add("write", lambda *_: self._cc_save_tone_preference())

        # About this niche (details panel)
        cc_help_box = tk.LabelFrame(opt_box, text="📖 About this niche",
                                    font=("Arial", 9, "bold"), fg="#1565C0")
        cc_help_box.pack(fill=tk.X, pady=(2, 4))
        self.cc_help_label = tk.Label(
            cc_help_box, text="Select a niche to see what it does.",
            font=("Arial", 8), justify=tk.LEFT, anchor=tk.W,
            wraplength=300, fg="#333")
        self.cc_help_label.pack(fill=tk.X, padx=6, pady=6)

        # Populate the niche dropdown now that it exists.
        self._cc_refresh_presets()
        # Restore last-selected tone from saved preferences
        self._cc_load_tone_preference()

        # Generate button
        tk.Button(left, text="🎬 Generate Case Commentary",
                  font=("Arial", 11, "bold"), bg="#4CAF50", fg="white",
                  command=self._cc_generate).pack(fill=tk.X, pady=(0, 5))
        self.cc_status_label = tk.Label(left, text="", fg="blue", wraplength=400)
        self.cc_status_label.pack(anchor=tk.W)

        # ── Batch Process Section ─────────────────────────────
        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=6)

        cc_batch_box = tk.LabelFrame(left, text="📦 Batch Process (Channel)",
                                     font=("Arial", 10, "bold"), padx=10, pady=8)
        cc_batch_box.pack(fill=tk.X, pady=(0, 6))

        sel_row = tk.Frame(cc_batch_box)
        sel_row.pack(fill=tk.X, pady=(0, 4))
        tk.Button(sel_row, text="📁 Select Channel",
                  command=self._cc_select_channel,
                  bg="#607D8B", fg="white", font=("Arial", 9)).pack(side=tk.LEFT)
        self._cc_batch_info_label = tk.Label(sel_row, text="No channel selected",
                                             fg="gray", font=("Arial", 9))
        self._cc_batch_info_label.pack(side=tk.LEFT, padx=8)

        self._cc_batch_list_text = tk.Text(cc_batch_box, height=5, width=50,
                                           state=tk.DISABLED, font=("Consolas", 9),
                                           bg="#fafafa", wrap=tk.NONE)
        self._cc_batch_list_text.pack(fill=tk.X, pady=(0, 4))
        self._cc_batch_list_text.tag_configure("waiting", foreground="gray")
        self._cc_batch_list_text.tag_configure("processing", foreground="#1565C0")
        self._cc_batch_list_text.tag_configure("done", foreground="#2E7D32")
        self._cc_batch_list_text.tag_configure("error", foreground="red")

        bbtn_row = tk.Frame(cc_batch_box)
        bbtn_row.pack(fill=tk.X)
        self._cc_batch_gen_btn = tk.Button(
            bbtn_row, text="▶ Generate All", state=tk.DISABLED,
            font=("Arial", 9, "bold"), bg="#2196F3", fg="white",
            command=self._cc_generate_all)
        self._cc_batch_gen_btn.pack(side=tk.LEFT, padx=2)
        self._cc_batch_save_btn = tk.Button(
            bbtn_row, text="💾 Save All to Excel", state=tk.DISABLED,
            font=("Arial", 9, "bold"), bg="#4CAF50", fg="white",
            command=self._cc_save_batch_excel)
        self._cc_batch_save_btn.pack(side=tk.LEFT, padx=2)
        self._cc_batch_prog_label = tk.Label(bbtn_row, text="", fg="blue",
                                             font=("Arial", 9))
        self._cc_batch_prog_label.pack(side=tk.RIGHT, padx=4)

        self._cc_batch_videos = []
        self._cc_batch_channel_path = None

        # ── Right panel: results ────────────────────────────
        right = tk.Frame(main_frame)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))

        # Results notebook with tabs
        self.cc_result_notebook = ttk.Notebook(right)
        self.cc_result_notebook.pack(fill=tk.BOTH, expand=True)

        # Summary tab
        self.cc_summary_frame = tk.Frame(self.cc_result_notebook)
        self.cc_result_notebook.add(self.cc_summary_frame, text="  Summary  ")
        self.cc_summary_text = tk.Text(self.cc_summary_frame, height=10, wrap=tk.WORD,
                                        font=("Arial", 11, "bold"), bg="#fafafa")
        self.cc_summary_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.cc_summary_text.tag_configure("hook", foreground="#c0392b", font=("Arial", 11, "bold"))
        self.cc_summary_text.tag_configure("body", foreground="#2c3e50", font=("Arial", 11))

        # Montage Clips tab
        self.cc_clips_frame = tk.Frame(self.cc_result_notebook)
        self.cc_result_notebook.add(self.cc_clips_frame, text="  Montage Clips  ")
        columns = ("#", "Start", "End", "Duration", "Description")
        self.cc_clips_tree = ttk.Treeview(self.cc_clips_frame, columns=columns,
                                           show="headings", height=8)
        for col in columns:
            self.cc_clips_tree.heading(col, text=col)
            width = 40 if col in ("#", "Duration") else 80 if col in ("Start", "End") else 250
            self.cc_clips_tree.column(col, width=width, anchor=tk.CENTER if col != "Description" else tk.W)
        self.cc_clips_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Commentary Spots tab
        self.cc_spots_frame = tk.Frame(self.cc_result_notebook)
        self.cc_result_notebook.add(self.cc_spots_frame, text="  Commentary Spots  ")
        columns = ("#", "Timestamp", "Emotion", "Text")
        self.cc_spots_tree = ttk.Treeview(self.cc_spots_frame, columns=columns,
                                           show="headings", height=8)
        for col in columns:
            self.cc_spots_tree.heading(col, text=col)
            width = 40 if col == "#" else 80 if col in ("Timestamp", "Emotion") else 300
            self.cc_spots_tree.column(col, width=width, anchor=tk.CENTER if col not in ("Text",) else tk.W)
        self.cc_spots_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Full Output tab
        self.cc_full_frame = tk.Frame(self.cc_result_notebook)
        self.cc_result_notebook.add(self.cc_full_frame, text="  Full Output  ")
        self.cc_full_text = tk.Text(self.cc_full_frame, height=20, wrap=tk.WORD,
                                     font=("Consolas", 9))
        self.cc_full_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Action buttons below result notebook
        btn_frame = tk.Frame(right)
        btn_frame.pack(fill=tk.X, pady=(5, 0))
        tk.Button(btn_frame, text="📋 Copy Summary",
                  command=self._cc_copy_summary).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_frame, text="💾 Save to Excel",
                  font=("Arial", 9, "bold"), bg="#4CAF50", fg="white",
                  command=self._cc_save_excel).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_frame, text="Clear",
                  command=self._cc_clear).pack(side=tk.RIGHT, padx=2)

        # Internal state
        self._cc_file_path = None
        self._cc_last_result = None  # {summary, clips: [...], spots: [...]}
        self._cc_video_url = None

    # ── Case Commentary helpers ───────────────────────────────────

    def _cc_browse_file(self):
        """Browse and select local video file(s) — supports multi-select."""
        from tkinter import filedialog
        from pathlib import Path
        paths = filedialog.askopenfilenames(
            title="Select Video File(s) (Ctrl+click for multiple)",
            filetypes=[("Video files", "*.mp4 *.mov *.avi *.webm *.mkv *.m4v"),
                       ("All files", "*.*")]
        )
        if not paths:
            return

        if len(paths) == 1:
            self._cc_file_path = paths[0]
            self.cc_url_text.delete("1.0", tk.END)
            self.cc_file_label.config(text=Path(paths[0]).name, fg="black")
            return

        # Multiple files → fill batch list
        first_parent = Path(paths[0]).parent
        self._cc_batch_channel_path = first_parent
        self._cc_batch_videos = [
            {"path": p, "name": Path(p).name, "status": "waiting", "result": None}
            for p in paths
        ]
        folder_name = first_parent.name or "Selected"
        self._cc_batch_info_label.config(
            text=f"📁 {folder_name} ({len(self._cc_batch_videos)} videos)", fg="black")
        self._cc_refresh_batch_list()
        self._cc_batch_gen_btn.config(state=tk.NORMAL)
        self._cc_batch_save_btn.config(state=tk.DISABLED)
        self._cc_batch_prog_label.config(
            text=f"Ready • 0/{len(self._cc_batch_videos)}", fg="blue")
        self.cc_status_label.config(
            text=f"✅ {len(self._cc_batch_videos)} videos loaded for batch processing",
            fg="green")

    def _cc_browse_channel(self):
        """Browse the channels folder — supports multi-select."""
        from tkinter import filedialog
        from pathlib import Path

        channels_dir = Path(__file__).parent.parent / "channels"
        channels_dir.mkdir(parents=True, exist_ok=True)

        paths = filedialog.askopenfilenames(
            title="Select Video(s) from Channels (Ctrl+click for multiple)",
            initialdir=str(channels_dir),
            filetypes=[("Video files", "*.mp4 *.mov *.avi *.webm *.mkv *.m4v"),
                       ("All files", "*.*")]
        )
        if not paths:
            return

        if len(paths) == 1:
            self._cc_file_path = paths[0]
            self._cc_video_url = None
            self.cc_url_text.delete("1.0", tk.END)
            try:
                rel = Path(paths[0]).relative_to(channels_dir)
                label = f"📂 {rel}"
            except ValueError:
                label = Path(paths[0]).name
            self.cc_file_label.config(text=label, fg="green")
            return

        # Multiple files → fill batch list
        first_parent = Path(paths[0]).parent
        self._cc_batch_channel_path = first_parent
        self._cc_batch_videos = [
            {"path": p, "name": Path(p).name, "status": "waiting", "result": None}
            for p in paths
        ]
        try:
            rel = first_parent.relative_to(channels_dir)
            label = str(rel)
        except ValueError:
            label = first_parent.name
        self._cc_batch_info_label.config(
            text=f"📂 {label} ({len(self._cc_batch_videos)} videos)", fg="black")
        self._cc_refresh_batch_list()
        self._cc_batch_gen_btn.config(state=tk.NORMAL)
        self._cc_batch_save_btn.config(state=tk.DISABLED)
        self._cc_batch_prog_label.config(
            text=f"Ready • 0/{len(self._cc_batch_videos)}", fg="blue")
        self.cc_status_label.config(
            text=f"✅ {len(self._cc_batch_videos)} videos loaded for batch processing",
            fg="green")

    def _cc_clear(self):
        """Clear all results."""
        self.cc_summary_text.delete("1.0", tk.END)
        for row in self.cc_clips_tree.get_children():
            self.cc_clips_tree.delete(row)
        for row in self.cc_spots_tree.get_children():
            self.cc_spots_tree.delete(row)
        self.cc_full_text.delete("1.0", tk.END)
        self._cc_last_result = None

    def _cc_copy_summary(self):
        """Copy the summary text to clipboard."""
        text = self.cc_summary_text.get("1.0", tk.END).strip()
        if text:
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            self.cc_status_label.config(text="✅ Summary copied to clipboard.", fg="green")

    # ── Case Commentary batch helpers ─────────────────────────────

    def _cc_select_channel(self):
        """Select a channel folder with videos/ subdirectory for batch."""
        from tkinter import filedialog
        from pathlib import Path

        base_dir = Path(__file__).parent.parent / "channels"
        base_dir.mkdir(parents=True, exist_ok=True)

        folder = filedialog.askdirectory(
            title="Select Channel Folder (with videos/ subfolder)",
            initialdir=str(base_dir))
        if not folder:
            return

        folder = Path(folder)
        if folder.name == "videos":
            videos_dir = folder
            channel_root = folder.parent
        else:
            videos_dir = folder / "videos"
            channel_root = folder

        if not videos_dir.is_dir():
            self.cc_status_label.config(
                text="❌ No 'videos/' folder found in this channel.", fg="red")
            return

        video_files = []
        for ext in ("*.mp4", "*.mov", "*.avi", "*.webm", "*.mkv", "*.m4v"):
            video_files.extend(sorted(videos_dir.glob(ext)))

        if not video_files:
            self.cc_status_label.config(
                text="❌ No video files found in channel.", fg="red")
            return

        self._cc_batch_channel_path = channel_root
        self._cc_batch_videos = [
            {"path": str(v), "name": v.name, "status": "waiting", "result": None}
            for v in video_files
        ]

        try:
            rel = channel_root.relative_to(base_dir.parent)
            label = str(rel)
        except ValueError:
            label = channel_root.name
        self._cc_batch_info_label.config(
            text=f"📂 {label} ({len(self._cc_batch_videos)} videos)", fg="black")
        self._cc_refresh_batch_list()
        self._cc_batch_gen_btn.config(state=tk.NORMAL)
        self._cc_batch_save_btn.config(state=tk.DISABLED)
        self._cc_batch_prog_label.config(
            text=f"Ready • 0/{len(self._cc_batch_videos)}", fg="blue")
        self.cc_status_label.config(
            text=f"✅ Channel loaded: {len(self._cc_batch_videos)} videos", fg="green")

    def _cc_refresh_batch_list(self):
        """Rebuild the batch video list Text widget."""
        w = self._cc_batch_list_text
        w.config(state=tk.NORMAL)
        w.delete("1.0", tk.END)
        icons = {"waiting": "⏳", "processing": "🔄", "done": "✅", "error": "❌"}
        for v in self._cc_batch_videos:
            w.insert(tk.END, f"{icons.get(v['status'], '⏳')} {v['name']}\n", v["status"])
        w.config(state=tk.DISABLED)

    def _cc_generate_all(self):
        """Batch-process every video in the selected channel."""
        if not self._cc_batch_videos:
            return

        for v in self._cc_batch_videos:
            v["status"] = "waiting"
            v["result"] = None
        self._cc_refresh_batch_list()
        self._cc_batch_gen_btn.config(state=tk.DISABLED)
        self._cc_batch_save_btn.config(state=tk.DISABLED)

        lang = self.cc_lang_var.get().strip()
        niche = self.cc_niche_var.get().strip()
        cc_slug = self._cc_selected_slug()
        # ── Vertical / Shorts ⇒ force the STORY-CUT prompt ──────────────
        # (see single-generate path for full rationale) Tool 2's story-cut
        # mode only works with the long-video prompt; the teaser prompt would
        # produce 5s flash-cuts + full-video commentary that get truncated.
        try:
            if bool(self.cc_vertical_var.get()):
                cc_slug = "case_commentary_longvideo"
        except Exception:
            pass

        def _task():
            try:
                from core.script_generator import ScriptGenerator

                api_keys = getattr(self, '_saved_api_keys', [])
                sa_path = self.sa_path_var.get().strip()

                if not api_keys and not sa_path:
                    self.root.after(0, lambda: self._cc_batch_prog_label.config(
                        text="❌ No API keys configured", fg="red"))
                    self.root.after(0, lambda: self._cc_batch_gen_btn.config(
                        state=tk.NORMAL))
                    return

                if sa_path and not api_keys:
                    gen = ScriptGenerator(service_account_path=sa_path)
                elif api_keys:
                    gen = ScriptGenerator(api_keys=api_keys)
                else:
                    return

                total = len(self._cc_batch_videos)
                done = 0
                errors = 0

                for i, v in enumerate(self._cc_batch_videos):
                    v["status"] = "processing"
                    self.root.after(0, lambda: self._cc_refresh_batch_list())
                    self.root.after(0, lambda i=i: self._cc_batch_prog_label.config(
                        text=f"⏳ Processing {i+1}/{total}...", fg="blue"))

                    try:
                        # Reuse same _cc_generate logic per video
                        def status(msg):
                            pass  # silence per-video status
                        local_path = gen._resolve_video(None, v["path"])
                        if isinstance(local_path, dict) and "error" in local_path:
                            v["result"] = {"error": local_path["error"]}
                            v["status"] = "error"
                            errors += 1
                            self.root.after(0, lambda: self._cc_refresh_batch_list())
                            continue

                        # Upload to Gemini File API (session-cached)
                        _cache_key = None
                        try:
                            _cache_key = str(Path(local_path).resolve())
                        except Exception:
                            pass
                        _cached = getattr(self, '_cc_upload_cache', {})
                        if _cache_key and _cache_key in _cached:
                            upload = _cached[_cache_key]
                        else:
                            upload = gen._upload_video(local_path)
                            if "error" not in upload and _cache_key:
                                _cached[_cache_key] = upload
                                self._cc_upload_cache = _cached
                        if "error" in upload:
                            v["result"] = {"error": upload["error"]}
                            v["status"] = "error"
                            errors += 1
                            self.root.after(0, lambda: self._cc_refresh_batch_list())
                            continue

                        prompt_template = gen.get_prompt_data(cc_slug)
                        if not prompt_template:
                            v["result"] = {"error": "Prompt not found"}
                            v["status"] = "error"
                            errors += 1
                            self.root.after(0, lambda: self._cc_refresh_batch_list())
                            continue

                        raw_template = prompt_template.get("narration_prompt", "")
                        prompt = raw_template.replace("{language}", lang)
                        prompt = prompt.replace("{niche_angle}", niche)
                        try:
                            _tgt_sec = self._cc_duration_to_sec(
                                getattr(self, 'cc_duration_var',
                                        None).get()) if getattr(
                                self, 'cc_duration_var', None) else 180
                        except (ValueError, AttributeError):
                            _tgt_sec = 180
                        prompt = prompt.replace("{target_duration}",
                                                str(_tgt_sec))

                        # ── Inject context / backstory and optional script ──
                        context = self.cc_context_text.get("1.0", tk.END).strip() if hasattr(self, 'cc_context_text') else ""
                        script_input = self.cc_script_text.get("1.0", tk.END).strip() if hasattr(self, 'cc_script_text') else ""
                        if context:
                            prompt += f"\n\n📖 CONTEXT / BACKSTORY (use this to shape the narrative):\n{context}\n"
                        # ── Auto-context: video metadata from Excel ──
                        auto_ctx = self._cc_read_video_excel_metadata(v["path"])
                        if auto_ctx:
                            prompt += "📹 VIDEO METADATA (title, description, captions, transcript — CRITICAL: use this info as the PRIMARY source for the story. Weave these facts into a gripping narrative covering who, what, where, when, why, and how. Do NOT only describe what the video shows — tell the full incident story from this metadata):\n" + auto_ctx + "\n"
                        if script_input:
                            prompt += f"\n\n📝 OPTIONAL SCRIPT / DESCRIPTION (incorporate this into the narration where relevant):\n{script_input}\n"

                        # Measure duration so long videos get down-sampled
                        # (avoids HTTP 400 "invalid argument" on the analysis call).
                        _bd = 0.0
                        try:
                            import subprocess, json as _json
                            _probe = subprocess.run(
                                ['ffprobe', '-v', 'error', '-show_entries',
                                 'format=duration', '-of', 'json', local_path],
                                capture_output=True, text=True, timeout=30)
                            if _probe.returncode == 0:
                                _bd = float(_json.loads(_probe.stdout).get(
                                    'format', {}).get('duration', 0))
                        except Exception:
                            pass

                        result = gen._call_gemini_with_file(prompt, upload,
                                                            timeout=600,
                                                            video_duration=_bd)
                        if "error" in result:
                            v["result"] = {"error": result["error"]}
                            v["status"] = "error"
                            errors += 1
                        else:
                            response_text = result.get("text", "").strip()
                            if response_text:
                                parsed = self._cc_parse_response(response_text)
                                v["result"] = parsed
                                v["status"] = "done"
                                done += 1
                            else:
                                v["result"] = {"error": "Empty response"}
                                v["status"] = "error"
                                errors += 1
                    except Exception as e:
                        v["status"] = "error"
                        v["result"] = {"error": str(e)}
                        errors += 1

                    self.root.after(0, lambda: self._cc_refresh_batch_list())

                self.root.after(0, lambda d=done, e=errors: self._cc_batch_prog_label.config(
                    text=f"✅ {d} success, {e} error{'s' if e != 1 else ''}",
                    fg="green" if errors == 0 else "#E65100"))
                self.root.after(0, lambda: self._cc_batch_save_btn.config(
                    state=tk.NORMAL))

            except Exception as e:
                self.root.after(0, lambda: self._cc_batch_prog_label.config(
                    text=f"❌ Batch error: {str(e)}", fg="red"))
                self.root.after(0, lambda: self._cc_batch_gen_btn.config(
                    state=tk.NORMAL))

        import threading
        threading.Thread(target=_task, daemon=True).start()

    def _cc_save_batch_excel(self):
        """Save all batch results to a single Excel file (appends if exists)."""
        from tkinter import filedialog, messagebox
        from datetime import datetime
        from pathlib import Path

        results = [v for v in self._cc_batch_videos
                   if v["result"] and "error" not in v["result"]]
        if not results:
            messagebox.showwarning("No Results",
                                   "No successful results to save.")
            return

        _prefix = "CCTV" if "cctv" in self._cc_selected_slug().lower() else "CC"
        _lang = self.cc_lang_var.get().strip()
        default_name = f"{_prefix}_batch"
        if self._cc_batch_channel_path:
            default_name = f"{_prefix}_{self._cc_batch_channel_path.name}"
        default_name += f"_{_lang}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        path = filedialog.asksaveasfilename(
            title="Save Batch Case Commentary to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=str(Path(__file__).parent.parent / "channels"),
            initialfile=default_name,
        )
        if not path:
            return

        try:
            import pandas as pd
            rows = []
            # Aspect ratio + vertical flag (same rules as single save).
            try:
                _aspect_b = (self.cc_aspect_var.get() or "").split()[0]
            except Exception:
                _aspect_b = "16:9"
            try:
                _vertical_b = bool(self.cc_vertical_var.get())
            except Exception:
                _vertical_b = False
            if _aspect_b == "9:16":
                _vertical_b = True
            elif _aspect_b in ("16:9", "1:1"):
                _vertical_b = False
            vertical_flag = "Yes" if _vertical_b else "No"
            _is_movies_b = (self._cc_selected_slug() == "case_commentary_movies")
            story_cut_flag = "Yes" if self._cc_selected_slug() in (
                "case_commentary_longvideo", "case_commentary_movies") else "No"
            mute_original_flag = "Yes" if _is_movies_b else "No"
            # Build one generator for thumbnail-prompt vision calls (reused
            # across all rows). None if no credentials → thumbnail prompt skipped.
            _batch_gen = None
            try:
                from core.script_generator import ScriptGenerator as _SGB
                _bk = getattr(self, '_saved_api_keys', [])
                _bsa = self.sa_path_var.get().strip()
                if _bsa and not _bk:
                    _batch_gen = _SGB(service_account_path=_bsa)
                elif _bk:
                    _batch_gen = _SGB(api_keys=_bk)
            except Exception as _bge:
                print(f"[thumbnail] batch generator init failed: {_bge}")

            for v in results:
                r = v["result"]
                video_name = v["name"]
                summary = r.get("summary", "")
                spots = sorted(r.get("spots", []), key=lambda s: s.get("seconds", 0))
                spots_texts = [s.get("text", "") for s in spots]
                all_text = summary + " " + " ".join(spots_texts)
                clips_ref = "; ".join(
                    f"{c['start_ts']}-{c['end_ts']}|{c['description']}"
                    for c in r.get("clips", []))
                commentary_ref = "; ".join(
                    f"{sp['timestamp']}|{sp['text']}" for sp in spots)
                # ── Thumbnail: extract the Gemini-chosen frame ──
                thumb_ts = r.get("thumbnail_ts", "") or ""
                thumb_text = r.get("thumbnail_text", "") or ""
                thumb_frame_path = ""
                thumb_prompt = ""
                thumb_ref = ""
                if thumb_ts and v.get("path"):
                    try:
                        _sec = r.get("thumbnail_sec", 0) or 0
                        _frame_out = Path(path).with_name(
                            f"{_extract_video_id(video_name)}_thumbframe.jpg")
                        import subprocess as _sp
                        _sp.run(
                            ['ffmpeg', '-y', '-loglevel', 'error',
                             '-ss', str(_sec), '-i', str(v["path"]),
                             '-frames:v', '1', '-q:v', '2', str(_frame_out)],
                            check=True, capture_output=True, timeout=60)
                        if _frame_out.is_file() and _frame_out.stat().st_size > 0:
                            thumb_frame_path = str(_frame_out)
                    except Exception:
                        thumb_frame_path = ""
                # ── Thumbnail styling prompt from the original cover ──
                # Batch works on local files; reconstruct a YouTube URL from
                # the 11-char video id so the original thumbnail can be pulled.
                # Non-YouTube ids simply fail the download → blank → fallback.
                _vid_b = _extract_video_id(video_name)
                if _batch_gen is not None and re.fullmatch(r'[A-Za-z0-9_-]{11}', _vid_b or ''):
                    try:
                        thumb_prompt, thumb_ref = self._cc_build_thumbnail_prompt(
                            _batch_gen,
                            f"https://www.youtube.com/watch?v={_vid_b}",
                            _vid_b, path,
                            title="", niche=self.cc_niche_var.get().strip(),
                            lang=self.cc_lang_var.get().strip(),
                            vertical=_vertical_b)
                    except Exception as _bpe:
                        print(f"[thumbnail] batch prompt build failed: {_bpe}")
                rows.append({
                    "Video ID": _extract_video_id(video_name),
                    "Title": "",
                    "Video File": video_name,
                    "Language": self.cc_lang_var.get().strip(),
                    "Niche": self.cc_niche_var.get().strip(),
                    "Custom Title": "",
                    "Custom Script": summary,
                    "Word Count": len(summary.split()),
                    "Montage Clips": clips_ref,
                    "Commentary Spots": commentary_ref,
                    "Vertical Format": vertical_flag,
                    "Aspect Ratio": _aspect_b,
                    "Story Cut": story_cut_flag,
                    "Mute Original": mute_original_flag,
                    "Thumbnail Time": thumb_ts,
                    "Thumbnail Text": thumb_text,
                    "Thumbnail Frame": thumb_frame_path,
                    "Thumbnail Prompt": thumb_prompt,
                    "Thumbnail Ref": thumb_ref,
                    "Voiceover Style": r.get("voiceover_style", ""),
                    "Voiceover Speed (WPM)": r.get("voiceover_speed", 0),
                    "Hashtag 1": "",
                    "Hashtag 2": "",
                    "Date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                })
            # Append/merge with existing file (skip duplicate Video IDs)
            existing_path = Path(path)
            if existing_path.exists():
                try:
                    old_df = pd.read_excel(path)
                    old_rows = old_df.to_dict("records")
                    existing_ids = {str(r.get("Video ID", "")) for r in old_rows if r.get("Video ID")}
                    deduped_new = [r for r in rows if str(r.get("Video ID", "")) not in existing_ids]
                    rows = old_rows + deduped_new
                except Exception:
                    pass
            df = pd.DataFrame(rows)
            df.to_excel(path, index=False)
            messagebox.showinfo("Saved",
                                f"✅ {len(rows)} scripts saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {str(e)}")

    # ── Case Commentary preset (niche) management ─────────────────
    # Each niche is a full prompt whose slug starts with this prefix. Adding a
    # new niche = adding a new such prompt (via the Edit/Add button) — no code.
    CC_SLUG_PREFIX = "case_commentary"

    def _cc_prompt_entries(self):
        """Return [(slug, display_name)] for every Case Commentary niche prompt.

        A niche is any prompt whose slug starts with CC_SLUG_PREFIX. The base
        'case_commentary' (Courtroom) always sorts first.
        """
        import os, json
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        path = os.path.join(base, "data", "script_prompts.json")
        entries = []
        try:
            with open(path, "r", encoding="utf-8") as f:
                prompts = json.load(f).get("prompts", {})
            for slug, info in prompts.items():
                if slug == self.CC_SLUG_PREFIX or slug.startswith(self.CC_SLUG_PREFIX + "_"):
                    entries.append((slug, info.get("name", slug)))
        except Exception:
            pass
        # Base courtroom prompt first, rest alphabetical by name
        entries.sort(key=lambda e: (e[0] != self.CC_SLUG_PREFIX, e[1].lower()))
        return entries

    def _cc_refresh_presets(self, select_slug=None):
        """Repopulate the niche dropdown from current prompts; keep selection."""
        if not hasattr(self, "cc_niche_menu"):
            return
        entries = self._cc_prompt_entries()
        names = [name for _slug, name in entries]
        self.cc_niche_menu["values"] = names
        # Decide what to select
        want = None
        if select_slug:
            want = next((n for s, n in entries if s == select_slug), None)
        if not want:
            cur = self.cc_niche_var.get()
            want = cur if cur in names else (names[0] if names else "")
        self.cc_niche_var.set(want)
        self._cc_update_preset_help()

    def _cc_selected_slug(self):
        """Resolve the selected niche display name back to its prompt slug."""
        name = self.cc_niche_var.get()
        for slug, dname in self._cc_prompt_entries():
            if dname == name:
                return slug
        return self.CC_SLUG_PREFIX

    def _cc_update_preset_help(self):
        """Show the description of the selected niche in the details panel."""
        if not hasattr(self, "cc_help_label"):
            return
        import os, json
        slug = self._cc_selected_slug()
        desc = ""
        try:
            base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            path = os.path.join(base, "data", "script_prompts.json")
            with open(path, "r", encoding="utf-8") as f:
                info = json.load(f).get("prompts", {}).get(slug, {})
            desc = info.get("description", "")
        except Exception:
            pass
        self.cc_help_label.config(
            text=desc or "Gemini watches the video and outputs a summary, "
                         "montage clips, and commentary spots.")

    def _cc_edit_prompt(self):
        """Open the prompt manager focused on Case Commentary niche prompts.

        New prompts added here are auto-tagged as Case Commentary niches so they
        immediately appear in the niche dropdown.
        """
        PromptManagerDialog(
            self.root, self,
            category="case_commentary",
            preselect_slug=self._cc_selected_slug(),
            on_close=self._cc_refresh_presets)

    def _cc_save_tone_preference(self):
        """Save the selected tone and voice model to gemini_config.json so they persist across restarts."""
        import json
        try:
            config_path = self._get_gemini_config_path()
            config = {}
            if config_path.exists():
                with open(config_path, 'r') as f:
                    config = json.load(f)
            config['cc_tone'] = self.cc_tone_var.get()
            config['cc_voice_model'] = self.cc_voice_model_var.get()
            config_path.parent.mkdir(parents=True, exist_ok=True)
            with open(config_path, 'w') as f:
                json.dump(config, f)
        except Exception:
            pass  # best-effort save

    def _cc_load_tone_preference(self):
        """Restore the last-selected tone and voice model from saved config."""
        import json
        try:
            config_path = self._get_gemini_config_path()
            if config_path.exists():
                with open(config_path, 'r') as f:
                    config = json.load(f)
                saved_tone = config.get('cc_tone', '')
                if saved_tone and saved_tone in _TONE_PACING:
                    self.cc_tone_var.set(saved_tone)
                saved_vm = config.get('cc_voice_model', '')
                if saved_vm and saved_vm in _VOICE_WPM:
                    self.cc_voice_model_var.set(saved_vm)
        except Exception:
            pass  # best-effort load

    def _cc_generate(self):
        """Main generation: download video, upload to Gemini, get structured output."""
        import threading

        url = self.cc_url_text.get("1.0", tk.END).strip()
        file_path = self._cc_file_path
        lang = self.cc_lang_var.get().strip()
        niche = self.cc_niche_var.get().strip()
        context = self.cc_context_text.get("1.0", tk.END).strip() if hasattr(self, 'cc_context_text') else ""
        script_input = self.cc_script_text.get("1.0", tk.END).strip() if hasattr(self, 'cc_script_text') else ""

        if not url and not file_path:
            self.cc_status_label.config(text="❌ Enter a YouTube URL or select a video file.", fg="red")
            return

        self._cc_clear()
        self._cc_video_url = url or file_path
        self.cc_status_label.config(text="⏳ Initializing...", fg="blue")

        def task():
            try:
                from core.script_generator import ScriptGenerator
                api_keys = getattr(self, '_saved_api_keys', [])
                sa_path = self.sa_path_var.get().strip()
                if not api_keys and not sa_path:
                    self.root.after(0, lambda: self.cc_status_label.config(
                        text="❌ Gemini API key not configured. Go to Settings.", fg="red"))
                    return

                if sa_path and not api_keys:
                    gen = ScriptGenerator(service_account_path=sa_path)
                elif api_keys:
                    gen = ScriptGenerator(api_keys=api_keys)
                else:
                    self.root.after(0, lambda: self.cc_status_label.config(
                        text="❌ No API key or service account configured.", fg="red"))
                    return

                def status(msg):
                    self.root.after(0, lambda m=msg: self.cc_status_label.config(text=m, fg="blue"))

                # 1. Resolve video
                status("📥 Downloading video...")
                local_path = gen._resolve_video(url, file_path)
                if isinstance(local_path, dict) and "error" in local_path:
                    self.root.after(0, lambda: self.cc_status_label.config(
                        text=f"❌ {local_path.get('error', 'Download failed')}", fg="red"))
                    return
                # Remember the resolved local file so the Excel export can grab
                # the Gemini-chosen thumbnail frame from it.
                self._cc_local_video = local_path

                # ── Measure video duration for word count calibration ──
                status("⏱️ Measuring video duration...")
                video_duration = 0.0
                try:
                    import subprocess, json
                    probe = subprocess.run(
                        ['ffprobe', '-v', 'error', '-show_entries',
                         'format=duration', '-of', 'json', local_path],
                        capture_output=True, text=True, timeout=30)
                    if probe.returncode == 0:
                        data = json.loads(probe.stdout)
                        video_duration = float(data.get('format', {}).get('duration', 0))
                except Exception:
                    pass

                # 2. Upload to Gemini File API (session-cached — URIs last ~48h)
                _cache_key = None
                try:
                    _cache_key = str(Path(local_path).resolve())
                except Exception:
                    pass
                _cached = getattr(self, '_cc_upload_cache', {})
                if _cache_key and _cache_key in _cached:
                    upload = _cached[_cache_key]
                    status("♻️ Using cached upload (same video)")
                else:
                    status("☁️ Uploading video to Gemini...")
                    upload = gen._upload_video(local_path, progress_callback=lambda m: status(m))
                    if "error" not in upload and _cache_key:
                        _cached[_cache_key] = upload
                        self._cc_upload_cache = _cached
                if "error" in upload:
                    self.root.after(0, lambda: self.cc_status_label.config(
                        text=f"❌ Upload failed: {upload.get('error')}", fg="red"))
                    return

                # 3. Build the case commentary prompt (selected niche)
                status("📝 Building analysis prompt...")
                cc_slug = self._cc_selected_slug()
                # ── Vertical / Shorts ⇒ force the STORY-CUT prompt ─────────
                # Tool 2 runs "story-cut mode" (keep only stitched segments,
                # drop the full source) whenever Vertical Format=Yes.  That
                # mode is ONLY compatible with the long-video prompt, whose
                # segments are 10-40s story beats budgeted to the target
                # duration and whose commentary spots sit INSIDE the cut.  The
                # teaser prompt's 5s flash-cuts + full-video commentary would
                # be kept as ~25s of clips while minutes of narration get
                # truncated to a stub.  So when vertical is on, use story-cut.
                # NOTE: the Movie Recap niche is ALSO a story-cut prompt (its
                # segments are full scenes budgeted to the target and its
                # commentary spots sit inside the cut), so it's already
                # vertical-compatible — do NOT clobber it back to the courtroom
                # long-video prompt. Only force longvideo for non-story-cut niches.
                _STORY_CUT_SLUGS = {"case_commentary_longvideo", "case_commentary_movies"}
                try:
                    if bool(self.cc_vertical_var.get()) and cc_slug not in _STORY_CUT_SLUGS:
                        cc_slug = "case_commentary_longvideo"
                except Exception:
                    pass
                prompt_template = gen.get_prompt_data(cc_slug)
                if not prompt_template:
                    self.root.after(0, lambda: self.cc_status_label.config(
                        text="❌ Case Commentary prompt not found.", fg="red"))
                    return

                raw_template = prompt_template.get("narration_prompt", "")
                prompt = raw_template.replace("{language}", lang)
                prompt = prompt.replace("{niche_angle}", niche)
                # Target-duration (MM:SS or minutes) for the long-video story cut.
                try:
                    _tgt_sec = self._cc_duration_to_sec(
                        getattr(self, 'cc_duration_var',
                                None).get()) if getattr(
                        self, 'cc_duration_var', None) else 180
                except (ValueError, AttributeError):
                    _tgt_sec = 180
                prompt = prompt.replace("{target_duration}",
                                        str(_tgt_sec))

                # ── Inject context / backstory and optional script ──
                # ── Auto-context: video metadata from Excel ──
                auto_ctx = self._cc_read_video_excel_metadata(local_path)
                if auto_ctx:
                    prompt += "📹 VIDEO METADATA (title, description, captions, transcript — CRITICAL: use this info as the PRIMARY source for the story. Weave these facts into a gripping narrative covering who, what, where, when, why, and how. Do NOT only describe what the video shows — tell the full incident story from this metadata):\n" + auto_ctx + "\n"
                if context:
                    prompt += f"\n\n📖 CONTEXT / BACKSTORY (use this to shape the narrative):\n{context}\n"
                if script_input:
                    prompt += f"\n\n📝 OPTIONAL SCRIPT / DESCRIPTION (incorporate this into the narration where relevant):\n{script_input}\n"

                # ── Inject word count constraint from voice model + tone calibration ──
                # In story-cut / vertical mode the FINAL short is only
                # ~target_duration long (the stitched segments), NOT the full
                # source.  Budget the narration to the target, otherwise we'd
                # tell Gemini to write minutes of script for a 3-min short and
                # the voiceover overshoots the video massively.
                # Story-cut niches (courtroom long-video, movie recap) always
                # produce a SHORT of ~target_duration — the final video is the
                # stitched segments, NOT the full source — regardless of aspect
                # ratio (16:9 / 9:16 / 1:1). So budget the narration to the
                # target whenever this is a story-cut niche OR vertical is on.
                # Previously this was gated on the vertical checkbox alone, so a
                # 16:9 long-video cut fell back to the full source duration and
                # the target-duration setting was effectively ignored.
                _wc_dur = video_duration
                try:
                    _is_story_cut = cc_slug in _STORY_CUT_SLUGS
                    if (_is_story_cut or bool(self.cc_vertical_var.get())) and _tgt_sec > 0:
                        _wc_dur = min(video_duration, _tgt_sec) if video_duration > 0 else _tgt_sec
                except Exception:
                    pass
                if _wc_dur > 0:
                    voice_model = getattr(self, 'cc_voice_model_var', None)
                    tone_style = getattr(self, 'cc_tone_var', None)
                    vm = voice_model.get() if voice_model else 'Zephyr'
                    ts = tone_style.get() if tone_style else 'Storytelling'
                    target_words = calc_target_word_count(_wc_dur, vm, ts)
                    wc_hint = (
                        f"\n\n📐 WORD COUNT CONSTRAINT (CRITICAL):\n"
                        f"The final short is ~{_wc_dur:.1f} seconds long. "
                        f"Your narration must fit within this duration at natural speaking pace.\n"
                        f"Voice model: {vm} (~{_VOICE_WPM.get(vm, 105)} WPM at 1.0x)\n"
                        f"Tone/style: {ts} (pacing factor {_TONE_PACING.get(ts, 0.85):.2f})\n"
                        f"Target total word count for the entire script: ~{target_words} words.\n"
                        f"Distribute this budget across the summary, clips, and commentary spots.\n"
                        f"DO NOT exceed {target_words} words total."
                    )
                    prompt += wc_hint

                # Qwen3 mode: reduce word limits so Gemini writes shorter text
                if getattr(self, 'cc_tts_engine_var', None) is not None and self.cc_tts_engine_var.get() == "qwen3":
                    replacements = {
                        "15-20 seconds (40-55 words)": "15-20 seconds (30-40 words)",
                        "max 12 words": "max 8 words",
                        "max 15 words": "max 10 words",
                        "40 to 55 words": "30 to 40 words",
                        "8-15 seconds (25-45 words)": "8-15 seconds (20-30 words)",
                        "25 to 45 words": "20 to 30 words",
                        "200 WPM": "135 WPM",
                    }
                    # max 10 words → max 7 words (only for CCTV, after generic max 12→8 so unmatched ones survive)
                    replacements["max 10 words"] = "max 7 words"
                    for old, new in replacements.items():
                        prompt = prompt.replace(old, new)

                # 4. Call Gemini with the uploaded video
                status("🤖 Gemini analyzing video... This may take a minute.")
                result = gen._call_gemini_with_file(prompt, upload,
                                                     timeout=600,
                                                     progress_callback=lambda m: status(m),
                                                     video_duration=video_duration)
                if "error" in result:
                    self.root.after(0, lambda: self.cc_status_label.config(
                        text=f"❌ Gemini error: {result.get('error')}", fg="red"))
                    return

                response_text = result.get("text", "").strip()
                if not response_text:
                    self.root.after(0, lambda: self.cc_status_label.config(
                        text="❌ Empty response from Gemini.", fg="red"))
                    return

                # 5. Parse the structured response
                parsed = self._cc_parse_response(response_text)
                self._cc_last_result = parsed

                # 6. Display results
                self.root.after(0, lambda: self._cc_display_result(parsed))
                self.root.after(0, lambda: self.cc_status_label.config(
                    text=f"✅ Done! Summary ({len(parsed.get('summary','').split())} words), "
                         f"{len(parsed.get('clips',[]))} montage clips, "
                         f"{len(parsed.get('spots',[]))} commentary spots.",
                    fg="green"))

            except Exception as e:
                import traceback
                traceback.print_exc()
                self.root.after(0, lambda: self.cc_status_label.config(
                    text=f"❌ Error: {str(e)}", fg="red"))

        threading.Thread(target=task, daemon=True).start()

    def _cc_build_thumbnail_prompt(self, gen, url, video_id, out_dir,
                                   title="", niche="", lang="", vertical=False):
        """Download the original thumbnail and have Gemini write a better prompt.

        Only recreates when the source has a PROPER published thumbnail
        (e.g. YouTube maxresdefault).  Otherwise returns ("", "") so the
        caller/Tool 2 falls back to the current frame-grab + text settings.

        Returns (thumb_prompt, thumb_ref_path).  Both empty on fallback.
        """
        if not url:
            return "", ""
        try:
            from core.thumbnail_source import fetch_original_thumbnail
        except Exception as _ie:
            print(f"[thumbnail] source module import failed: {_ie}")
            return "", ""

        ref_out = Path(out_dir).with_name(
            f"{(video_id or 'thumb')}_origthumb.jpg")

        def _log(level, msg):
            print(f"[thumbnail] {level}: {msg}")

        try:
            info = fetch_original_thumbnail(url, ref_out, log=_log)
        except Exception as _fe:
            print(f"[thumbnail] fetch failed: {_fe}")
            return "", ""

        if not info.get("is_proper") or not info.get("path"):
            # No real designed cover — use current settings (frame + text).
            return "", ""

        try:
            res = gen.analyze_thumbnail(
                info["path"], title=title, niche=niche,
                language=lang, vertical=vertical)
        except Exception as _ae:
            print(f"[thumbnail] analyze failed: {_ae}")
            return "", info["path"]

        if "error" in res:
            print(f"[thumbnail] analyze error: {res['error']}")
            return "", info["path"]
        return res.get("prompt", ""), info["path"]

    def _cc_parse_response(self, text):
        """Parse Gemini's structured output into {summary, clips, spots, voiceover_style, voiceover_speed}."""
        result = {"summary": "", "summary_emotion": "", "disclaimer": "", "clips": [], "spots": [],
                  "voiceover_style": "", "voiceover_speed": 0,
                  "thumbnail_ts": "", "thumbnail_sec": 0, "thumbnail_text": ""}

        if not text:
            return result

        # Split by section headers
        sections = re.split(r'^===+\s*', text, flags=re.MULTILINE)

        current_section = None
        for chunk in sections:
            chunk = chunk.strip()
            if not chunk:
                continue

            upper = chunk.upper()

            if upper.startswith("CASE SUMMARY") or upper.startswith("SUMMARY"):
                current_section = "summary"
                # Extract the text after the header
                content = chunk.split("\n", 1)
                if len(content) > 1:
                    raw = content[1].strip()
                    # Gemini sometimes glues the VOICEOVER STYLE / VOICEOVER
                    # SPEED lines (and any trailing metadata) onto the end of
                    # the summary block instead of a separate section. Cut the
                    # summary off at the first VOICEOVER/THUMBNAIL/section
                    # marker so those never leak into the spoken narration.
                    raw = re.split(
                        r'\n\s*(?:VOICEOVER\s+STYLE|VOICEOVER\s+SPEED|THUMBNAIL|MONTAGE\s+CLIPS|COMMENTARY\s+SPOTS|DISCLAIMER)\b',
                        raw, maxsplit=1, flags=re.IGNORECASE)[0].strip()
                    # Optional leading [emotion] tag on the summary
                    emo_match = re.match(r'^\[(\w+)\]\s*(.*)', raw, re.DOTALL)
                    if emo_match:
                        result["summary_emotion"] = emo_match.group(1).strip().lower()
                        result["summary"] = emo_match.group(2).strip()
                    else:
                        result["summary"] = raw

            elif upper.startswith("DISCLAIMER"):
                current_section = "disclaimer"
                content = chunk.split("\n", 1)
                if len(content) > 1:
                    raw = content[1].strip()
                    # Parse: [DISCLAIMER] text
                    disc_match = re.match(r'^\[DISCLAIMER\]\s*(.*)', raw, re.DOTALL)
                    if disc_match:
                        result["disclaimer"] = disc_match.group(1).strip()
                    else:
                        result["disclaimer"] = raw

            elif upper.startswith("MONTAGE CLIPS") or upper.startswith("CLIPS"):
                current_section = "clips"
                content = chunk.split("\n", 1)
                if len(content) > 1:
                    lines = content[1].strip().split("\n")
                    for line in lines:
                        line = line.strip()
                        if not line or line.startswith("-") or line.startswith("="):
                            continue
                        # Parse: MM:SS-MM:SS | description
                        match = re.match(r'(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})\s*[|\|]\s*(.+)', line)
                        if match:
                            start_ts, end_ts, desc = match.groups()
                            # Calculate duration
                            start_sec = self._ts_to_sec(start_ts)
                            end_sec = self._ts_to_sec(end_ts)
                            duration = end_sec - start_sec
                            result["clips"].append({
                                "start_ts": start_ts,
                                "end_ts": end_ts,
                                "start_sec": start_sec,
                                "end_sec": end_sec,
                                "duration": duration,
                                "description": desc.strip(),
                            })

            elif upper.startswith("COMMENTARY SPOTS") or upper.startswith("SPOTS"):
                current_section = "spots"
                content = chunk.split("\n", 1)
                if len(content) > 1:
                    lines = content[1].strip().split("\n")
                    for line in lines:
                        line = line.strip()
                        if not line or line.startswith("-") or line.startswith("="):
                            continue
                        # Parse: MM:SS | [emotion] | commentary text
                        # or legacy: MM:SS | commentary text
                        match = re.match(
                            r'(\d{1,2}:\d{2})\s*[|\|]\s*'
                            r'(?:\[(\w+)\]\s*[|\|]\s*)?(.+)',
                            line)
                        if match:
                            ts, emotion, text_val = match.groups()
                            result["spots"].append({
                                "timestamp": ts,
                                "seconds": self._ts_to_sec(ts),
                                "emotion": (emotion or '').strip().lower(),
                                "text": text_val.strip(),
                            })
                    break

        # Extract voiceover style/speed from the full text (may appear after
        # the COMMENTARY SPOTS section, with or without --- wrapper)
        vo_match = re.search(
            r'VOICEOVER\s+STYLE:\s*(.+?)\s*\n'
            r'VOICEOVER\s+SPEED:\s*(\d+)',
            text, re.IGNORECASE
        )
        if vo_match:
            result["voiceover_style"] = vo_match.group(1).strip()
            result["voiceover_speed"] = int(vo_match.group(2))

        # Thumbnail pick (may appear anywhere; parsed from full text so the
        # early `break` in the section loop above can't swallow it).
        #   THUMBNAIL: MM:SS | catchy overlay text
        thumb_match = re.search(
            r'THUMBNAIL\s*:?\s*(\d{1,2}:\d{2})\s*[|\|]\s*(.+)',
            text, re.IGNORECASE)
        if thumb_match:
            t_ts = thumb_match.group(1).strip()
            result["thumbnail_ts"] = t_ts
            result["thumbnail_sec"] = self._ts_to_sec(t_ts)
            result["thumbnail_text"] = thumb_match.group(2).strip()

        return result

    @staticmethod
    def _ts_to_sec(ts):
        """Convert MM:SS to seconds."""
        parts = ts.split(":")
        return int(parts[0]) * 60 + int(parts[1])

    @staticmethod
    def _cc_duration_to_sec(val):
        """Parse a duration string to seconds.
        Accepts:
        - "MM:SS"  (e.g. "2:50", "03:00")
        - number string representing minutes (e.g. "3" = 180s)
        Default: 180s.
        """
        val = str(val).strip()
        if ':' in val:
            parts = val.split(':')
            try:
                return int(parts[0]) * 60 + int(parts[1])
            except (ValueError, IndexError):
                return 180
        try:
            return int(float(val) * 60)
        except ValueError:
            return 180

    def _cc_display_result(self, parsed):
        """Populate the result widgets with parsed data."""
        # Summary
        summary = parsed.get("summary", "")
        self.cc_summary_text.delete("1.0", tk.END)
        if summary:
            # Try to highlight the first sentence as the hook
            lines = summary.split("\n")
            if len(lines) >= 1:
                first_line = lines[0].strip()
                self.cc_summary_text.insert(tk.END, first_line + "\n", "hook")
                rest = "\n".join(lines[1:]).strip()
                if rest:
                    self.cc_summary_text.insert(tk.END, "\n" + rest, "body")
            else:
                self.cc_summary_text.insert(tk.END, summary, "body")
        else:
            self.cc_summary_text.insert(tk.END, "(No summary generated)")

        # Montage clips
        for row in self.cc_clips_tree.get_children():
            self.cc_clips_tree.delete(row)
        clips = parsed.get("clips", [])
        for i, clip in enumerate(clips, 1):
            dur_str = f"{clip['duration']:.1f}s"
            self.cc_clips_tree.insert("", tk.END, values=(
                i, clip["start_ts"], clip["end_ts"],
                dur_str, clip["description"]
            ))

        # Commentary spots
        for row in self.cc_spots_tree.get_children():
            self.cc_spots_tree.delete(row)
        spots = parsed.get("spots", [])
        for i, spot in enumerate(spots, 1):
            emotion = spot.get("emotion", "") or ""
            self.cc_spots_tree.insert("", tk.END, values=(
                i, spot["timestamp"], emotion, spot["text"]
            ))

        # Full output
        self.cc_full_text.delete("1.0", tk.END)
        self.cc_full_text.insert(tk.END, "=== CASE SUMMARY ===\n")
        self.cc_full_text.insert(tk.END, summary + "\n\n")
        disclaimer = parsed.get("disclaimer", "")
        if disclaimer:
            self.cc_full_text.insert(tk.END, "=== DISCLAIMER ===\n")
            self.cc_full_text.insert(tk.END, f"[DISCLAIMER] {disclaimer}\n\n")
        self.cc_full_text.insert(tk.END, "=== MONTAGE CLIPS ===\n")
        for clip in clips:
            self.cc_full_text.insert(tk.END,
                f"{clip['start_ts']}-{clip['end_ts']} | {clip['description']} "
                f"({clip['duration']:.1f}s)\n")
        self.cc_full_text.insert(tk.END, "\n=== COMMENTARY SPOTS ===\n")
        for spot in spots:
            emo = spot.get("emotion", "") or ""
            self.cc_full_text.insert(tk.END,
                f"{spot['timestamp']} | {emo} | {spot['text']}\n" if emo
                else f"{spot['timestamp']} | {spot['text']}\n")

        # Select Summary tab by default
        self.cc_result_notebook.select(0)

    # ── Excel export ─────────────────────────────────────────────

    # ── Auto-context: read video metadata from Excel ───────────────────

    def _cc_read_video_excel_metadata(self, video_path):
        try:
            from pathlib import Path
            import openpyxl
            vpath = Path(video_path)
            folder = vpath.parent
            vid = vpath.stem
            excel_files = list(folder.glob("results_clean_*.xlsx"))
            if not excel_files:
                return ""
            wb = openpyxl.load_workbook(excel_files[0], read_only=True, data_only=True)
            ws = wb.active
            if not ws:
                return ""
            headers = {str(ws.cell(1, c).value or "").strip().lower(): c
                       for c in range(1, ws.max_column + 1)}
            col_vid        = headers.get("video id", 0)
            col_title      = headers.get("title", 0)
            col_desc       = headers.get("description", 0)
            col_transcript = headers.get("speech transcript", 0)
            col_captions   = headers.get("captions", 0)
            col_overlay    = headers.get("overlay text", 0)
            if not col_vid:
                return ""
            for row in range(2, ws.max_row + 1):
                row_vid = str(ws.cell(row, col_vid).value or "").strip()
                if row_vid == vid:
                    title       = str(ws.cell(row, col_title).value or "") if col_title else ""
                    description = str(ws.cell(row, col_desc).value or "") if col_desc else ""
                    transcript  = str(ws.cell(row, col_transcript).value or "") if col_transcript else ""
                    captions    = str(ws.cell(row, col_captions).value or "") if col_captions else ""
                    overlay     = str(ws.cell(row, col_overlay).value or "") if col_overlay else ""
                    parts = []
                    if title:
                        parts.append("📌 TITLE:\n" + title)
                    if description:
                        parts.append("📄 DESCRIPTION:\n" + description)
                    if overlay:
                        parts.append("🖼️ OVERLAY TEXT (on-screen text visible in video):\n" + overlay)
                    if transcript:
                        parts.append("🎙️ SPEECH TRANSCRIPT:\n" + transcript)
                    if captions:
                        parts.append("💬 CAPTIONS:\n" + captions)
                    wb.close()
                    if parts:
                        return "\n\n".join(parts)
                    return ""
            wb.close()
        except Exception:
            pass
        return ""

    def _cc_save_excel(self):
        """Save the case commentary script to Excel — same format as other tabs."""
        from tkinter import filedialog, messagebox

        if not self._cc_last_result:
            self.cc_status_label.config(text="❌ Nothing to save — generate commentary first.", fg="red")
            return

        summary = self._cc_last_result.get("summary", "")
        if not summary:
            self.cc_status_label.config(text="❌ No summary text to save.", fg="red")
            return

        url = self._cc_video_url or self.cc_url_text.get("1.0", tk.END).strip()
        lang = self.cc_lang_var.get().strip()
        niche = self.cc_niche_var.get().strip()

        # Determine file prefix based on selected niche slug
        _slug = self._cc_selected_slug()
        _prefix = "CCTV" if "cctv" in _slug.lower() else "CC"

        # Extract video ID from URL / file path
        _vid_id = _extract_yt_id(url) or (_extract_video_id(Path(url).name) if url else "unknown")

        # Vertical / shorts reframe flag for Tool 2 (Video Automation Studio).
        # Explicit checkbox wins; otherwise the long-video story-cut slug implies vertical.
        # ── Aspect ratio → derives the vertical flag ──────────────────
        # Movie Recap exposes 16:9 / 9:16 / 1:1. The legacy "Vertical Format"
        # flag stays True only for 9:16 (Tool 2's reframe path); 16:9 and 1:1
        # are non-vertical. The explicit vertical checkbox still forces 9:16.
        try:
            _aspect = (self.cc_aspect_var.get() or "").split()[0]  # "16:9" / "9:16" / "1:1"
        except Exception:
            _aspect = "16:9"
        try:
            _vertical = bool(self.cc_vertical_var.get())
        except Exception:
            _vertical = False
        if _aspect == "9:16":
            _vertical = True
        elif _aspect in ("16:9", "1:1"):
            _vertical = False
        vertical_flag = "Yes" if _vertical else "No"

        # Story-cut mode = keep only stitched segments, drop the source.
        # Both the courtroom long-video and the movie niche use it. The movie
        # niche additionally mutes the original audio (narration-only recap).
        _is_movies = (self._cc_selected_slug() == "case_commentary_movies")
        story_cut_flag = "Yes" if self._cc_selected_slug() in (
            "case_commentary_longvideo", "case_commentary_movies") else "No"
        mute_original_flag = "Yes" if _is_movies else "No"

        path = filedialog.asksaveasfilename(
            title="Save Case Commentary to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=str(Path(__file__).parent.parent / "channels"),
            initialfile=f"{_prefix}_{_vid_id}_{lang}.xlsx"
        )
        if not path:
            return

        # Build the script in standard [timestamp] | [emotion] | text | (words) | [beat] format
        script_lines = []

        # Line 1: the summary at 00:00 as the hook/opening
        summary = self._cc_last_result.get("summary", "")
        summary_emo = self._cc_last_result.get("summary_emotion", "") or ""
        summary_words = len(summary.split())
        emo_part = f"| [{summary_emo}] " if summary_emo else "| "
        script_lines.append(
            f"[00:00] {emo_part}{summary} | ({summary_words} words) | [hook]"
        )

        # Line 2: disclaimer (right after the hook, before commentary)
        disclaimer = self._cc_last_result.get("disclaimer", "")
        if disclaimer:
            disc_words = len(disclaimer.split())
            script_lines.append(
                f"[00:05] | [disclaimer] {disclaimer} | ({disc_words} words) | [disclaimer]"
            )

        # NOTE: Commentary spots are NOT embedded here — they go into the
        # separate "Commentary Spots" column below.  Putting them in both
        # the Custom Script AND Commentary Spots columns makes Tool 2 speak
        # each spot twice (once from the TTS of this script, once from the
        # Commentary Spots TTS pipeline).  The Commentary Spots column is
        # the single source of truth for spots.  Custom Script = summary only.
        # Add a [00:10] placeholder to leave room for disclaimer/future lines.

        full_script = "\n".join(script_lines)
        disc_words = len(disclaimer.split()) if disclaimer else 0
        _cc_spots = self._cc_last_result.get("spots", [])
        total_words = sum(len(s.split()) for s in [summary] + [sp["text"] for sp in _cc_spots]) + disc_words

        # Montage clips — parsed by the video editing tool to create the intro
        clips_ref = "; ".join(
            f"{c['start_ts']}-{c['end_ts']}|{c['description']}"
            for c in self._cc_last_result.get("clips", [])
        )

        # Commentary spots — parsed by the video editing tool to create
        # text overlays during the main video portion
        commentary_ref = "; ".join(
            f"{sp['timestamp']}|{sp['text']}"
            for sp in sorted(
                self._cc_last_result.get("spots", []),
                key=lambda s: s.get("seconds", 0))
        )

        import pandas as pd
        # Extract video_id from URL (supports youtube.com/watch?v=XXX,
        # youtu.be/XXX, youtube.com/shorts/XXX, etc.)
        video_id = ""
        if url:
            vid_match = re.search(r'(?:v=|youtu\.be/|shorts/)([A-Za-z0-9_-]{11})', url)
            if vid_match:
                video_id = vid_match.group(1)
        # Fallback: extract video ID from local file path when no URL
        if not video_id:
            _local_vid = getattr(self, "_cc_file_path", "") or getattr(self, "_cc_local_video", "") or ""
            if _local_vid:
                video_id = _extract_video_id(Path(_local_vid).name)

        # Lightweight title fetch via yt-dlp (no download, cached by yt-dlp)
        title = ""
        if video_id and (("youtube.com" in url or "youtu.be" in url) or not url):
            try:
                import yt_dlp
                with yt_dlp.YoutubeDL({'quiet': True, 'no_warnings': True, 'noplaylist': True}) as ydl:
                    info = ydl.extract_info(f"https://www.youtube.com/watch?v={video_id}", download=False)
                    title = info.get("title", "") or ""
            except Exception:
                pass  # non-critical — title stays empty and user can fill manually

        # ── Thumbnail: extract the Gemini-chosen frame next to the Excel ──
        thumb_ts = self._cc_last_result.get("thumbnail_ts", "") or ""
        thumb_text = self._cc_last_result.get("thumbnail_text", "") or ""
        thumb_frame_path = ""
        if thumb_ts:
            try:
                _local_vid = getattr(self, "_cc_local_video", "") or ""
                if _local_vid and Path(_local_vid).is_file():
                    _sec = self._cc_last_result.get("thumbnail_sec", 0) or 0
                    _frame_out = Path(path).with_name(
                        f"{(video_id or 'thumb')}_thumbframe.jpg")
                    import subprocess as _sp
                    _sp.run(
                        ['ffmpeg', '-y', '-loglevel', 'error',
                         '-ss', str(_sec), '-i', str(_local_vid),
                         '-frames:v', '1', '-q:v', '2', str(_frame_out)],
                        check=True, capture_output=True, timeout=60)
                    if _frame_out.is_file() and _frame_out.stat().st_size > 0:
                        thumb_frame_path = str(_frame_out)
            except Exception as _te:
                # Non-critical — the time/text still ship; Tool 2 can grab
                # the frame itself from the video if needed.
                print(f"[thumbnail] frame extract failed: {_te}")

        # ── Thumbnail styling prompt: recreate the ORIGINAL cover, better ──
        # Only when the source has a proper published thumbnail; otherwise
        # these stay empty and Tool 2 uses its frame-grab + text settings.
        thumb_prompt = ""
        thumb_ref = ""
        try:
            from core.script_generator import ScriptGenerator as _SG
            _api_keys = getattr(self, '_saved_api_keys', [])
            _sa_path = self.sa_path_var.get().strip()
            _gen = None
            if _sa_path and not _api_keys:
                _gen = _SG(service_account_path=_sa_path)
            elif _api_keys:
                _gen = _SG(api_keys=_api_keys)
            if _gen is not None:
                # Reconstruct a YouTube URL from the video ID so
                # fetch_original_thumbnail can download the cover even
                # when the user uploaded a local file (not a YouTube URL).
                _thumb_url = url
                if video_id and re.fullmatch(r'[A-Za-z0-9_-]{11}', video_id):
                    if not re.search(r'(?:v=|youtu\.be/|shorts/)' + re.escape(video_id), url):
                        _thumb_url = f"https://www.youtube.com/watch?v={video_id}"
                thumb_prompt, thumb_ref = self._cc_build_thumbnail_prompt(
                    _gen, _thumb_url, video_id, path,
                    title=title, niche=niche, lang=lang, vertical=_vertical)
        except Exception as _tpe:
            print(f"[thumbnail] prompt build failed: {_tpe}")

        rows = [{
            "Video ID": video_id,
            "Title": title,
            "Video URL": url,
            "Language": lang,
            "Niche": niche,
            "Custom Title": "",
            "Custom Script": full_script,
            "Word Count": total_words,
            "Montage Clips": clips_ref,
            "Commentary Spots": commentary_ref,
            "Vertical Format": vertical_flag,
            "Aspect Ratio": _aspect,
            "Story Cut": story_cut_flag,
            "Mute Original": mute_original_flag,
            "Thumbnail Time": thumb_ts,
            "Thumbnail Text": thumb_text,
            "Thumbnail Frame": thumb_frame_path,
            "Thumbnail Prompt": thumb_prompt,
            "Thumbnail Ref": thumb_ref,
            "Voiceover Style": self._cc_last_result.get("voiceover_style", ""),
            "Voiceover Speed (WPM)": self._cc_last_result.get("voiceover_speed", 0),
            "Hashtag 1": "",
            "Hashtag 2": "",
            "Date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        }]

        try:
            import pandas as pd
            from pathlib import Path as _P2
            existing_path = _P2(path)
            if existing_path.exists():
                try:
                    old_df = pd.read_excel(path)
                    old_rows = old_df.to_dict("records")
                    existing_ids = {str(r.get("Video ID", "")) for r in old_rows if r.get("Video ID")}
                    deduped_new = [r for r in rows if str(r.get("Video ID", "")) not in existing_ids]
                    rows = old_rows + deduped_new
                except Exception:
                    pass
            df = pd.DataFrame(rows)
            df.to_excel(path, index=False)
            messagebox.showinfo("Saved", f"Commentary script saved to:\n{path}")
            self.cc_status_label.config(text=f"✅ Saved to: {path}", fg="green")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {str(e)}")

    def create_settings_tab(self):
        """Settings tab - platform settings and preferences"""
        # Instagram Settings
        ig_section = tk.LabelFrame(self._settings_scroller.inner, text="📱 Instagram Settings",
                                  font=("Arial", 11, "bold"), padx=15, pady=15)
        ig_section.pack(fill=tk.X, padx=10, pady=10)

        status_frame = tk.Frame(ig_section)
        status_frame.pack(fill=tk.X, pady=5)

        tk.Label(status_frame, text="Authentication Status:",
                font=("Arial", 10, "bold")).pack(side=tk.LEFT)

        self.ig_status_label = tk.Label(status_frame, text="Not logged in",
                                       font=("Arial", 10), fg='red')
        self.ig_status_label.pack(side=tk.LEFT, padx=10)

        btn_frame = tk.Frame(ig_section)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="🔐 Login to Instagram",
                 command=self.show_instagram_login,
                 bg='#9C27B0', fg='white', font=("Arial", 10, "bold"),
                 width=20).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="🌐 Extract Cookies from Browser",
                 command=self.extract_browser_cookies,
                 bg='#FF9800', fg='white', font=("Arial", 10, "bold"),
                 width=25).pack(side=tk.LEFT, padx=5)

        tk.Label(ig_section, text="Note: Instagram login is required for profile scanning and metadata extraction",
                font=("Arial", 9, "italic"), fg='#666').pack()

        # Account Scraper Section
        scraper_section = tk.LabelFrame(self._settings_scroller.inner, text="📋 Account URL Scraper",
                                       font=("Arial", 10, "bold"), padx=10, pady=10)
        scraper_section.pack(fill=tk.X, pady=(15, 5))

        scraper_inner = tk.Frame(scraper_section)
        scraper_inner.pack(fill=tk.X)

        tk.Label(scraper_inner, text="Instagram Username:",
                font=("Arial", 9, "bold")).grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)

        self.ig_scrape_entry = tk.Entry(scraper_inner, font=("Arial", 10), width=25)
        self.ig_scrape_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        self.ig_scrape_entry.insert(0, "")

        tk.Button(scraper_inner,
                 text="🔍 Scrape All Videos",
                 command=self._scrape_instagram_account,
                 bg='#E1306C', fg='white', font=("Arial", 9, "bold"),
                 width=18).grid(row=0, column=2, padx=10, pady=5)

        tk.Label(scraper_section,
                text="Scrape ALL video URLs from an account — then load them for bulk processing",
                font=("Arial", 8, "italic"), fg='#888').pack(anchor=tk.W, padx=8, pady=(0, 3))

        # General Settings
        general_section = tk.LabelFrame(self._settings_scroller.inner, text="⚙️ General Settings",
                                       font=("Arial", 11, "bold"), padx=15, pady=15)
        general_section.pack(fill=tk.X, padx=10, pady=10)

        tk.Label(general_section, text="Default Settings:",
                font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=5)

        tk.Label(general_section, text="• Videos save to: channels/platform/channelname/videos/",
                font=("Arial", 9)).pack(anchor=tk.W, padx=20, pady=2)
        tk.Label(general_section, text="• Results save to: channels/platform/channelname/",
                font=("Arial", 9)).pack(anchor=tk.W, padx=20, pady=2)
        tk.Label(general_section, text="• Excel reports auto-generated after processing",
                font=("Arial", 9)).pack(anchor=tk.W, padx=20, pady=2)

    def create_log_tab(self):
        """Activity Log tab"""
        log_frame = tk.Frame(self.log_tab)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Log controls
        control_frame = tk.Frame(log_frame)
        control_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(control_frame, text="Activity Log",
                font=("Arial", 11, "bold")).pack(side=tk.LEFT)

        tk.Button(control_frame, text="Clear Log", command=self.clear_log,
                 bg='#9E9E9E', fg='white').pack(side=tk.RIGHT, padx=5)

        # Log text area
        self.log_text = scrolledtext.ScrolledText(log_frame, height=25,
                                                  bg='#f5f5f5', fg='#333',
                                                  font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def create_status_bar(self):
        """Status bar at bottom"""
        self.status_bar = tk.Frame(self.root, bg='#34495e', height=30)
        self.status_bar.pack(fill=tk.X, side=tk.BOTTOM)
        self.status_bar.pack_propagate(False)

        self.status_label = tk.Label(self.status_bar, text="Ready",
                                     bg='#34495e', fg='white',
                                     font=("Arial", 9), anchor=tk.W)
        self.status_label.pack(side=tk.LEFT, padx=10)

    def update_feature_status(self):
        """Update the feature status label"""
        download = self.features['download_video'].get()
        download_audio = self.features['download_audio'].get()
        ocr = self.features['extract_ocr'].get()
        speech = self.features['extract_speech'].get()
        captions = self.features['extract_captions'].get()
        metadata = self.features['extract_metadata'].get()

        if captions and not download and not download_audio and not ocr and not speech:
            self.feature_status.config(
                text="✅ YouTube Captions mode: No download needed",
                fg='green'
            )
        elif ocr and not download and not download_audio:
            self.feature_status.config(
                text="⚠️ Warning: OCR requires video download",
                fg='orange'
            )
        elif not any([ocr, speech, metadata, captions]):
            self.feature_status.config(
                text="⚠️ Select at least one extraction feature",
                fg='red'
            )
        elif download and ocr and speech and metadata:
            self.feature_status.config(
                text="✅ Full extraction mode: All features enabled",
                fg='green'
            )
        elif download_audio and not download:
            self.feature_status.config(
                text="🎵 Audio-only mode: Download audio + speech transcription",
                fg='blue'
            )
        elif not download and not download_audio:
            self.feature_status.config(
                text="📝 Metadata-only mode: Fast extraction without downloads",
                fg='blue'
            )
        else:
            enabled = []
            if ocr: enabled.append("OCR")
            if speech: enabled.append("Speech")
            if metadata: enabled.append("Metadata")
            self.feature_status.config(
                text=f"Enabled: {', '.join(enabled)}",
                fg='green'
            )

    def _on_gemini_model_change(self, event=None):
        """Update ScriptGenerator MODEL_NAME when Gemini model dropdown changes."""
        model = self.gemini_model.get()
        try:
            from core.script_generator import ScriptGenerator
            ScriptGenerator.MODEL_NAME = model
            self.log(f"🤖 Gemini model set to: {model}")
        except Exception as e:
            self.log(f"⚠️ Failed to set Gemini model: {e}")

    def _on_whisper_model_change(self, event=None):
        """Update WHISPER_MODEL when Whisper model dropdown changes."""
        model = self.whisper_model.get()
        try:
            import config as cfg
            cfg.WHISPER_MODEL = model
            # Also update the extractor module's local import reference
            from core import extractor as ext
            ext.WHISPER_MODEL = model
            # Force whisper to reload on next transcription
            self.processor.extractor.whisper_model = None
            self.log(f"🎙️ Whisper model set to: {model} (will reload on next transcription)")
        except Exception as e:
            self.log(f"⚠️ Failed to set Whisper model: {e}")

    def log(self, message, replace_last=False):
        """Add message to log, optionally replacing the last line"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        if replace_last:
            # Remove the last line so we overwrite it
            last_start = self.log_text.index("end-1l linestart")
            last_end = self.log_text.index("end-1c")
            # Don't delete if it's the very first line (nothing to replace)
            if last_start != "1.0":
                self.log_text.delete(last_start, last_end)
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.status_label.config(text=message[:80])

    def clear_log(self):
        """Clear the log"""
        self.log_text.delete(1.0, tk.END)

    def stop_process(self):
        """Stop processing"""
        self.stop_processing = True
        self.log("🛑 Stop requested - will stop after current video...")

    def browse_url_file(self):
        """Browse and load URLs from file"""
        file_path = filedialog.askopenfilename(
            title="Select URL List File",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()

            urls = []
            for line in content.split('\n'):
                line = line.strip()
                if line and not line.startswith('#'):
                    if ',' in line:
                        urls.extend([u.strip() for u in line.split(',') if u.strip()])
                    else:
                        urls.append(line)

            if not urls:
                messagebox.showwarning("Empty File", "No URLs found in the file.")
                return

            # Detect channel folder from path
            file_path_obj = Path(file_path)
            if 'channels' in file_path_obj.parts:
                try:
                    channels_idx = file_path_obj.parts.index('channels')
                    if len(file_path_obj.parts) > channels_idx + 2:
                        platform = file_path_obj.parts[channels_idx + 1]
                        channel_name = file_path_obj.parts[channels_idx + 2]
                        channel_folder = file_path_obj.parent

                        self.processor.current_channel_folder = channel_folder
                        self.log(f"📁 Detected channel folder: {platform}/{channel_name}")
                except (ValueError, IndexError):
                    pass

            # Clear and insert into Text widget (not Entry)
            self.input_text.delete("1.0", tk.END)
            self.input_text.insert("1.0", '\n'.join(urls))  # Use newlines instead of commas
            self.log(f"📁 Loaded {len(urls)} URL(s) from file: {file_path_obj.name}")

        except Exception as e:
            messagebox.showerror("File Error", f"Failed to read file:\n{str(e)}")
            self.log(f"❌ Failed to load URL file: {str(e)}")

    def browse_folder(self):
        """Browse folder for local videos"""
        folder_path = filedialog.askdirectory(title="Select Folder with Downloaded Videos")

        if not folder_path:
            return

        self.log(f"📁 Selected folder: {folder_path}")

        self.stop_processing = False
        self.progress_label.config(text="Processing folder...")

        thread = threading.Thread(target=self._process_folder_thread, args=(folder_path,))
        thread.daemon = True
        thread.start()

    def process_input(self):
        """Process URLs from input"""
        # Get text from Text widget (not Entry widget)
        url_input = self.input_text.get("1.0", tk.END).strip()
        if not url_input:
            messagebox.showwarning("Input Required", "Please enter a URL, channel, or video IDs")
            return

        platform = self.platform_var.get()

        # Auto-detect platform from the first URL (overrides dropdown)
        first_url = url_input.split('\n')[0].split(',')[0].strip().lower()
        if any(d in first_url for d in ('rednote.com', 'xiaohongshu.com')):
            platform = 'xiaohongshu'
        elif 'tiktok.com' in first_url:
            platform = 'tiktok'
        elif 'instagram.com' in first_url:
            platform = 'instagram'
        elif 'facebook.com' in first_url or 'fb.com' in first_url:
            platform = 'facebook'
        elif 'threads.com' in first_url or 'threads.net' in first_url:
            platform = 'threads'
        elif 'bilibili.com' in first_url:
            platform = 'bilibili'
        elif 'youtube.com' in first_url or 'youtu.be' in first_url:
            platform = 'youtube'
        elif 'pornhub.com' in first_url or 'xvideos.com' in first_url or 'hardgif.com' in first_url:
            platform = 'other'

        self.stop_processing = False
        self.progress_label.config(text="Starting processing...")
        self.log("Starting processing...")

        thread = threading.Thread(target=self._process_thread, args=(url_input, platform))
        thread.daemon = True
        thread.start()

    def _process_thread(self, url_input, platform):
        """Processing thread"""
        failed_count = 0
        success_count = 0

        try:
            # Reset channel folder for each new run so it doesn't carry over
            # from a previous platform's downloads (e.g. youtube → other)
            self.processor.current_channel_folder = None
            urls = self.processor.parse_input(url_input, platform)

            if not urls:
                self.log("❌ No valid URLs found")
                self.progress_label.config(text="No URLs to process")
                return

            total = len(urls)
            self.log(f"📋 Found {total} video(s) to process")

            # Get feature selections
            download_video = self.features['download_video'].get()
            download_audio = self.features['download_audio'].get()
            voice_only = self.features['voice_only'].get()
            extract_ocr = self.features['extract_ocr'].get()
            extract_speech = self.features['extract_speech'].get()
            extract_captions = self.features['extract_captions'].get()
            audio_format = self.audio_format.get()
            video_quality = self.video_quality.get()
            force_reprocess = self.features['force_reprocess'].get()

            # Track skipped URLs
            skipped_urls = []
            failed_urls = []

            for idx, url in enumerate(urls, 1):
                if self.stop_processing:
                    self.log("❌ Processing stopped by user")
                    break

                self.log(f"\n▶️ [{idx}/{total}] {url}")
                self.progress_var.set((idx / total) * 100)
                self.progress_label.config(text=f"Processing {idx}/{total}")

                try:
                    # Check if already processed (unless force reprocess is enabled)
                    if not force_reprocess and self.processor.db.is_processed(url):
                        self.log(f"⏭️  Already processed, skipping")
                        skipped_urls.append(url)
                        success_count += 1
                        continue

                    result = self.processor.process_video(
                        url, platform, self.log,
                        download_video=download_video,
                        download_audio=download_audio,
                        audio_format=audio_format,
                        video_quality=video_quality,
                        voice_only=voice_only,
                        extract_ocr=extract_ocr,
                        extract_speech=extract_speech,
                        extract_captions=extract_captions,
                        force_reprocess=force_reprocess
                    )

                    if result == "skipped":
                        skipped_urls.append(url)
                    else:
                        success_count += 1
                except Exception as e:
                    failed_count += 1
                    failed_urls.append((url, str(e)))
                    self.log(f"❌ ERROR: {str(e)}")
                    import traceback
                    self.log(f"   Traceback: {traceback.format_exc()[:500]}")  # Limit traceback length

            # Summary
            self.log(f"\n{'='*60}")
            self.log(f"📊 PROCESSING SUMMARY")
            self.log(f"{'='*60}")
            self.log(f"✅ Processed: {success_count}/{total}")
            self.log(f"⏭️  Skipped: {len(skipped_urls)}/{total}")
            self.log(f"❌ Failed: {failed_count}/{total}")

            if skipped_urls:
                self.log(f"\n⏭️  Skipped URLs ({len(skipped_urls)}):")
                for url in skipped_urls[:10]:  # Show first 10
                    self.log(f"   • {url}")
                if len(skipped_urls) > 10:
                    self.log(f"   ... and {len(skipped_urls) - 10} more")

            if failed_urls:
                self.log(f"\n❌ Failed URLs ({len(failed_urls)}):")
                for url, error in failed_urls[:10]:  # Show first 10
                    self.log(f"   • {url}")
                    self.log(f"     Error: {error[:100]}")
                if len(failed_urls) > 10:
                    self.log(f"   ... and {len(failed_urls) - 10} more")

            self.log(f"{'='*60}\n")
            self.progress_var.set(0)
            self.progress_label.config(text="Complete!")

            # Auto-generate Excel if enabled
            if success_count > 0 and self.features['auto_excel'].get():
                self.log("📊 Generating Excel report...")
                try:
                    excel_path = self.processor.exporter.generate_excel_report(
                        channel_folder=self.processor.current_channel_folder
                    )
                    if excel_path:
                        self.log(f"✅ Excel report created: {excel_path}")
                except Exception as e:
                    self.log(f"⚠️  Excel generation failed: {str(e)}")

        except Exception as e:
            self.log(f"❌ Fatal error: {str(e)}")
        finally:
            self.progress_label.config(text="Ready")

    def _process_folder_thread(self, folder_path):
        """Process folder thread"""
        from pathlib import Path

        failed_count = 0
        success_count = 0

        try:
            video_extensions = ['.mp4', '.avi', '.mov', '.mkv', '.flv', '.wmv', '.webm']
            video_files = []

            for ext in video_extensions:
                video_files.extend(Path(folder_path).glob(f'*{ext}'))

            if not video_files:
                self.log("❌ No video files found in folder")
                return

            total = len(video_files)
            self.log(f"Found {total} video file(s) to process")

            for idx, video_path in enumerate(video_files, 1):
                if self.stop_processing:
                    self.log("❌ Processing stopped by user")
                    break

                self.log(f"Processing {idx}/{total}: {video_path.name}")
                self.progress_var.set((idx / total) * 100)

                try:
                    result = self.processor.process_local_video(str(video_path), self.log)
                    if result != "skipped":
                        success_count += 1
                except Exception as e:
                    failed_count += 1
                    self.log(f"❌ Failed: {str(e)}")

            self.log(f"✅ Processing complete: {success_count} succeeded, {failed_count} failed")
            self.progress_var.set(0)

            if success_count > 0 and self.features['auto_excel'].get():
                self.log("📊 Generating Excel report...")
                try:
                    excel_path = self.processor.exporter.generate_excel_report()
                    if excel_path:
                        self.log(f"✅ Excel report created: {excel_path}")
                except Exception as e:
                    self.log(f"⚠️  Excel generation failed: {str(e)}")

        except Exception as e:
            self.log(f"❌ Fatal error: {str(e)}")
        finally:
            self.progress_label.config(text="Ready")

    def toggle_all_columns(self, select):
        """Select or deselect all metadata columns"""
        for var in self.metadata_columns.values():
            var.set(select)

    def start_metadata_scan(self):
        """Start metadata scan from Metadata tab"""
        url = self.metadata_url.get().strip()
        if not url:
            messagebox.showwarning("Input Required", "Please enter a channel/playlist/profile URL")
            return

        # Switch to log tab to show progress
        self.notebook.select(self.log_tab)

        platform = self.platform_var.get()

        # Auto-detect platform from URL if it clearly matches a known platform.
        # This overrides the dropdown so users don't have to manually switch
        # when scanning xiaohongshu/rednote, tiktok, etc.
        url_lower = url.lower()
        if 'rednote.com' in url_lower or 'xiaohongshu.com' in url_lower:
            platform = 'xiaohongshu'
        elif 'tiktok.com' in url_lower:
            platform = 'tiktok'
        elif 'instagram.com' in url_lower:
            platform = 'instagram'
        elif 'facebook.com' in url_lower or 'fb.com' in url_lower or 'fb.watch' in url_lower:
            platform = 'facebook'
        elif 'threads.com' in url_lower or 'threads.net' in url_lower:
            platform = 'threads'
        elif 'bilibili.com' in url_lower:
            platform = 'bilibili'
        elif 'youtube.com' in url_lower or 'youtu.be' in url_lower:
            platform = 'youtube'
        elif 'pornhub.com' in url_lower or 'xvideos.com' in url_lower or 'hardgif.com' in url_lower:
            platform = 'other'

        # Platform-specific checks
        if platform == 'instagram':
            username_file = Path(__file__).parent.parent / "data" / "ig_username.txt"
            if not username_file.exists():
                messagebox.showwarning(
                    "Instagram Login Required",
                    "Please login via Settings tab first"
                )
                return

        self.log("🔍 Starting metadata scan...")

        thread = threading.Thread(target=self._metadata_scan_thread,
                                 args=(url, platform))
        thread.daemon = True
        thread.start()

    def _metadata_scan_thread(self, url_input, platform):
        """Metadata scan thread"""
        from core.metadata_scanner import MetadataScanner
        import re
        from config import BASE_DIR

        try:
            scanner = MetadataScanner()

            self.log(f"📡 Scanning: {url_input}")

            # Parse popularity filter values
            try:
                min_views = int(self.min_views_var.get().strip())
            except ValueError:
                min_views = 0
            try:
                top_n = int(self.top_n_var.get().strip())
            except ValueError:
                top_n = 0

            # Show what filter is being applied
            if min_views > 0 or top_n > 0:
                self.log(f"🔥 Popularity filter active: min_views={min_views:,}" +
                        (f", top_n={top_n}" if top_n > 0 else ""))

            if platform == 'youtube':
                result = scanner.scan_youtube_channel(
                    url_input,
                    filter_shorts=False,
                    max_videos=None,
                    progress_callback=self.log,
                    min_views=min_views,
                    top_n=top_n
                )
            elif platform == 'tiktok':
                result = scanner.scan_tiktok_profile(
                    url_input,
                    max_videos=None,
                    progress_callback=self.log
                )
            elif platform == 'instagram':
                result = scanner.scan_instagram_profile(
                    url_input,
                    max_videos=50,
                    progress_callback=self.log
                )
            elif platform == 'facebook':
                result = scanner.scan_facebook_page(
                    url_input,
                    max_videos=None,
                    progress_callback=self.log
                )
            elif platform == 'xiaohongshu':
                result = scanner.scan_xiaohongshu_profile(
                    url_input,
                    max_videos=None,  # Get all videos
                    progress_callback=self.log
                )
            else:
                raise Exception(f"Unsupported platform: {platform}")

            channel_name = result.get('channel_name', 'Unknown')
            videos = result.get('videos', [])
            self.log(f"✅ Found {len(videos)} videos from {channel_name}")

            # Auto-create channel folder structure
            safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', channel_name)
            safe_name = safe_name[:50]  # Limit length

            channel_folder = BASE_DIR / "channels" / platform / safe_name
            channel_folder.mkdir(parents=True, exist_ok=True)

            # Create subfolders
            videos_folder = channel_folder / "videos"
            reports_folder = channel_folder / "reports"
            videos_folder.mkdir(exist_ok=True)
            reports_folder.mkdir(exist_ok=True)

            self.log(f"📁 Auto-created folder: channels/{platform}/{safe_name}/")

            # Get selected columns from GUI
            selected_columns = {k: v.get() for k, v in self.metadata_columns.items()}

            # Export with channel name in filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            self.log("📊 Exporting to Excel...")
            excel_file = channel_folder / f"metadata_{safe_name}_{timestamp}.xlsx"
            scanner.export_to_excel([result], str(excel_file), selected_columns=selected_columns)
            self.log(f"✅ Excel created: {excel_file.name}")

            self.log("📝 Exporting URLs to TXT...")
            txt_file = channel_folder / f"urls_{safe_name}_{timestamp}.txt"
            scanner.export_urls_to_txt([result], str(txt_file))
            self.log(f"✅ URL list created: {txt_file.name}")

            self.log(f"{'='*50}")
            self.log(f"✅ Metadata scan complete!")
            self.log(f"📁 Results saved to: {channel_folder}")
            self.log(f"{'='*50}")

            # Ask to load URLs
            def ask_load():
                if messagebox.askyesno(
                    "Scan Complete",
                    f"Metadata scan complete!\n\n"
                    f"✅ Channel: {channel_name}\n"
                    f"🎥 {len(videos)} video(s) found\n"
                    f"📁 Saved to: channels/{platform}/{safe_name}/\n\n"
                    f"Load URLs for processing?",
                    parent=self.root
                ):
                    with open(txt_file, 'r', encoding='utf-8') as f:
                        urls = [line.strip() for line in f if line.strip()]

                    self.processor.current_channel_folder = channel_folder
                    self.input_text.delete("1.0", tk.END)
                    self.input_text.insert("1.0", '\n'.join(urls))
                    self.log(f"📥 Loaded {len(urls)} URLs into input field")

                    # Switch to Quick Process tab
                    self.notebook.select(0)

            self.root.after(100, ask_load)

        except Exception as e:
            self.log(f"❌ Metadata scan error: {str(e)}")

    def export_data(self):
        """Export data to Excel"""
        self.log("📊 Generating Excel report...")
        try:
            excel_path = self.processor.exporter.generate_excel_report()
            if excel_path:
                self.log(f"✅ Excel report created: {excel_path}")
                messagebox.showinfo("Export Complete", f"Excel report created:\n{excel_path}")
        except Exception as e:
            self.log(f"❌ Excel generation failed: {str(e)}")
            messagebox.showerror("Export Error", f"Failed to generate Excel report:\n{str(e)}")

    def check_instagram_auth(self):
        """Check Instagram authentication status.

        Checks ALL cookie files (legacy cookies.txt + data/cookies/*.txt)
        so the user sees how many accounts are available for rotation.
        """
        from platforms.instagram import ig_multi_status

        ms = ig_multi_status()
        valid = [c for c in ms["accounts"] if c["logged_in"]]
        total = len(ms["accounts"])

        if valid:
            count_str = f"{len(valid)} account{'s' if len(valid) > 1 else ''}"
            first_user = valid[0].get("user_id") or ""
            detail = f" ({count_str})" if len(valid) > 1 else ""
            display = first_user if first_user else "session"
            label = f"✅ Instagram: {display}{detail}"
            self.auth_label.config(text=label[:45])
            self.ig_status_label.config(text=label, fg='green')
            self.log(f"✅ Instagram: {count_str} authenticated — auto-rotate on failure")
        else:
            # Show why the first cookie failed as the reason
            first = ms["accounts"][0] if ms["accounts"] else {"reason": "no cookie files"}
            hint = first["reason"]
            label = f"❌ Instagram: {hint}"[:40]
            self.auth_label.config(text=label)
            self.ig_status_label.config(text=f"❌ {hint}", fg='red')
            self.log(f"Instagram: No valid session in {total} file(s) — {hint}")

    def _scrape_instagram_account(self):
        """Scrape ALL video URLs from an Instagram account using yt-dlp."""
        username = self.ig_scrape_entry.get().strip()
        if not username or username == "username":
            messagebox.showwarning("Input Required", "Enter an Instagram username first.", parent=self.root)
            return

        # Validate username
        import re
        if 'instagram.com' in username:
            m = re.search(r'instagram\.com/([^/?]+)', username)
            if m:
                username = m.group(1)
            else:
                messagebox.showerror("Invalid Input", "Could not extract username from that URL.", parent=self.root)
                return

        username = username.lstrip('@')
        if not username or username in ('reel', 'p', 'tv', 'stories', 'reels'):
            messagebox.showerror("Invalid Username", "That doesn't look like a valid Instagram profile.", parent=self.root)
            return

        def task():
            try:
                from platforms.instagram import scrape_instagram_account_urls
                urls = scrape_instagram_account_urls(
                    username,
                    progress_callback=self.log,
                )
                if not urls:
                    self.log("No URLs found.")
                    return

                # Write to a txt file
                from datetime import datetime
                from pathlib import Path
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                out_dir = Path(__file__).parent.parent / "channels" / "instagram" / username
                out_dir.mkdir(parents=True, exist_ok=True)
                txt_path = out_dir / f"urls_{username}_{ts}.txt"
                with open(txt_path, 'w', encoding='utf-8') as f:
                    f.write('\n'.join(urls))

                self.log(f"\n{'='*50}")
                self.log(f"✅ Scraped {len(urls)} URLs from @{username}")
                self.log(f"📁 Saved to: channels/instagram/{username}/urls_{username}_{ts}.txt")
                self.log(f"{'='*50}")

                # Load into input text
                def ask_load():
                    if messagebox.askyesno(
                        "Scrape Complete",
                        f"✅ Found {len(urls)} posts from @{username}\n\n"
                        f"Load them into the input field for bulk processing?",
                        parent=self.root
                    ):
                        self.input_text.delete("1.0", tk.END)
                        self.input_text.insert("1.0", '\n'.join(urls))
                        self.log(f"📥 Loaded {len(urls)} URLs into input field")
                        self.notebook.select(0)

                self.root.after(100, ask_load)

            except Exception as e:
                self.log(f"❌ Scrape failed: {str(e)}")
                messagebox.showerror("Scrape Failed", str(e), parent=self.root)

        import threading
        t = threading.Thread(target=task, daemon=True)
        t.start()
        self.log(f"🚀 Scraping @{username} in background...")

    def extract_browser_cookies(self):
        """Extract Instagram cookies from browser using yt-dlp"""
        from pathlib import Path

        data_dir = Path(__file__).parent.parent / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        cookies_path = data_dir / "cookies.txt"

        browsers = [
            ("chrome", "Google Chrome"),
            ("brave", "Brave"),
            ("edge", "Microsoft Edge"),
            ("firefox", "Firefox"),
        ]

        for browser_key, browser_name in browsers:
            try:
                self.log(f"🌐 Extracting cookies from {browser_name}...")
                # Use yt-dlp to extract cookies and save to our cookies.txt
                # This only runs the cookie extraction, not a real download
                import yt_dlp
                ydl_opts = {
                    'cookiesfrombrowser': (browser_key,),
                    'cookiefile': str(cookies_path),
                    'quiet': True,
                    'no_warnings': True,
                }
                with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                    # Extract info for Instagram homepage (no actual download)
                    ydl.extract_info('https://www.instagram.com/', download=False)

                # Check if Instagram cookies were written
                if cookies_path.exists() and cookies_path.stat().st_size > 0:
                    content = cookies_path.read_text(encoding='utf-8')
                    if 'sessionid' in content and '.instagram.com' in content:
                        self.log(f"✅ Instagram cookies extracted from {browser_name}!")
                        self.ig_status_label.config(
                            text=f"✅ Auth via {browser_name} cookies",
                            fg='green')
                        self.auth_label.config(
                            text=f"✅ Instagram: Auth ({browser_name})")
                        return
                self.log(f"⚠️  No Instagram session found in {browser_name}")
            except ImportError:
                self.log("❌ yt-dlp not installed")
                break
            except Exception as e:
                self.log(f"⚠️  {browser_name}: {str(e)[:80]}")

        self.log("❌ Could not extract Instagram cookies from any browser")
        self.log("   Make sure you're logged into instagram.com in one of your browsers")

    def show_instagram_login(self):
        """Show Instagram login dialog"""
        login_window = tk.Toplevel(self.root)
        login_window.title("Instagram Login")
        login_window.geometry("400x250")
        login_window.resizable(False, False)
        login_window.transient(self.root)
        login_window.grab_set()

        tk.Label(login_window, text="Instagram Authentication",
                font=("Arial", 14, "bold")).pack(pady=20)

        form_frame = tk.Frame(login_window, padx=20, pady=10)
        form_frame.pack()

        tk.Label(form_frame, text="Username:").grid(row=0, column=0, sticky=tk.W, pady=5)
        username_entry = tk.Entry(form_frame, width=30)
        username_entry.grid(row=0, column=1, pady=5)

        tk.Label(form_frame, text="Password:").grid(row=1, column=0, sticky=tk.W, pady=5)
        password_entry = tk.Entry(form_frame, width=30, show="*")
        password_entry.grid(row=1, column=1, pady=5)

        status_label = tk.Label(login_window, text="", fg="red")
        status_label.pack(pady=5)

        def do_login():
            username = username_entry.get().strip()
            password = password_entry.get().strip()

            if not username or not password:
                status_label.config(text="Please enter both username and password")
                return

            status_label.config(text="Logging in...", fg="blue")
            login_window.update()

            try:
                scraper = self.processor.scrapers['instagram']
                scraper.login(username, password)

                status_label.config(text="✅ Login successful!", fg="green")
                self.check_instagram_auth()
                self.log("✅ Instagram login successful")

                login_window.after(1000, login_window.destroy)
            except Exception as e:
                status_label.config(text=f"Login failed: {str(e)}", fg="red")
                self.log(f"❌ Instagram login failed: {str(e)}")

        button_frame = tk.Frame(login_window)
        button_frame.pack(pady=10)

        tk.Button(button_frame, text="Login", command=do_login,
                 bg='#4CAF50', fg='white', width=12).pack(side=tk.LEFT, padx=5)

        tk.Button(button_frame, text="Cancel", command=login_window.destroy,
                 bg='#f44336', fg='white', width=12).pack(side=tk.LEFT, padx=5)

        username_entry.focus()

    # ─── Script Generator methods ────────────────────────────────────────

    def _start_script_generation(self):
        """Start the script generation thread — first ask user to pick an Excel file"""
        api_keys = getattr(self, '_saved_api_keys', [])
        sa_path = self.sa_path_var.get().strip()
        if not api_keys and not sa_path:
            messagebox.showwarning("API Key Required",
                                   "Please add at least one Gemini API key OR select a Service Account JSON file.\n"
                                   "Get a key at: https://aistudio.google.com/apikey")
            return

        # Ask user to select the Excel file with downloaded video data
        # Default to the last-used directory, or the channels folder
        default_dir = self._last_excel_dir
        if not default_dir:
            # Fall back to the channels/youtube directory
            base = Path(__file__).resolve().parent.parent
            default_dir = str(base / "channels")
        excel_path = filedialog.askopenfilename(
            title="Select Excel File with Video Data",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=default_dir,
        )
        if not excel_path:
            self.log("⏹️ Script generation cancelled — no Excel file selected.")
            return
        # Remember this directory for next time
        self._last_excel_dir = str(Path(excel_path).parent)

        # Switch to log tab
        self.notebook.select(self.log_tab)
        self.log(f"🚀 Starting script generation from: {excel_path}")
        lang = self.script_lang_var.get().lower().strip()
        self.log(f"🌐 Target language: {lang.capitalize()}")
        self.log(f"📋 {len(api_keys)} API key(s) available (auto-fallback on quota exhaustion)")

        thread = threading.Thread(
            target=self._script_generation_thread,
            args=(excel_path, lang),
            daemon=True,
        )
        thread.start()

    def _script_generation_thread(self, excel_path, lang):
        """Background thread for script generation — reads from Excel, saves to lang-specific columns"""
        import pandas as pd
        from pathlib import Path
        from core.script_generator import ScriptGenerator
        from core.exporter import DataExporter
        from openpyxl import load_workbook

        api_keys = getattr(self, '_saved_api_keys', [])
        sa_path = self.sa_path_var.get().strip()
        niche = self.niche_var.get()

        try:
            prompts_path = self.get_prompts_config_path()

            # Init the generator
            if sa_path and not api_keys:
                gen = ScriptGenerator(service_account_path=sa_path,
                                      prompts_config_path=prompts_path)
                self.log("🔑 Using service account authentication")
            elif api_keys:
                gen = ScriptGenerator(api_keys=api_keys,
                                      prompts_config_path=prompts_path)
                self.log(f"🔑 {len(api_keys)} key(s) loaded. Active: {gen.active_key_label}")
            else:
                self.log("❌ No API key or service account configured.")
                return

            # Set the active prompt template
            prompt_slug = self._get_active_prompt_slug()
            if gen.get_prompt_data(prompt_slug):
                gen.set_active_prompt(prompt_slug)
                self.log(f"📝 Script style: {gen.get_active_prompt_name()}")

            # ── Read the Excel file ─────────────────────────────────
            df = pd.read_excel(excel_path)
            # Normalize column names (strip whitespace)
            df.columns = df.columns.str.strip()
            cols = list(df.columns)
            self.log(f"📊 Loaded Excel: {len(df)} rows, columns: {', '.join(cols[:8])}{'…' if len(cols) > 8 else ''}")

            # Map known column variations → internal names
            col_map = {}
            for raw_col in cols:
                cl = raw_col.lower().strip()
                if cl in ("video id", "video_id"):
                    col_map["video_id"] = raw_col
                elif cl in ("title", "video title"):
                    col_map["title"] = raw_col
                elif cl in ("speech transcript", "transcript", "speech_text", "speech text"):
                    col_map["speech_text"] = raw_col
                elif cl in ("duration (sec)", "video_duration", "duration sec", "duration"):
                    col_map["video_duration"] = raw_col
                elif cl in ("description", "video description"):
                    col_map["description"] = raw_col
                elif cl in ("captions",):
                    col_map["captions"] = raw_col
                elif cl in ("custom script", "generated_script"):
                    col_map["generated_script"] = raw_col
                elif cl in ("custom title", "suggested_title", "suggested title"):
                    col_map["suggested_title"] = raw_col

            missing = [k for k in ("video_id", "speech_text") if k not in col_map]
            if missing:
                self.log(f"❌ Excel missing required columns: {missing}. Found: {cols}")
                self.root.after(100, lambda: messagebox.showerror(
                    "Missing Columns",
                    f"Excel needs a 'Video ID' and 'Speech Transcript' column.\n\nFound: {cols}"
                ))
                return

            # Filter rows to only those needing scripts
            pending_rows = []
            for idx, row in df.iterrows():
                transcript = str(row.get(col_map["speech_text"], "") or "")
                if not transcript.strip():
                    continue
                # Skip if already has a generated script
                if col_map.get("generated_script"):
                    raw_val = row.get(col_map["generated_script"], "")
                    # Handle NaN from empty Excel cells (NaN != itself)
                    if isinstance(raw_val, float) and raw_val != raw_val:
                        existing = ""
                    else:
                        existing = str(raw_val).strip()
                    if existing:
                        continue
                pending_rows.append((idx, row))

            if not pending_rows:
                self.log("✅ No videos pending script generation (all already have scripts or no transcripts).")
                return

            self.log(f"📝 Found {len(pending_rows)} video(s) needing scripts.")
            updated_count = 0
            error_count = 0
            results_list = []

            for i, (df_idx, row) in enumerate(pending_rows, 1):
                if self.stop_processing:
                    self.log("⏹️ Script generation cancelled by user.")
                    break

                video_id = str(row.get(col_map["video_id"], "?"))
                title = str(row.get(col_map.get("title", ""), video_id)) if col_map.get("title") else video_id
                transcript = str(row.get(col_map["speech_text"], "") or "")
                duration = float(row.get(col_map.get("video_duration", ""), 0) or 0) if col_map.get("video_duration") else 0

                # Target word count based on video duration (~160 WPM)
                # so the script fills the clip naturally without overflow
                target_wpm = 110 if getattr(self, 'tts_engine_var', None) is not None and self.tts_engine_var.get() == "qwen3" else 160
                word_count = int(duration * target_wpm / 60) if duration > 0 else len(transcript.split())
                wpm = round(word_count / (duration / 60), 1) if duration > 0 else 0.0

                self.log(f"[{i}/{len(pending_rows)}] Generating script for {video_id} "
                        f"(key {gen.active_key_label})...")

                result = gen.generate_script_with_retry(
                    transcript=transcript,
                    title=title,
                    duration=duration,
                    word_count=word_count,
                    wpm=wpm,
                    niche_angle=niche,
                    language=lang,
                    max_retries=2,
                )

                if "error" in result:
                    self.log(f"❌ [{i}/{len(pending_rows)}] {video_id}: {result['error']}")
                    error_count += 1
                    # If all API keys exhausted, stop immediately — don't waste time on remaining videos
                    if gen.all_keys_exhausted:
                        self.log("⏹️ All API keys exhausted — stopping early. Add more keys or wait for quota reset.")
                        break
                    continue

                gen_script = result.get("script", "")
                gen_wc = result.get("generated_word_count", 0)
                sug_title = result.get("suggested_title", "")
                h1 = result.get("hashtag_1", "")
                h2 = result.get("hashtag_2", "")
                # Build a short description — title + hashtags in the target language
                gen_description = sug_title
                if h1 or h2:
                    gen_description += "\n\n"
                    if h1:
                        gen_description += f"#{h1} "
                    if h2:
                        gen_description += f"#{h2}"

                # Store collected data — will write to per-language file at end
                results_list.append({
                    "Video ID": video_id,
                    "Title": title,
                    "Video Title": sug_title,
                    "Description": (str(row.get(col_map.get("description", ""), "") or "")
                        if col_map.get("description") else ""),
                    "Speech Transcript": (str(row.get(col_map.get("speech_text", ""), "") or "")),
                    "Captions": (str(row.get(col_map.get("captions", ""), "") or "")
                        if col_map.get("captions") else ""),
                    "Custom Title": re.sub(r'#\S+\s*', '', sug_title).strip(),
                    "Custom Description": gen_description,
                    "Custom Script": gen_script,
                    "Voiceover Style": result.get("voiceover_style", ""),
                    "Voiceover Speed (WPM)": result.get("voiceover_speed", ""),
                })

                updated_count += 1
                self.log(f"✅ [{i}/{len(pending_rows)}] {video_id} — {gen_wc} words ({lang})")

            self.log(f"{'='*50}")
            self.log(f"✅ Script generation complete: {updated_count} updated, {error_count} errors")

            # ── Build per-language Excel file (append mode) ────────
            if results_list:
                # Determine channel name from source Excel parent folder
                src = Path(excel_path)
                channel_name = src.parent.name
                safe_channel = "".join(c if c.isalnum() or c in " -_" else "_" for c in channel_name)[:40]
                lang_name = lang.capitalize()
                out_filename = f"{lang_name} {safe_channel}.xlsx"
                out_path = src.parent / out_filename

                # Build set of existing Video IDs to avoid duplicates
                existing_ids = set()
                new_rows = []
                if out_path.exists():
                    try:
                        old_df = pd.read_excel(out_path)
                        for _, old_row in old_df.iterrows():
                            vid = old_row.get("Video ID", "")
                            if vid:
                                existing_ids.add(str(vid))
                        for item in results_list:
                            if str(item.get("Video ID", "")) not in existing_ids:
                                new_rows.append(item)
                    except Exception:
                        new_rows = results_list
                else:
                    new_rows = results_list

                if not new_rows:
                    self.log(f"📁 All {len(existing_ids)} videos already in {out_filename} — nothing new")
                else:
                    if existing_ids:
                        # Append to existing file
                        wb = load_workbook(out_path)
                        ws = wb.active
                        headers = list(new_rows[0].keys())
                        start_row = ws.max_row + 1
                        for row_offset, item in enumerate(new_rows):
                            for col_idx, key in enumerate(headers, 1):
                                ws.cell(row=start_row + row_offset, column=col_idx,
                                        value=item.get(key, ""))
                        # Apply text wrapping to new rows
                        for row_cells in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
                            for cell in row_cells:
                                cell.alignment = Alignment(wrap_text=True, vertical="top")
                        wb.save(out_path)
                        wb.close()
                    else:
                        # Fresh file
                        out_df = pd.DataFrame(new_rows)
                        out_df.to_excel(out_path, index=False)
                        wb = load_workbook(out_path)
                        ws = wb.active
                        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
                            for cell in row_cells:
                                cell.alignment = Alignment(wrap_text=True, vertical="top")
                        wb.save(out_path)
                        wb.close()

                    self.log(f"📁 Per-language Excel saved/appended: {out_filename} ({len(new_rows)} new)")
                self.root.after(100, lambda p=str(out_path), c=updated_count, lg=lang: messagebox.showinfo(
                    "Script Generation Complete",
                    f"✅ {c} scripts generated ({lg.capitalize()})\n\n📁 Saved to:\n{p}"
                ))
            else:
                self.log("⚠️ No results to save.")

        except Exception as e:
            self.log(f"❌ Script generation error: {str(e)}")
            import traceback
            self.log(traceback.format_exc())

    # ── Script Style / Prompt Template Management ─────────────────

    # Prompt slugs valid for each mode (single source of truth)
    REWRITE_MODE_SLUGS = [
        "simple_rewrite",
        "movies_commentary",
        "replace_existing_narration",
        "dialogue_to_narration",
        "educational_facts",
        "courtroom_legal",
        "heartwarming",
    ]
    WRITE_STORY_MODE_SLUGS = [
        "movies_with_dialogue",
        "movies_with_voiceover",
        "movies_voiceover_timed",
        "courtroom_legal",
        "heartwarming",
        "cctv_surveillance",
		"ocean_mysteries",
        "funny_compilations",
    ]

    # Per-preset help: shown in the "About this preset" panel.
    # Keyed by slug. Each entry: when to use + what it does + a short example.
    PRESET_HELP = {
        # ── Rewrite mode (transcript-based) ──────────────────────
        "simple_rewrite": (
            "WHEN: The video's existing script is fine but a bit flat.\n"
            "DOES: Rewrites your transcript into a punchier version with a strong "
            "hook — keeps the SAME length and meaning, just better wording.\n"
            "EXAMPLE: \"Today we look at the case\" → \"This case shocked everyone "
            "in the courtroom — here's why.\""
        ),
        "movies_commentary": (
            "WHEN: You want a dramatic recap of a movie/story clip.\n"
            "DOES: Turns the transcript into cinematic third-person narration that "
            "recaps the events with tension and pacing.\n"
            "EXAMPLE: A fight scene becomes \"He had one chance left — and he knew it.\""
        ),
        "replace_existing_narration": (
            "WHEN: The clip ALREADY has a voiceover/narration and you want a better one.\n"
            "DOES: Reads the existing narration transcript and writes a stronger "
            "cinematic replacement, matching the original length.\n"
            "EXAMPLE: Swaps a dull voiceover for a gripping one of the same duration."
        ),
        "dialogue_to_narration": (
            "WHEN: You have a TRANSCRIPT of a dialogue-only clip (no narration yet).\n"
            "DOES: Converts character dialogue into flowing third-person story "
            "narration. (For working from the actual VIDEO instead, use Write-Story mode.)\n"
            "EXAMPLE: \"You can't do this!\" / \"Watch me.\" → \"She begged him to "
            "stop — but he had already made up his mind.\""
        ),
        "educational_facts": (
            "WHEN: How-to, science facts, survival tips, or any explainer content.\n"
            "DOES: Rewrites the transcript into clear, engaging educational narration "
            "that keeps viewers watching.\n"
            "EXAMPLE: \"Water boils at 100C\" → \"Most people get this wrong — here's "
            "what actually happens at 100 degrees.\""
        ),
        "courtroom_legal": (
            "WHEN: Courtroom / legal case clips — trials, disputes, verdicts.\n"
            "DOES: Turns the transcript into a compelling legal-story narration that "
            "explains the case as a human drama.\n"
            "EXAMPLE: \"The defendant pleaded not guilty\" → \"He looked the judge in "
            "the eye and said the two words that would change everything.\""
        ),
        "heartwarming": (
            "WHEN: Wholesome human/animal interest — rescues, kind acts, reunions.\n"
            "DOES: Rewrites the transcript into warm, emotional narration that makes "
            "viewers feel something.\n"
            "EXAMPLE: A dog rescue becomes \"After 3 years lost, he finally heard the "
            "voice he'd never forgotten.\""
        ),
        # ── Write Story mode (Gemini watches the video) ──────────
        "movies_with_dialogue": (
            "WHEN: A movie/TV clip that has ONLY dialogue and no narration.\n"
            "DOES: Gemini WATCHES the video, understands the theme, and writes a "
            "complete scene-by-scene story script that replaces all audio.\n"
            "EXAMPLE: Watches a tense argument scene → writes a full narrated story "
            "around it from scratch."
        ),
        "movies_with_voiceover": (
            "WHEN: A clip that ALREADY has a voiceover, and you want to replace it.\n"
            "DOES: Gemini watches the video and writes a new voiceover script matching "
            "the clip's exact word count so it fits the timing.\n"
            "EXAMPLE: Replaces a creator's voiceover with your own version, same length."
        ),
        "movies_voiceover_timed": (
            "WHEN: You need the narration to line up PRECISELY with the video timing.\n"
            "DOES: Gemini watches the video and writes a scene-by-scene script with "
            "timestamps, so each line matches what's on screen at that moment.\n"
            "EXAMPLE: \"[00:03] He opens the door... [00:07] and freezes.\" — narration "
            "synced to the actual scene cuts."
        ),
        "cctv_surveillance": (
            "WHEN: CCTV / surveillance footage — crime, near-misses, hero moments, "
            "caught-red-handed, instant karma.\n"
            "DOES: Gemini finds the most shocking moment for a hook-sized intro clip, "
            "then narrates the WHOLE video in grounded news-anchor third person, "
            "ending on a freeze-frame CTA.\n"
            "EXAMPLE: \"00:00 | Nobody noticed what he was actually doing. | [hook]\" → "
            "full narration → \"...What would you have done? Comment below. | [cta]\""
        ),
        "ocean_mysteries": (
            "WHEN: Ocean/maritime compilation footage — storms, ships, dark seas, "
            "haunted vessels, near-disasters.\n"
            "DOES: Gemini watches the whole compilation, identifies every clip and the "
            "unifying theme, then narrates each scene with a two-line formula: "
            "what happens (action) + what it means (moral/lesson).\n"
            "EXAMPLE: \"A cargo ship vanishes behind a wave the size of a building. | [observe]\" → "
            "\"The ocean does not care how confident you were before you left the shore. | [lesson]\""
        ),
        "funny_compilations": (
            "WHEN: Funny fails/bloopers compilations — many short (~4s) clips of "
            "fails, pranks, animal chaos, silly moments.\n"
            "DOES: Opens with a warm 15-20s welcome + a light 'just for entertainment' "
            "disclaimer, then drops ONE short funny reaction per clip timed to the joke "
            "('oh come on!', 'what a shot!', a 'hahaha' only when earned) — reacts, "
            "never describes.\n"
            "EXAMPLE: \"00:00 | Welcome to the fun world — let's forget our stress and "
            "enjoy the funniest clips around the world. Quick note, this is all just for "
            "fun to refresh your mind. | [intro]\" → \"00:18 | Ohh, what a shot! | [react]\" → "
            "\"00:23 | Hahaha, don't do that again, baby! | [react]\""
        ),
    }

    def _preset_help_for_name(self, display_name, mode_slugs):
        """Return the help text for a preset given its display name + which mode."""
        try:
            import os, json
            base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            config_path = os.path.join(base, "data", "script_prompts.json")
            with open(config_path, "r", encoding="utf-8") as f:
                prompts = json.load(f).get("prompts", {})
            for slug in mode_slugs:
                info = prompts.get(slug)
                if info and info.get("name") == display_name:
                    return self.PRESET_HELP.get(
                        slug, "(No description available for this preset.)")
        except Exception:
            pass
        return "Select a preset to see what it does."


    def _refresh_script_styles(self):
        """Refresh the Rewrite-mode preset combobox from the prompts config.

        Only transcript-based prompts (REWRITE_MODE_SLUGS) are shown — video-only
        and Case-Commentary prompts are excluded so the list stays unambiguous.
        """
        try:
            import os, json
            base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            config_path = os.path.join(base, "data", "script_prompts.json")
            if os.path.isfile(config_path):
                with open(config_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                prompts = data.get("prompts", {})
                active = data.get("active_prompt", "movies_commentary")
                names = []
                active_name = None
                # Preserve REWRITE_MODE_SLUGS order, skip any missing slug
                for slug in self.REWRITE_MODE_SLUGS:
                    info = prompts.get(slug)
                    if not info:
                        continue
                    name = info.get("name", slug)
                    names.append(name)
                    if slug == active:
                        active_name = name
                self.script_style_combo["values"] = names
                # Default to the active prompt if it's in this mode, else first item
                if active_name:
                    self.script_style_var.set(active_name)
                elif names:
                    self.script_style_var.set(names[0])
                # Refresh the details panel for the now-selected preset
                self._update_rewrite_preset_help()
        except Exception:
            pass

    def _update_rewrite_preset_help(self):
        """Show the help text for the selected Rewrite-mode preset."""
        if not hasattr(self, "rewrite_help_label"):
            return
        text = self._preset_help_for_name(
            self.script_style_var.get(), self.REWRITE_MODE_SLUGS)
        self.rewrite_help_label.config(text=text)

    def _get_active_prompt_slug(self):
        """Resolve the currently selected style name back to a prompt slug"""
        try:
            import os, json
            base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            config_path = os.path.join(base, "data", "script_prompts.json")
            if os.path.isfile(config_path):
                with open(config_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                prompts = data.get("prompts", {})
                selected_name = self.script_style_var.get()
                for slug, info in prompts.items():
                    if info.get("name") == selected_name:
                        return slug
            return "movies_commentary"
        except Exception:
            return "movies_commentary"

    def _slug_for_name(self, display_name, mode_slugs):
        """Resolve a preset display name back to its slug, limited to one mode.

        Returns None if nothing matches (e.g. nothing selected yet).
        """
        try:
            import os, json
            base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            config_path = os.path.join(base, "data", "script_prompts.json")
            with open(config_path, "r", encoding="utf-8") as f:
                prompts = json.load(f).get("prompts", {})
            for slug in mode_slugs:
                info = prompts.get(slug)
                if info and info.get("name") == display_name:
                    return slug
        except Exception:
            pass
        return None

    def _open_prompt_manager(self, mode_slugs=None, preselect_slug=None,
                             prefer_master=False):
        """Open the prompt-template manager.

        When opened from a mode's "Manage Presets" button, ``mode_slugs`` limits
        the editor to that mode's presets and ``preselect_slug`` opens the one
        currently chosen in the dropdown. With no args, all presets are shown.

        ``prefer_master`` (Write Story mode) makes the editor read/write the
        master ``.txt`` file for slugs that have one, so edits actually reach
        Gemini instead of going to the ignored JSON copy.
        """
        PromptManagerDialog(self.root, self,
                            mode_slugs=mode_slugs,
                            preselect_slug=preselect_slug,
                            prefer_master=prefer_master)

    def get_prompts_config_path(self):
        """Return the path to script_prompts.json"""
        import os
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(base, "data", "script_prompts.json")


class PromptManagerDialog:
    """Tkinter dialog for managing script prompt templates"""

    def __init__(self, parent, dashboard, mode_slugs=None, preselect_slug=None,
                 prefer_master=False, category=None, on_close=None):
        self.parent = parent
        self.dashboard = dashboard
        self.config_path = dashboard.get_prompts_config_path()
        self.data = self._load()
        # When opened from a mode, only that mode's presets are listed.
        self.mode_slugs = list(mode_slugs) if mode_slugs else None
        self.preselect_slug = preselect_slug
        # Write Story mode: edit the master .txt file (what Gemini actually reads)
        # for slugs that have one, instead of the ignored JSON copy.
        self.prefer_master = prefer_master
        # Category mode (e.g. "case_commentary"): list only prompts whose slug
        # carries that prefix; new prompts are auto-prefixed so they join the set.
        self.category = category
        # Callback fired when the window closes (e.g. refresh a dropdown).
        self.on_close = on_close
        # Slug currently loaded in the editor, and whether it's a master .txt.
        self.current_slug = None
        self.editing_master = False
        # Slugs currently shown in the listbox, in display order (set by refresh).
        self.visible_slugs = []
        # Slugs created during this session (kept visible even when mode-filtered).
        self._session_added = []

        self.window = tk.Toplevel(parent)
        self.window.title("Case Commentary Niche Prompts" if self.category
                          else "Script Template Manager")
        self.window.geometry("750x550")
        self.window.transient(parent)
        self.window.grab_set()
        if self.on_close:
            self.window.protocol("WM_DELETE_WINDOW", self._handle_close)

        # ── Listbox on the left ─────────────────────────────
        main_frame = tk.Frame(self.window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        list_frame = tk.Frame(main_frame)
        list_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))

        tk.Label(list_frame, text="Script Styles:", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        self.listbox = tk.Listbox(list_frame, width=30, height=18,
                                   font=("Arial", 10))
        self.listbox.pack(side=tk.LEFT, fill=tk.Y)
        scrollbar = tk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox.config(yscrollcommand=scrollbar.set)
        self.listbox.bind("<<ListboxSelect>>", self._on_select)

        # ── Buttons between list and editor ─────────────────
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5)
        tk.Button(btn_frame, text="➕ Add New", font=("Arial", 9),
                  bg='#4CAF50', fg='white', width=12,
                  command=self._add_prompt).pack(pady=3)
        tk.Button(btn_frame, text="🗑️ Delete", font=("Arial", 9),
                  bg='#f44336', fg='white', width=12,
                  command=self._delete_prompt).pack(pady=3)
        tk.Button(btn_frame, text="📋 Duplicate", font=("Arial", 9),
                  bg='#FF9800', fg='white', width=12,
                  command=self._duplicate_prompt).pack(pady=3)

        # ── Editor on the right ─────────────────────────────
        editor_frame = tk.Frame(main_frame)
        editor_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Banner: tells the user which store this preset is saved to.
        self.source_banner = tk.Label(
            editor_frame, text="", font=("Arial", 8, "bold"),
            anchor=tk.W, justify=tk.LEFT, wraplength=420)
        self.source_banner.pack(fill=tk.X, pady=(0, 4))

        # Slug (internal key)
        slug_row = tk.Frame(editor_frame)
        slug_row.pack(fill=tk.X, pady=2)
        tk.Label(slug_row, text="Slug (internal key):", font=("Arial", 9, "bold"),
                 width=18, anchor=tk.W).pack(side=tk.LEFT)
        self.slug_var = tk.StringVar()
        self.slug_entry = tk.Entry(slug_row, textvariable=self.slug_var,
                                    font=("Arial", 9), width=25)
        self.slug_entry.pack(side=tk.LEFT, padx=5)

        # Name (display name)
        name_row = tk.Frame(editor_frame)
        name_row.pack(fill=tk.X, pady=2)
        tk.Label(name_row, text="Display Name:", font=("Arial", 9, "bold"),
                 width=18, anchor=tk.W).pack(side=tk.LEFT)
        self.name_var = tk.StringVar()
        tk.Entry(name_row, textvariable=self.name_var,
                 font=("Arial", 9), width=40).pack(side=tk.LEFT, padx=5)

        # Description
        desc_row = tk.Frame(editor_frame)
        desc_row.pack(fill=tk.X, pady=2)
        tk.Label(desc_row, text="Description:", font=("Arial", 9, "bold"),
                 width=18, anchor=tk.W).pack(side=tk.LEFT)
        self.desc_var = tk.StringVar()
        tk.Entry(desc_row, textvariable=self.desc_var,
                 font=("Arial", 9), width=40).pack(side=tk.LEFT, padx=5)

        # Narration prompt text area
        tk.Label(editor_frame, text="Narration Prompt (use {title}, {transcript}, {target_word_count}, {niche_angle}, {word_count_tolerance}):",
                 font=("Arial", 9, "bold"), anchor=tk.W).pack(fill=tk.X, pady=(10, 2))

        nar_frame = tk.Frame(editor_frame)
        nar_frame.pack(fill=tk.BOTH, expand=True, pady=2)
        self.narration_text = tk.Text(nar_frame, height=10, width=55,
                                       font=("Consolas", 9), wrap=tk.WORD)
        nar_scroll = tk.Scrollbar(nar_frame, orient=tk.VERTICAL, command=self.narration_text.yview)
        self.narration_text.config(yscrollcommand=nar_scroll.set)
        self.narration_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        nar_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Metadata prompt text area
        tk.Label(editor_frame, text="Metadata Prompt (use {transcript}, {script}, {niche_angle}):",
                 font=("Arial", 9, "bold"), anchor=tk.W).pack(fill=tk.X, pady=(5, 2))

        meta_frame = tk.Frame(editor_frame)
        meta_frame.pack(fill=tk.BOTH, expand=True, pady=2)
        self.metadata_text = tk.Text(meta_frame, height=5, width=55,
                                      font=("Consolas", 9), wrap=tk.WORD)
        meta_scroll = tk.Scrollbar(meta_frame, orient=tk.VERTICAL, command=self.metadata_text.yview)
        self.metadata_text.config(yscrollcommand=meta_scroll.set)
        self.metadata_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        meta_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Save button
        btn_row = tk.Frame(editor_frame)
        btn_row.pack(fill=tk.X, pady=5)
        tk.Button(btn_row, text="💾 Save Changes", font=("Arial", 10, "bold"),
                  bg='#1565C0', fg='white', width=20,
                  command=self._save_current).pack(side=tk.LEFT, padx=5)
        tk.Label(btn_row, text="  Available placeholders: {title} {transcript} {target_word_count} {niche_angle} {word_count_tolerance}",
                 font=("Arial", 8), fg="#666").pack(side=tk.LEFT)

        # Populate the listbox
        self._refresh_listbox()

    def _load(self):
        """Load prompts from config file"""
        import json, os
        if os.path.isfile(self.config_path):
            try:
                with open(self.config_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {"active_prompt": "", "prompts": {}}

    def _save(self):
        """Save prompts to config file"""
        import json
        try:
            with open(self.config_path, "w", encoding="utf-8") as f:
                json.dump(self.data, f, indent=2, ensure_ascii=False)
            self.dashboard._refresh_script_styles()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Save Error", f"Failed to save prompts:\n{e}")

    def _save_json_only(self):
        """Write the JSON config without re-deriving dropdowns (used when only
        the name/description of a master-backed preset changed)."""
        import json
        try:
            with open(self.config_path, "w", encoding="utf-8") as f:
                json.dump(self.data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Save Error", f"Failed to save prompts:\n{e}")

    def _refresh_listbox(self):
        """Refresh the listbox with current prompt names.

        If this manager was opened for a specific mode, only that mode's presets
        are shown (in mode order). The preset that was selected in the dropdown
        is auto-selected so its text loads straight into the editor.
        """
        self.listbox.delete(0, tk.END)
        prompts = self.data.get("prompts", {})
        active = self.data.get("active_prompt", "")

        # Decide which slugs to show, and in what order.
        if self.category:
            # Category mode: every prompt whose slug carries the prefix.
            pref = self.category
            ordered = [s for s in prompts
                       if s == pref or s.startswith(pref + "_")]
            ordered.sort(key=lambda s: (s != pref, prompts[s].get("name", s).lower()))
        elif self.mode_slugs:
            ordered = [s for s in self.mode_slugs if s in prompts]
            # Include any not-yet-saved extras added this session (e.g. duplicates)
            extras = [s for s in self._session_added if s in prompts and s not in ordered]
            ordered += extras
        else:
            ordered = list(prompts.keys())
        self.visible_slugs = ordered

        for slug in ordered:
            info = prompts.get(slug, {})
            name = info.get("name", slug)
            marker = " ★" if slug == active else ""
            self.listbox.insert(tk.END, f"  {name}{marker}")

        # Pick what to select: the dropdown's preset first, else active, else first.
        target = None
        if self.preselect_slug and self.preselect_slug in ordered:
            target = self.preselect_slug
            self.preselect_slug = None  # only force-select once, on open
        elif active in ordered:
            target = active
        elif ordered:
            target = ordered[0]

        if target is not None:
            idx = ordered.index(target)
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(idx)
            self.listbox.see(idx)
            self.listbox.event_generate("<<ListboxSelect>>")

    def _handle_close(self):
        """Fire the on_close callback (e.g. refresh a dropdown) then close."""
        try:
            if self.on_close:
                self.on_close()
        finally:
            self.window.destroy()

    def _on_select(self, event):
        """Load selected prompt into the editor.

        In Write Story mode (prefer_master), slugs backed by a master .txt file
        load their narration text FROM that file — because that's what Gemini
        actually reads. Editing then saves back to the file.
        """
        selection = self.listbox.curselection()
        if not selection:
            return
        idx = selection[0]
        prompts = self.data.get("prompts", {})
        if idx >= len(self.visible_slugs):
            return
        slug = self.visible_slugs[idx]
        info = prompts.get(slug, {})

        self.current_slug = slug
        self.editing_master = False
        master_text = None
        if self.prefer_master:
            try:
                from core.script_generator import ScriptGenerator
                if ScriptGenerator.is_master_backed(slug):
                    p = ScriptGenerator.get_master_prompt_path(slug)
                    master_text = p.read_text(encoding="utf-8")
                    self.editing_master = True
            except Exception:
                master_text = None

        self.slug_var.set(slug)
        self.name_var.set(info.get("name", ""))
        self.desc_var.set(info.get("description", ""))
        self.narration_text.delete("1.0", tk.END)
        self.narration_text.insert(
            "1.0", master_text if master_text is not None
            else info.get("narration_prompt", ""))
        self.metadata_text.delete("1.0", tk.END)
        self.metadata_text.insert("1.0", info.get("metadata_prompt", ""))
        self._update_source_banner()

    def _update_source_banner(self):
        """Show which store the current preset saves to (master .txt vs JSON)."""
        if not hasattr(self, "source_banner"):
            return
        if self.editing_master:
            try:
                from core.script_generator import ScriptGenerator
                fname = ScriptGenerator.get_master_prompt_path(self.current_slug).name
            except Exception:
                fname = "master prompt file"
            self.source_banner.config(
                text=("🎬 Master prompt (Write Story) — Gemini reads this when "
                      "watching the video. Saving writes to:\n" + fname),
                fg="#1565C0")
        elif self.prefer_master:
            self.source_banner.config(
                text=("📝 This preset has no master file — uses the JSON prompt "
                      "below. Saving writes to script_prompts.json."),
                fg="#555")
        else:
            self.source_banner.config(text="", fg="#555")

    def _save_current(self):
        """Save the currently edited prompt template"""
        prompts = self.data.get("prompts", {})
        slug = self.slug_var.get().strip()
        if not slug:
            from tkinter import messagebox
            messagebox.showwarning("Missing Slug", "Enter a slug (internal key) for this template.")
            return
        name = self.name_var.get().strip() or slug
        description = self.desc_var.get().strip()
        narration = self.narration_text.get("1.0", tk.END).strip()
        metadata = self.metadata_text.get("1.0", tk.END).strip()

        # Category mode: keep the slug under the category prefix so the new
        # niche shows up in the Case Commentary dropdown. Don't rename the base.
        if self.category and slug != self.category \
                and not slug.startswith(self.category + "_"):
            slug = f"{self.category}_{slug}"

        if not narration:
            from tkinter import messagebox
            messagebox.showwarning("Missing Prompt", "The narration prompt cannot be empty.")
            return

        # ── Master-prompt path: write back to the .txt Gemini actually reads ──
        if self.editing_master and slug == self.current_slug:
            from tkinter import messagebox
            try:
                from core.script_generator import ScriptGenerator
                p = ScriptGenerator.get_master_prompt_path(slug)
                p.write_text(narration, encoding="utf-8")
            except Exception as e:
                messagebox.showerror("Save Error",
                                     f"Failed to save master prompt file:\n{e}")
                return
            # Keep the JSON name/description in sync (used for the dropdown label).
            if slug in prompts:
                prompts[slug]["name"] = name
                prompts[slug]["description"] = description
                self.data["prompts"] = prompts
                self._save_json_only()
            self._refresh_listbox()
            messagebox.showinfo(
                "Saved", f"Master prompt '{name}' saved to:\n{p.name}\n\n"
                "Write Story mode will now use your edited version.")
            return

        if slug in prompts:
            prompts[slug].update({
                "name": name,
                "description": description,
                "narration_prompt": narration,
                "metadata_prompt": metadata,
            })
        else:
            prompts[slug] = {
                "name": name,
                "description": description,
                "narration_prompt": narration,
                "metadata_prompt": metadata,
            }
            self.data["active_prompt"] = slug
            if slug not in self._session_added:
                self._session_added.append(slug)

        self.data["prompts"] = prompts
        self._save()
        self._refresh_listbox()
        from tkinter import messagebox
        messagebox.showinfo("Saved", f"Template '{name}' saved successfully!")

    def _add_prompt(self):
        """Clear the editor to add a new prompt template.

        In Case Commentary category mode, seed the new niche from the base
        courtroom prompt so it keeps the required 3-section output format the
        parser depends on — the user just rewrites the ROLE/focus for the niche.
        """
        self.current_slug = None
        self.editing_master = False

        if self.category == "case_commentary":
            prompts = self.data.get("prompts", {})
            base = prompts.get("case_commentary", {})
            # Suggest a unique slug like case_commentary_movies
            i = 1
            new_slug = f"case_commentary_niche{i}"
            while new_slug in prompts:
                i += 1
                new_slug = f"case_commentary_niche{i}"
            self.slug_var.set(new_slug)
            self.name_var.set("New Niche (rename me)")
            self.desc_var.set("Describe what this niche does.")
            self.narration_text.delete("1.0", tk.END)
            self.narration_text.insert("1.0", base.get("narration_prompt", ""))
            self.metadata_text.delete("1.0", tk.END)
            self.metadata_text.insert("1.0", base.get("metadata_prompt", ""))
            self._update_source_banner()
            return

        self.slug_var.set("")
        self.name_var.set("")
        self.desc_var.set("")
        self.narration_text.delete("1.0", tk.END)
        self.narration_text.insert("1.0",
            "You are a professional script writer.\n\n"
            "Original Video Title: {title}\n\n"
            "RAW TRANSCRIPT:\n{transcript}\n\n"
            "Write the script now (no preamble, no explanations — just the script):"
        )
        self.metadata_text.delete("1.0", tk.END)
        self.metadata_text.insert("1.0",
            'Based on the following, generate a title and 2 hashtags.\n\n'
            'Transcript:\n{transcript}\n\n'
            'Cinematic Script:\n{script}\n\n'
            'Return JSON: {{"suggested_title": "...", "hashtag_1": "...", "hashtag_2": "..."}}'
        )
        self._update_source_banner()

    def _delete_prompt(self):
        """Delete the selected prompt template"""
        selection = self.listbox.curselection()
        if not selection:
            return
        idx = selection[0]
        prompts = self.data.get("prompts", {})
        if idx >= len(self.visible_slugs):
            return
        slug = self.visible_slugs[idx]

        if len(prompts) <= 1:
            from tkinter import messagebox
            messagebox.showwarning("Cannot Delete", "Cannot delete the last template. Add a new one first.")
            return

        from tkinter import messagebox
        if not messagebox.askyesno("Confirm Delete", f"Delete template '{prompts[slug].get('name', slug)}'?"):
            return

        was_active = (slug == self.data.get("active_prompt", ""))
        del prompts[slug]
        if was_active and prompts:
            self.data["active_prompt"] = next(iter(prompts.keys()))
        self.data["prompts"] = prompts
        self._save()
        self._refresh_listbox()
        self.slug_var.set("")
        self.name_var.set("")
        self.desc_var.set("")
        self.narration_text.delete("1.0", tk.END)
        self.metadata_text.delete("1.0", tk.END)

    def _duplicate_prompt(self):
        """Duplicate the selected prompt template"""
        selection = self.listbox.curselection()
        if not selection:
            return
        idx = selection[0]
        prompts = self.data.get("prompts", {})
        if idx >= len(self.visible_slugs):
            return
        slug = self.visible_slugs[idx]
        info = prompts[slug]

        import re
        base_slug = slug.rstrip("0123456789_")
        copy_num = 1
        while f"{base_slug}_copy{copy_num}" in prompts:
            copy_num += 1
        new_slug = f"{base_slug}_copy{copy_num}"

        prompts[new_slug] = {
            "name": f"{info.get('name', slug)} (Copy)",
            "description": info.get("description", ""),
            "narration_prompt": info.get("narration_prompt", ""),
            "metadata_prompt": info.get("metadata_prompt", ""),
        }
        self.data["prompts"] = prompts
        self._session_added.append(new_slug)
        self._save()
        self._refresh_listbox()
